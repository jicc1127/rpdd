# -*- coding: utf-8 -*-
"""
module mh_rpdu
MHFarm　繁殖成績をまとめるtools
ｖ1.0　2021/11/3
変更履歴
#v1.01 2021/12/8############
#fpyDF_RPDCow( wbN, sheetorg, sheetN, bDate ) 
dysfrmclving をtimedelta->intに変換した
#fpydysoffstAI( wbN, sheetN )
daysoffstAI をtimedelta->intに変換した
#v1.02 2021/12/9#############
#fpyNewSheetHeifer(wbN, sheetN)
"分娩日" -> "日齢"に変更
#fpyAgeinDays( wbN, sheetN )
日齢計算の関数を追加　(tools)
#fpyDF_RPDHeifer( wbN, sheetorg, sheetN, bDate )
日齢をorgシートからコピーするように追加
上記の関数よりこちらが上位
月齢に関しては、orgシート1か月=30.5日のため、
月齢関数fpyAgeinMnths(wbN, sheetN)を優先した
"""
import openpyxl
import datetime
import fmstls
#fpyNewSheetCow############################################################### 
"""
fpyNewSheetCow : 繁殖成績のデータを入力する列名だけのsheetを作成する。
ｖ1.0
2021/10/2

@author: jicc

"""
#! python3
def fpyNewSheetCow(wbN, sheetN):
    """
    繁殖成績のデータを入力する列名だけのsheetを作成する。
    Parameters
    ----------
   wbN : 　str          ワークブック名　"MH_RPDu.xlsx"
        sheetを作成するワークブック
    sheetN : str　　　　　　シート名:"yyyymmddCow00"
    												
        作成するシート名

    Returns
    -------
    None.

    """
   #import openpyxl
    
    wb = openpyxl.load_workbook(wbN)
    #sheetN = wb[sheetN]
    wb.create_sheet(title=sheetN, index=0)
    sheet = wb[sheetN]
    
    
    sheet.cell(row=1, column=1).value = "LineNo"
    sheet.cell(row=1, column=2).value = "Group"
    sheet.cell(row=1, column=3).value = "Stage"
    sheet.cell(row=1, column=4).value = "牛ｺｰﾄﾞ"
    sheet.cell(row=1, column=5).value = "牛群"
    sheet.cell(row=1, column=6).value = "個体識別番号"
    sheet.cell(row=1, column=7).value = "生年月日"
    sheet.cell(row=1, column=8).value = "産次"
    sheet.cell(row=1, column=9).value = "分娩日"
    sheet.cell(row=1, column=10).value = "分娩後日数"
    sheet.cell(row=1, column=11).value = "授精回数"
    sheet.cell(row=1, column=12).value = "初回授精日"
    sheet.cell(row=1, column=13).value = "初回授精日数"
    sheet.cell(row=1, column=14).value = "最終授精日"
    sheet.cell(row=1, column=15).value = "授精後日数"
    sheet.cell(row=1, column=16).value = "分娩予定日"
    sheet.cell(row=1, column=17).value = "空胎日数"
    sheet.cell(row=1, column=18).value = "基準日"
 
    wb.save(wbN)

#fpyNewSheetCow_s###############################################################
"""
fpyNewSheetCow : 繁殖成績のデータを入力する列名だけのsheetを作成する。
sheet version
ｖ1.0
2023/7/16
@author: jicc

"""
#! python3
def fpyNewSheetCow_s(sheet):
    """
    繁殖成績のデータを入力する列名だけのsheetを作成する。
    Parameters
    ----------
  
    sheet : worksheet.worksheet.Worksheet
         worksheet object　　　　　　
         シート名:"yyyymmddHeifer00"

    Returns
    -------
    sheet : worksheet.worksheet.Worksheet

    """
    
    sheet.cell(row=1, column=1).value = "LineNo"
    sheet.cell(row=1, column=2).value = "Group"
    sheet.cell(row=1, column=3).value = "Stage"
    sheet.cell(row=1, column=4).value = "牛ｺｰﾄﾞ"
    sheet.cell(row=1, column=5).value = "牛群"
    sheet.cell(row=1, column=6).value = "個体識別番号"
    sheet.cell(row=1, column=7).value = "生年月日"
    sheet.cell(row=1, column=8).value = "産次"
    sheet.cell(row=1, column=9).value = "分娩日"
    sheet.cell(row=1, column=10).value = "分娩後日数"
    sheet.cell(row=1, column=11).value = "授精回数"
    sheet.cell(row=1, column=12).value = "初回授精日"
    sheet.cell(row=1, column=13).value = "初回授精日数"
    sheet.cell(row=1, column=14).value = "最終授精日"
    sheet.cell(row=1, column=15).value = "授精後日数"
    sheet.cell(row=1, column=16).value = "分娩予定日"
    sheet.cell(row=1, column=17).value = "空胎日数"
    sheet.cell(row=1, column=18).value = "基準日"
 
    return sheet
    
#fpyNewSheetHeifer############################################################    
"""
fpyNewSheetHeifer : 繁殖成績のデータを入力する列名だけのsheetを作成する。
ｖ1.0
2021/10/2

@author: jicc

"""
#! python3
def fpyNewSheetHeifer(wbN, sheetN):
    """
    
    Parameters
    ----------
   wbN : 　str          ワークブック名　"MH_RPDu.xlsx"
        sheetを作成するワークブック
    sheetN : str　　　　　　シート名:"yyyymmddHeifer00"
    													
        作成するシート名

    Returns
    -------
    None.

    """
   #import openpyxl
    
    wb = openpyxl.load_workbook(wbN)
    #sheetN = wb[sheetN]
    wb.create_sheet(title=sheetN, index=0)
    sheet = wb[sheetN]
    
    
    sheet.cell(row=1, column=1).value = "LineNo"
    sheet.cell(row=1, column=2).value = "Group"
    sheet.cell(row=1, column=3).value = "Stage"
    sheet.cell(row=1, column=4).value = "牛ｺｰﾄﾞ"
    sheet.cell(row=1, column=5).value = "牛群"
    sheet.cell(row=1, column=6).value = "個体識別番号"
    sheet.cell(row=1, column=7).value = "生年月日"
    sheet.cell(row=1, column=8).value = "産次"
    sheet.cell(row=1, column=9).value = "日齢"
    sheet.cell(row=1, column=10).value = "月齢"
    sheet.cell(row=1, column=11).value = "授精回数"
    sheet.cell(row=1, column=12).value = "初回授精日"
    sheet.cell(row=1, column=13).value = "初回授精月齢"
    sheet.cell(row=1, column=14).value = "最終授精日"
    sheet.cell(row=1, column=15).value = "授精後日数"
    sheet.cell(row=1, column=16).value = "分娩予定日"
    sheet.cell(row=1, column=17).value = "空胎月数"
    sheet.cell(row=1, column=18).value = "基準日"
 
    wb.save(wbN)

#fpyNewSheetHeifer_s#####################################################
"""
fpyNewSheetHeifer_s : 繁殖成績のデータを入力する列名だけのsheetを作成する。
sheet version
ｖ1.0
2023/6/29

@author: jicc

"""
#! python3
def fpyNewSheetHeifer_s(sheet):
    """
    
    Parameters
    ----------
  
    sheet : worksheet.worksheet.Worksheet
         worksheet object　　　　　　
         シート名:"yyyymmddHeifer00"

    Returns
    -------
    sheet : worksheet.worksheet.Worksheet

    """
   
    sheet.cell(row=1, column=1).value = "LineNo"
    sheet.cell(row=1, column=2).value = "Group"
    sheet.cell(row=1, column=3).value = "Stage"
    sheet.cell(row=1, column=4).value = "牛ｺｰﾄﾞ"
    sheet.cell(row=1, column=5).value = "牛群"
    sheet.cell(row=1, column=6).value = "個体識別番号"
    sheet.cell(row=1, column=7).value = "生年月日"
    sheet.cell(row=1, column=8).value = "産次"
    sheet.cell(row=1, column=9).value = "日齢"
    sheet.cell(row=1, column=10).value = "月齢"
    sheet.cell(row=1, column=11).value = "授精回数"
    sheet.cell(row=1, column=12).value = "初回授精日"
    sheet.cell(row=1, column=13).value = "初回授精月齢"
    sheet.cell(row=1, column=14).value = "最終授精日"
    sheet.cell(row=1, column=15).value = "授精後日数"
    sheet.cell(row=1, column=16).value = "分娩予定日"
    sheet.cell(row=1, column=17).value = "空胎月数"
    sheet.cell(row=1, column=18).value = "基準日"
 
    return sheet

#fpyDF_RPDCow################################################################    
"""
fpyDF_RPDCow:繁殖データ用DF作成の関数
    Umotionの　経産牛一覧plusyyyymmdd　から作成する
    v1.0
    2021/9/12
    @author:jicc

"""
#! python3
def fpyDF_RPDCow( wbN, sheetorg, sheetN, bDate ):
    """
    Parameters
    ----------
    wbN : ワークブック名
        "MH_RPDu.xlsx"
    sheetorg : データ参照シート
        "yyyymmddCoworg"
    sheetN : 経産牛データシート
        "yyyymmddCow00"
    bDate : 基準日
        "yyyy/mm/dd"
    
    Returns
    -------
    None.

    """
    
    #import openpyxl
    #import datetime

    wb = openpyxl.load_workbook(wbN)
    sheetorg = wb[sheetorg]

    sheetN = wb[sheetN]
    max_row = sheetorg.max_row
    bDate = datetime.datetime.strptime(bDate, '%Y/%m/%d')
    
    for row_num in range(2, max_row + 1):     #先頭行をスキップ
    
      #LineNo      
      sheetN.cell(row=row_num, column=1).value = row_num - 1
      
      #Stage
      sheetN.cell(row=row_num, column=3).value = sheetorg.cell(row=row_num, 
                                                            column=3).value
                                                  #状態
      #牛ｺｰﾄﾞ
      sheetN.cell(row=row_num, column=4).value = sheetorg.cell(row=row_num, 
                                                            column=4).value
                                                  #牛番号
      #牛群      
      sheetN.cell(row=row_num, column=5).value = sheetorg.cell(row=row_num, 
                                                            column=1).value
                                                  #牛群
      #個体識別番号
      sheetN.cell(row=row_num, column=6).value = sheetorg.cell(row=row_num, 
                                                            column=5).value
                                                  #個体識別番号
      #生年月日
      sheetN.cell(row=row_num, column=7).value = sheetorg.cell(row=row_num, 
                                                            column=18).value
                                                  #出生日
      #産次
      sheetN.cell(row=row_num, column=8).value = sheetorg.cell(row=row_num, 
                                                            column=6).value
                                                  #産次
      #分娩日
      clvingDate = sheetorg.cell(row=row_num, column=15).value #分娩日
      
      sheetN.cell(row=row_num, column=9).value = clvingDate
                                                  
      #分娩後日数
      dysfrmclving = bDate - clvingDate			#分娩後日数を計算
      #timedelta (datetime - datetime)
      dysfrmclving = dysfrmclving.days		#add 2021/12/8  v1.01
      #int
      sheetN.cell(row=row_num, column=10).value = dysfrmclving
                                                  #搾乳日数 sheetorg.cell(row=row_num, column=7).value
                                                  #の使用を中止　
                                                  #基準日=分娩日のとき　搾乳日数=Noneとなる
                                                  #2021/11/3
      #授精回数
      sheetN.cell(row=row_num, column=11).value = sheetorg.cell(row=row_num, 
                                                            column=10).value
                                                  #授精回数
      #初回授精日
      sheetN.cell(row=row_num, column=12).value = sheetorg.cell(row=row_num, 
                                                            column=16).value
                                                  #初回授精日
      #最終授精日
      sheetN.cell(row=row_num, column=14).value = sheetorg.cell(row=row_num, 
                                                            column=17).value
                                                  #授精日
      #授精後日数
      sheetN.cell(row=row_num, column=15).value = sheetorg.cell(row=row_num, 
                                                            column=11).value
                                                  #授精後日数
      #分娩予定日
      sheetN.cell(row=row_num, column=16).value = sheetorg.cell(row=row_num, 
                                                            column=14).value
                                                  #分娩予定日
       #空胎日数
      sheetN.cell(row=row_num, column=17).value = sheetorg.cell(row=row_num, 
                                                            column=9).value
                                                  #空胎日数
        #鑑定待ちも　空胎日数==分娩後日数となっている
        
      #基準日
      sheetN.cell(row=row_num, column=18).value = bDate
      
      #memo
      #sheetN.cell(row=row_num, column=19).value = sheetorg.cell(row=row_num, 
      #                                                      column=2).value      
                                                  #個体ラベル
                                                  
    
    wb.save(wbN)

#fpyDF_RPDCow_s################################################################    
"""
fpyDF_RPDCow:繁殖データ用DF作成の関数
    Umotionの　経産牛一覧plusyyyymmdd　から作成する
    v1.0
    2021/9/12
    @author:jicc

"""
#! python3
def fpyDF_RPDCow_s( sorg, sheet, bDate ):
    """
    Parameters
    ----------
    sorg : worksheet.worksheet.Worksheet
        "yyyymmddHeiferorg"
    sheet : worksheet.worksheet.Worksheet
         worksheet object　　　　　　
         シート名:"yyyymmddHeifer00"
    bDate : 基準日
        "yyyy/mm/dd"
    
    Returns
    -------
    sheet : worksheet.worksheet.Worksheet

    """

    bDate = datetime.datetime.strptime(bDate, '%Y/%m/%d')
    
    for row_num in range(2, sorg.max_row + 1):     #先頭行をスキップ
    
      #LineNo      
      sheet.cell(row=row_num, column=1).value = row_num - 1
      
      #Stage
      sheet.cell(row=row_num, column=3).value = sorg.cell(row=row_num, 
                                                            column=3).value
                                                  #状態
      #牛ｺｰﾄﾞ
      sheet.cell(row=row_num, column=4).value = sorg.cell(row=row_num, 
                                                            column=4).value
                                                  #牛番号
      #牛群      
      sheet.cell(row=row_num, column=5).value = sorg.cell(row=row_num, 
                                                            column=1).value
                                                  #牛群
      #個体識別番号
      sheet.cell(row=row_num, column=6).value = sorg.cell(row=row_num, 
                                                            column=5).value
                                                  #個体識別番号
      #生年月日
      sheet.cell(row=row_num, column=7).value = sorg.cell(row=row_num, 
                                                            column=18).value
                                                  #出生日
      #産次
      sheet.cell(row=row_num, column=8).value = sorg.cell(row=row_num, 
                                                            column=6).value
                                                  #産次
      #分娩日
      clvingDate = sorg.cell(row=row_num, column=15).value #分娩日
      
      sheet.cell(row=row_num, column=9).value = clvingDate
                                                  
      #分娩後日数
      dysfrmclving = bDate - clvingDate			#分娩後日数を計算
      #timedelta (datetime - datetime)
      dysfrmclving = dysfrmclving.days		#add 2021/12/8  v1.01
      #int
      sheet.cell(row=row_num, column=10).value = dysfrmclving
                                                  #搾乳日数 sorg.cell(row=row_num, column=7).value
                                                  #の使用を中止　
                                                  #基準日=分娩日のとき　搾乳日数=Noneとなる
                                                  #2021/11/3
      #授精回数
      sheet.cell(row=row_num, column=11).value = sorg.cell(row=row_num, 
                                                            column=10).value
                                                  #授精回数
      #初回授精日
      sheet.cell(row=row_num, column=12).value = sorg.cell(row=row_num, 
                                                            column=16).value
                                                  #初回授精日
      #最終授精日
      sheet.cell(row=row_num, column=14).value = sorg.cell(row=row_num, 
                                                            column=17).value
                                                  #授精日
      #授精後日数
      sheet.cell(row=row_num, column=15).value = sorg.cell(row=row_num, 
                                                            column=11).value
                                                  #授精後日数
      #分娩予定日
      sheet.cell(row=row_num, column=16).value = sorg.cell(row=row_num, 
                                                            column=14).value
                                                  #分娩予定日
       #空胎日数
      sheet.cell(row=row_num, column=17).value = sorg.cell(row=row_num, 
                                                            column=9).value
                                                  #空胎日数
        #鑑定待ちも　空胎日数==分娩後日数となっている
        
      #基準日
      sheet.cell(row=row_num, column=18).value = bDate
      
      #memo
      #sheet.cell(row=row_num, column=19).value = sorg.cell(row=row_num, 
      #                                                      column=2).value      
                                                  #個体ラベル
                                                  
    
    return sheet

#fpyDF_RPDHeiferv1_0########################################################
"""
fpyDF_RPDHeiferv1_0:繁殖データ用DF作成の関数
    Umotionの　未経産牛一覧yyyymmdd　から作成する
    2022/12/5までの未経産牛一覧yyyymmdd　で使用する。
    v1.0
    2021/10/3
    @author:jicc

"""

#! python3
def fpyDF_RPDHeiferv1_0( wbN, sheetorg, sheetN, bDate ):
    """
    Parameters
    ----------
    wbN : ワークブック名
        "MH_RPDu.xlsx"
    sheetorg : データ参照シート
        "yyyymmddHeiferorg"
    sheetN : 経産牛データシート
        "yyyymmddHeifer00"
    bDate : 基準日
        "yyyy/mm/dd"
    
    Returns
    -------
    None.

    """
    
    #import openpyxl
    #import datetime

    wb = openpyxl.load_workbook(wbN)
    sheetorg = wb[sheetorg]
    sheetN = wb[sheetN]
    max_row = sheetorg.max_row
    bDate = datetime.datetime.strptime(bDate, '%Y/%m/%d')
    
    for row_num in range(2, max_row + 1):     #先頭行をスキップ
    
      #LineNo      
      sheetN.cell(row=row_num, column=1).value = row_num - 1
      
      #Stage
      sheetN.cell(row=row_num, column=3).value = sheetorg.cell(row=row_num, 
                                                            column=2).value
                                                  #状態
      #牛ｺｰﾄﾞ
      sheetN.cell(row=row_num, column=4).value = sheetorg.cell(row=row_num, 
                                                            column=3).value
                                                  #牛番号
      #牛群      
      sheetN.cell(row=row_num, column=5).value = sheetorg.cell(row=row_num, 
                                                            column=1).value
                                                  #牛群
      #個体識別番号
      sheetN.cell(row=row_num, column=6).value = sheetorg.cell(row=row_num, 
                                                            column=4).value
                                                  #個体識別番号
      #生年月日
      sheetN.cell(row=row_num, column=7).value = sheetorg.cell(row=row_num, 
                                                            column=5).value
                                                  #出生日
      #産次
      sheetN.cell(row=row_num, column=8).value = 0
                                                  #産次
      #日齢
      sheetN.cell(row=row_num, column=9).value = sheetorg.cell(row=row_num, 
                                                            column=6).value
                                                  #日齢
     #分娩後日数
     # sheetN.cell(row=row_num, column=10).value = sheetorg.cell(row=row_num, 
      #                                                      column=7).value
      #生後月齢として後でツールを追加する 
                                           #搾乳日数
      #授精回数
      sheetN.cell(row=row_num, column=11).value = sheetorg.cell(row=row_num, 
                                                            column=12).value
                                                  #授精回数
      #初回授精日
      sheetN.cell(row=row_num, column=12).value = sheetorg.cell(row=row_num, 
                                                            column=16).value
                                                  #初回授精日
      #最終授精日
      sheetN.cell(row=row_num, column=14).value = sheetorg.cell(row=row_num, 
                                                            column=17).value
                                                  #授精日
      #授精後日数
      sheetN.cell(row=row_num, column=15).value = sheetorg.cell(row=row_num, 
                                                            column=13).value
                                                  #授精後日数
      #分娩予定日
      sheetN.cell(row=row_num, column=16).value = sheetorg.cell(row=row_num, 
                                                            column=15).value
                                                  #分娩予定日
       #空胎日数
      #sheetN.cell(row=row_num, column=17).value = sheetorg.cell(row=row_num, 
      #                                                      column=9).value
                                                  #空胎日数
        #鑑定待ちも　空胎日数==分娩後日数となっている
        
      #基準日
      sheetN.cell(row=row_num, column=18).value = bDate
      
      #memo
      #sheetN.cell(row=row_num, column=19).value = sheetorg.cell(row=row_num, 
      #                                                      column=2).value      
                                                  #個体ラベル
                                                  
    
    wb.save(wbN)

#fpyDF_RPDHeifer#########################################################
"""
fpyDF_RPDHeifer:繁殖データ用DF作成の関数
    Umotionの　未経産牛一覧yyyymmdd　から作成する
    未経産牛一覧20221211から使用する
    v2.0
    2023/2/5
    @author:jicc

"""

#! python3
def fpyDF_RPDHeifer( wbN, sheetorg, sheetN, bDate ):
    """
    Parameters
    ----------
    wbN : ワークブック名
        "MH_RPDu.xlsx"
    sheetorg : データ参照シート
        "yyyymmddHeiferorg"
    sheetN : 経産牛データシート
        "yyyymmddHeifer00"
    bDate : 基準日
        "yyyy/mm/dd"
    
    Returns
    -------
    None.

    """
    
    #import openpyxl
    #import datetime

    wb = openpyxl.load_workbook(wbN)
    sheetorg = wb[sheetorg]
    sheetN = wb[sheetN]
    max_row = sheetorg.max_row
    bDate = datetime.datetime.strptime(bDate, '%Y/%m/%d')
    
    for row_num in range(2, max_row + 1):     #先頭行をスキップ
    
      #LineNo      
      sheetN.cell(row=row_num, column=1).value = row_num - 1
      
      #Stage
      sheetN.cell(row=row_num, column=3).value = sheetorg.cell(row=row_num, 
                                                            column=2).value
                                                  #状態
      #牛ｺｰﾄﾞ
      sheetN.cell(row=row_num, column=4).value = sheetorg.cell(row=row_num, 
                                                            column=3).value
                                                  #牛番号
      #牛群      
      sheetN.cell(row=row_num, column=5).value = sheetorg.cell(row=row_num, 
                                                            column=1).value
                                                  #牛群
      #個体識別番号
      sheetN.cell(row=row_num, column=6).value = sheetorg.cell(row=row_num, 
                                                            column=4).value
                                                  #個体識別番号
      #生年月日
      sheetN.cell(row=row_num, column=7).value = sheetorg.cell(row=row_num, 
                                                            column=5).value
                                                  #出生日
      #産次
      sheetN.cell(row=row_num, column=8).value = 0
                                                  #産次
      #日齢
      sheetN.cell(row=row_num, column=9).value = sheetorg.cell(row=row_num, 
                                                            column=6).value
                                                  #日齢
     #分娩後日数
     # sheetN.cell(row=row_num, column=10).value = sheetorg.cell(row=row_num, 
      #                                                      column=7).value
      #生後月齢として後でツールを追加する 
                                           #搾乳日数
      #授精回数
      sheetN.cell(row=row_num, column=11).value = sheetorg.cell(row=row_num, 
                                                            column=10).value
                                                  #授精回数
      #初回授精日
      sheetN.cell(row=row_num, column=12).value = sheetorg.cell(row=row_num, 
                                                            column=14).value
                                                  #初回授精日
      #最終授精日
      sheetN.cell(row=row_num, column=14).value = sheetorg.cell(row=row_num, 
                                                            column=15).value
                                                  #授精日
      #授精後日数
      sheetN.cell(row=row_num, column=15).value = sheetorg.cell(row=row_num, 
                                                            column=11).value
                                                  #授精後日数
      #分娩予定日
      sheetN.cell(row=row_num, column=16).value = sheetorg.cell(row=row_num, 
                                                            column=13).value
                                                  #分娩予定日
       #空胎日数
      #sheetN.cell(row=row_num, column=17).value = sheetorg.cell(row=row_num, 
      #                                                      column=9).value
                                                  #空胎日数
        #鑑定待ちも　空胎日数==分娩後日数となっている
        
      #基準日
      sheetN.cell(row=row_num, column=18).value = bDate
      
      #memo
      #sheetN.cell(row=row_num, column=19).value = sheetorg.cell(row=row_num, 
      #                                                      column=2).value      
                                                  #個体ラベル
                                                  
    
    wb.save(wbN)

#fpyDF_RPDHeifer_s#######################################################
"""
fpyDF_RPDHeifer_s:繁殖データ用DF作成の関数
    Umotionの　未経産牛一覧yyyymmdd　から作成する
    未経産牛一覧20221211から使用する
    sheet version
    v2.0
    2023/7/2
    @author:jicc

"""

#! python3
def fpyDF_RPDHeifer_s( sorg, sheet, bDate ):
    """
    Parameters
    ----------
    sorg : worksheet.worksheet.Worksheet
        "yyyymmddHeiferorg"
    sheet : worksheet.worksheet.Worksheet
         worksheet object　　　　　　
         シート名:"yyyymmddHeifer00"
    bDate : 基準日
        "yyyy/mm/dd"
    
    Returns
    -------
    sheet : worksheet.worksheet.Worksheet

    """
    
    #max_row = sorg.max_row
    bDate = datetime.datetime.strptime(bDate, '%Y/%m/%d')
    
    for row_num in range(2, sorg.max_row + 1):     #先頭行をスキップ
    
      #LineNo      
      sheet.cell(row=row_num, column=1).value = row_num - 1
      
      #Stage
      sheet.cell(row=row_num, column=3).value = sorg.cell(row=row_num, 
                                                            column=2).value
                                                  #状態
      #牛ｺｰﾄﾞ
      sheet.cell(row=row_num, column=4).value = sorg.cell(row=row_num, 
                                                            column=3).value
                                                  #牛番号
      #牛群      
      sheet.cell(row=row_num, column=5).value = sorg.cell(row=row_num, 
                                                            column=1).value
                                                  #牛群
      #個体識別番号
      sheet.cell(row=row_num, column=6).value = sorg.cell(row=row_num, 
                                                            column=4).value
                                                  #個体識別番号
      #生年月日
      sheet.cell(row=row_num, column=7).value = sorg.cell(row=row_num, 
                                                            column=5).value
                                                  #出生日
      #産次
      sheet.cell(row=row_num, column=8).value = 0
                                                  #産次
      #日齢
      sheet.cell(row=row_num, column=9).value = sorg.cell(row=row_num, 
                                                            column=6).value
                                                  #日齢
     #分娩後日数
     # sheet.cell(row=row_num, column=10).value = sorg.cell(row=row_num, 
      #                                                      column=7).value
      #生後月齢として後でツールを追加する 
                                           #搾乳日数
      #授精回数
      sheet.cell(row=row_num, column=11).value = sorg.cell(row=row_num, 
                                                            column=10).value
                                                  #授精回数
      #初回授精日
      sheet.cell(row=row_num, column=12).value = sorg.cell(row=row_num, 
                                                            column=14).value
                                                  #初回授精日
      #最終授精日
      sheet.cell(row=row_num, column=14).value = sorg.cell(row=row_num, 
                                                            column=15).value
                                                  #授精日
      #授精後日数
      sheet.cell(row=row_num, column=15).value = sorg.cell(row=row_num, 
                                                            column=11).value
                                                  #授精後日数
      #分娩予定日
      sheet.cell(row=row_num, column=16).value = sorg.cell(row=row_num, 
                                                            column=13).value
                                                  #分娩予定日
       #空胎日数
      #sheetN.cell(row=row_num, column=17).value = sheetorg.cell(row=row_num, 
      #                                                      column=9).value
                                                  #空胎日数
        #鑑定待ちも　空胎日数==分娩後日数となっている
        
      #基準日
      sheet.cell(row=row_num, column=18).value = bDate
      
      #memo
      #sheetN.cell(row=row_num, column=19).value = sheetorg.cell(row=row_num, 
      #                                                      column=2).value      
                                                  #個体ラベル
                                                  
    
    return sheet
   
"""
fpyidNo_9to10 : ９～10桁耳標の数値を文字列として、
    9桁の耳標に1桁目に０を加え10桁とする
ｖ1.0
2021/4/29
@author: jicc
"""

#! python3
def fpyidNo_9to10( wbN, sheetN, col ):
    """
    Parameters
    ----------
    wbN : 　str  
        対象となるExcelファイル名　拡張子.xlsxをつけ、　''でくくる。
    sheetN : str
        対象となるsheet
    col : int
        変更する10桁耳標の列

    Returns
    -------
    None.

    """
       
    #import openpyxl
    #import datetime
    
    wb = openpyxl.load_workbook(wbN)
    sheetN = wb[sheetN]   #wb.get_sheet_by_name(sheetN)
    max_row = sheetN.max_row
        
    for row_num in range(2, max_row + 1):     #先頭行をスキップ
        
        idNo = sheetN.cell(row=row_num, column=col).value
        idNo = str(idNo)
        if len(idNo) == 9:
            sheetN.cell(row=row_num, column=col).value = '0' + idNo 
        else:
            sheetN.cell(row=row_num, column=col).value = idNo 
              
    
    wb.save(wbN)

#fpygrouoping###############################################################
"""
fpygrouoping : Group のcodeに従って、グループ分けし、Group列,Stage列を書き換える。
                Umotion 経産牛一覧plusより
ｖ1.0
2021/9/4
@author: jicc
"""

#! python3
def fpygrouoping( wbN, sheetN, VWP ):
    """
    Parameters
    ----------
    wbN : 　str  
        対象となるExcelファイル名　拡張子.xlsxをつけ、　''でくくる。
    sheetN : str
        対象となるsheet
    VWP : int
        VWP
    Returns
    -------
    None.

    """
       
    #import openpyxl
    #import datetime
    
    wb = openpyxl.load_workbook(wbN)
    sheetN = wb[sheetN]			#.get_sheet_by_name(sheetN)
    max_row = sheetN.max_row
        
    for row_num in range(2, max_row + 1):     #先頭行をスキップ
        
        stage = sheetN.cell(row=row_num, column=3).value
        #Stageの値を取得
        if stage == 'フレッシュ':
            AItimes = sheetN.cell(row=row_num, column=11).value
            #授精回数
            dysfrmclving = sheetN.cell(row=row_num, column=10).value
            #分娩後日数
            if AItimes == 0:
                if dysfrmclving <= VWP:
                   sheetN.cell(row=row_num, column=2).value = 1 
                   sheetN.cell(row=row_num, column=3).value = '待機'
                
                elif dysfrmclving > VWP:
                   sheetN.cell(row=row_num, column=2).value = 2 
                   sheetN.cell(row=row_num, column=3).value = 'AI待ち 未授精'
                
                else:
                   sheetN.cell(row=row_num, column=2).value = 'error' 
        
            elif AItimes >= 1:
                sheetN.cell(row=row_num, column=2).value = 3 
                sheetN.cell(row=row_num, column=3).value = 'AI待ち 授精済'
            
            else:
                sheetN.cell(row=row_num, column=2).value = 'error'
                
        elif stage == '未授精':
            #AItimes = sheetN.cell(row=row_num, column=11).value
            #授精回数
            dysfrmclving = sheetN.cell(row=row_num, column=10).value
            #分娩後日数
            if dysfrmclving <= VWP:
                   sheetN.cell(row=row_num, column=2).value = 1 
                   sheetN.cell(row=row_num, column=3).value = '待機'
            elif dysfrmclving > VWP:
                   sheetN.cell(row=row_num, column=2).value = 2 
                   sheetN.cell(row=row_num, column=3).value = 'AI待ち 未授精'
       
        elif stage == '未受胎(－)':
            sheetN.cell(row=row_num, column=2).value = 3 
            sheetN.cell(row=row_num, column=3).value = 'AI待ち 授精済'
                
        elif stage == '授精':
            sheetN.cell(row=row_num, column=2).value = 4 
            sheetN.cell(row=row_num, column=3).value = '妊娠鑑定予定'
            
        elif stage == '受胎(＋)':
            sheetN.cell(row=row_num, column=2).value = 5 
            sheetN.cell(row=row_num, column=3).value = '妊娠鑑定+'
        
        elif stage == '乾乳前期':
            sheetN.cell(row=row_num, column=2).value = 6 
            sheetN.cell(row=row_num, column=3).value = '乾乳'
        
        elif stage == '繁殖除外':
            sheetN.cell(row=row_num, column=2).value = 7 
            sheetN.cell(row=row_num, column=3).value = '繁殖対象外'
        else:
            sheetN.cell(row=row_num, column=2).value = 'error'

        
    #max_row = str(max_row)
    #R_max_row = '"R" + max_row' 
    #add_sort_condition('A2:R_max_row', descending=False)
    #add_filter_column(self, 1, vals, blank=False)   
  
    wb.save(wbN)

#fpygrouoping_s###############################################################
"""
fpygrouoping_s : Group のcodeに従って、グループ分けし、Group列,Stage列を書き換える。
                Umotion 経産牛一覧plusより
sheet version
ｖ1.0
2023/7/16
@author: jicc
"""
#! python3
def fpygrouoping_s( sheet, VWP ):
    """
    Parameters
    ----------
   sheet : worksheet.worksheet.Worksheet
        worksheet object　　　　　　
        シート名:"yyyymmddHeifer00"
    VWP : int
        volantary waiting period
        50
        
    Returns
    -------
    sheet : worksheet.worksheet.Worksheet

    """
        
    for row_num in range(2, sheet.max_row + 1):     #先頭行をスキップ
        
        stage = sheet.cell(row=row_num, column=3).value
        #Stageの値を取得
        if stage == 'フレッシュ':
            AItimes = sheet.cell(row=row_num, column=11).value
            #授精回数
            dysfrmclving = sheet.cell(row=row_num, column=10).value
            #分娩後日数
            if AItimes == 0:
                if dysfrmclving <= VWP:
                   sheet.cell(row=row_num, column=2).value = 1 
                   sheet.cell(row=row_num, column=3).value = '待機'
                
                elif dysfrmclving > VWP:
                   sheet.cell(row=row_num, column=2).value = 2 
                   sheet.cell(row=row_num, column=3).value = 'AI待ち 未授精'
                
                else:
                   sheet.cell(row=row_num, column=2).value = 'error' 
        
            elif AItimes >= 1:
                sheet.cell(row=row_num, column=2).value = 3 
                sheet.cell(row=row_num, column=3).value = 'AI待ち 授精済'
            
            else:
                sheet.cell(row=row_num, column=2).value = 'error'
                
        elif stage == '未授精':
            #AItimes = sheet.cell(row=row_num, column=11).value
            #授精回数
            dysfrmclving = sheet.cell(row=row_num, column=10).value
            #分娩後日数
            if dysfrmclving <= VWP:
                   sheet.cell(row=row_num, column=2).value = 1 
                   sheet.cell(row=row_num, column=3).value = '待機'
            elif dysfrmclving > VWP:
                   sheet.cell(row=row_num, column=2).value = 2 
                   sheet.cell(row=row_num, column=3).value = 'AI待ち 未授精'
       
        elif stage == '未受胎(－)':
            sheet.cell(row=row_num, column=2).value = 3 
            sheet.cell(row=row_num, column=3).value = 'AI待ち 授精済'
                
        elif stage == '授精':
            sheet.cell(row=row_num, column=2).value = 4 
            sheet.cell(row=row_num, column=3).value = '妊娠鑑定予定'
            
        elif stage == '受胎(＋)':
            sheet.cell(row=row_num, column=2).value = 5 
            sheet.cell(row=row_num, column=3).value = '妊娠鑑定+'
        
        elif stage == '乾乳前期':
            sheet.cell(row=row_num, column=2).value = 6 
            sheet.cell(row=row_num, column=3).value = '乾乳'
        
        elif stage == '繁殖除外':
            sheet.cell(row=row_num, column=2).value = 7 
            sheet.cell(row=row_num, column=3).value = '繁殖対象外'
        else:
            sheet.cell(row=row_num, column=2).value = 'error'

        
    #max_row = str(max_row)
    #R_max_row = '"R" + max_row' 
    #add_sort_condition('A2:R_max_row', descending=False)
    #add_filter_column(self, 1, vals, blank=False)   
  
    return sheet

#fpygrouopingH############################################################
"""
fpygrouopingH : Group のcodeに従って、グループ分けし、Group列,Stage列を書き換える。
                Umotion 未経産牛一覧より
ｖ1.0
2021/10/3
@author: jicc
"""

#! python3
def fpygrouopingH( wbN, sheetN, VWPm, VWPM ):
    """
    Parameters
    ----------
    wbN : 　str  
        対象となるExcelファイル名　拡張子.xlsxをつけ、　''でくくる。
    sheetN : str
        対象となるsheet
    VWPm : int
        授精待機開始月齢
    VWPM : int
        授精開始月齢
   
    Returns
    -------
    None.

    """
       
    #import openpyxl
    #import datetime
    
    wb = openpyxl.load_workbook(wbN)
    sheetN = wb[sheetN]			#.get_sheet_by_name(sheetN)
    max_row = sheetN.max_row
        
    for row_num in range(2, max_row + 1):     #先頭行をスキップ
        
        stage = sheetN.cell(row=row_num, column=3).value  #Stage
        area = sheetN.cell(row=row_num, column=5).value    #牛群
        AItimes = sheetN.cell(row=row_num, column=11).value #授精回数
   
        if area == '8.預託':
            sheetN.cell(row=row_num, column=2).value = 9
            sheetN.cell(row=row_num, column=3).value = '預託'
        else:
            if AItimes == 0:
                ageinmonths = sheetN.cell(row=row_num, column=10).value
                #月齢
                if ageinmonths < VWPm:
                    sheetN.cell(row=row_num, column=2).value = 8
                    sheetN.cell(row=row_num, column=3).value = '哺育'
                elif ageinmonths >= VWPm and ageinmonths < VWPM:
                    sheetN.cell(row=row_num, column=2).value = 1
                    sheetN.cell(row=row_num, column=3).value = '待機'
                elif ageinmonths >= VWPM:
                    sheetN.cell(row=row_num, column=2).value = 2
                    sheetN.cell(row=row_num, column=3).value = 'AI待ち 未授精'
            elif AItimes > 0:
                    if stage == '未受胎(－)':
                        sheetN.cell(row=row_num, column=2).value = 3
                        sheetN.cell(row=row_num, column=3).value = 'AI待ち 授精済'
                    elif stage == '授精':
                        sheetN.cell(row=row_num, column=2).value = 4
                        sheetN.cell(row=row_num, column=3).value = '妊娠鑑定予定'
                    elif stage == '受胎(＋)':
                        sheetN.cell(row=row_num, column=2).value = 5
                        sheetN.cell(row=row_num, column=3).value = '妊娠鑑定+'
                    else:
                        sheetN.cell(row=row_num, column=2).value = 'Error'
            else:
                sheetN.cell(row=row_num, column=2).value = 'Error'
     
    wb.save(wbN)
    
#fpygrouopingH_s############################################################
"""
fpygrouopingH_s : Group のcodeに従って、グループ分けし、Group列,Stage列を書き換える。
                Umotion 未経産牛一覧より
ｖ1.0
2021/10/3
@author: jicc
"""

#! python3
def fpygrouopingH_s( sheet, VWPm, VWPM ):
    """
    Parameters
    ----------
    sheet : worksheet.worksheet.Worksheet
         worksheet object　　　　　　
         シート名:"yyyymmddHeifer00"
    VWPm : int
        授精待機開始月齢
    VWPM : int
        授精開始月齢
   
    Returns
    -------
    None.

    """
        
    for row_num in range(2, sheet.max_row + 1):     #先頭行をスキップ
        
        stage = sheet.cell(row=row_num, column=3).value  #Stage
        area = sheet.cell(row=row_num, column=5).value    #牛群
        AItimes = sheet.cell(row=row_num, column=11).value #授精回数
   
        if area == '8.預託':
            sheet.cell(row=row_num, column=2).value = 9
            sheet.cell(row=row_num, column=3).value = '預託'
        else:
            if AItimes == 0:
                ageinmonths = sheet.cell(row=row_num, column=10).value
                #月齢
                if ageinmonths < VWPm:
                    sheet.cell(row=row_num, column=2).value = 8
                    sheet.cell(row=row_num, column=3).value = '哺育'
                elif ageinmonths >= VWPm and ageinmonths < VWPM:
                    sheet.cell(row=row_num, column=2).value = 1
                    sheet.cell(row=row_num, column=3).value = '待機'
                elif ageinmonths >= VWPM:
                    sheet.cell(row=row_num, column=2).value = 2
                    sheet.cell(row=row_num, column=3).value = 'AI待ち 未授精'
            elif AItimes > 0:
                    if stage == '未受胎(－)':
                        sheet.cell(row=row_num, column=2).value = 3
                        sheet.cell(row=row_num, column=3).value = 'AI待ち 授精済'
                    elif stage == '授精':
                        sheet.cell(row=row_num, column=2).value = 4
                        sheet.cell(row=row_num, column=3).value = '妊娠鑑定予定'
                    elif stage == '受胎(＋)':
                        sheet.cell(row=row_num, column=2).value = 5
                        sheet.cell(row=row_num, column=3).value = '妊娠鑑定+'
                    else:
                        sheet.cell(row=row_num, column=2).value = 'Error'
            else:
                sheet.cell(row=row_num, column=2).value = 'Error'
     
    return sheet

#fpydysoffstAI#############################################################
"""
fpydysoffstAI : 初回授精日数を計算する

ｖ1.0
2021/5/2
@author: jicc
"""

#! python3
def fpydysoffstAI( wbN, sheetN ):
    """
    Parameters
    ----------
    wbN : 　str  
        対象となるExcelファイル名　拡張子.xlsxをつけ、　''でくくる。
    sheetN : str
        対象となるsheet
    
    Returns
    -------
    None.

    """
       
    #import openpyxl
    #import datetime
    
    wb = openpyxl.load_workbook(wbN)
    sheetN = wb[sheetN]     		#.get_sheet_by_name(sheetN)
    max_row = sheetN.max_row
        
    for row_num in range(2, max_row + 1):     #先頭行をスキップ
        
        clvingDate = sheetN.cell(row=row_num, column=9).value #分娩日
        fstAIDate = sheetN.cell(row=row_num, column=12).value #初回授精日
        if (clvingDate != None) and (fstAIDate != None):
            daysoffstAI = fstAIDate - clvingDate
            #timedelta (datetime - datetime)
            daysoffstAI = daysoffstAI.days		#add 2021/12/8 v1.01
            #int
            #初回授精日数
            sheetN.cell(row=row_num, column=13).value = daysoffstAI
            #初回授精日数を入力
        
    wb.save(wbN)

#fpydysoffstAI_s#############################################################
"""
fpydysoffstAI_s : 初回授精日数を計算する
sheet version
ｖ1.0
2023/7/16
@author: jicc
"""

#! python3
def fpydysoffstAI_s( sheet ):
    """
    Parameters
    ----------
    sheet : worksheet.worksheet.Worksheet
         worksheet object　　　　　　
         シート名:"yyyymmddCow00"
    
    Returns
    
    -------
    sheet : worksheet.worksheet.Worksheet
         worksheet object　　　　　　
         シート名:"yyyymmddCow00"

    """
    for row_num in range(2, sheet.max_row + 1):     #先頭行をスキップ
        
        clvingDate = sheet.cell(row=row_num, column=9).value #分娩日
        fstAIDate = sheet.cell(row=row_num, column=12).value #初回授精日
        if (clvingDate != None) and (fstAIDate != None):
            daysoffstAI = fstAIDate - clvingDate
            #timedelta (datetime - datetime)
            daysoffstAI = daysoffstAI.days		#add 2021/12/8 v1.01
            #int
            #初回授精日数
            sheet.cell(row=row_num, column=13).value = daysoffstAI
            #初回授精日数を入力
        
    return sheet

#fpymnthsoffstAI##########################################################
"""
fpymnthsoffstAI : 初回授精月齢を計算する

ｖ1.0
2021/10/4
@author: jicc
"""

#! python3
def fpymnthsoffstAI( wbN, sheetN ):
    """
    Parameters
    ----------
    wbN : 　str  
        対象となるExcelファイル名　拡張子.xlsxをつけ、　''でくくる。
    sheetN : str
        対象となるsheet
    
    Returns
    -------
    None.

    """
       
    #import openpyxl
    #import datetime
    
    wb = openpyxl.load_workbook(wbN)
    sheetN = wb[sheetN]     		#.get_sheet_by_name(sheetN)
    max_row = sheetN.max_row
        
    for row_num in range(2, max_row + 1):     #先頭行をスキップ
        
        birthday = sheetN.cell(row=row_num, column=7).value #分娩日
        fstAIDate = sheetN.cell(row=row_num, column=12).value #初回授精日
        if (birthday != None) and (fstAIDate != None):
            daysoffstAI = fstAIDate - birthday
            daysoffstAI = daysoffstAI.days
            #初回授精日数
            monthsoffstAI = daysoffstAI/30
            #初回授精
            sheetN.cell(row=row_num, column=13).value = monthsoffstAI
            #初回授精月齢を入力
        
    wb.save(wbN)


#fpymnthsoffstAI_s##########################################################
"""
fpymnthsoffstAI : 初回授精月齢を計算する
sheet version
ｖ1.0
2023/7/6
@author: jicc
"""

#! python3
def fpymnthsoffstAI_s( sheet ):
    """
    Parameters
    ----------
    sheet : worksheet.worksheet.Worksheet
         worksheet object　　　　　　
         シート名:"yyyymmddHeifer00"
    
    Returns
    
    -------
    sheet : worksheet.worksheet.Worksheet
         worksheet object　　　　　　
         シート名:"yyyymmddHeifer00"

    """
      
    for row_num in range(2, sheet.max_row + 1):     #先頭行をスキップ
        
        birthday = sheet.cell(row=row_num, column=7).value #分娩日
        fstAIDate = sheet.cell(row=row_num, column=12).value #初回授精日
        if (birthday != None) and (fstAIDate != None):
            daysoffstAI = fstAIDate - birthday
            daysoffstAI = daysoffstAI.days
            #初回授精日数
            monthsoffstAI = daysoffstAI/30
            #初回授精
            sheet.cell(row=row_num, column=13).value = monthsoffstAI
            #初回授精月齢を入力
        
    return sheet

#fpyopenMnths###############################################################
"""
fpyopenMnths : 空胎月数を計算する
ｖ1.0
2021/10/4
@author: jicc
"""

#! python3
def fpyopenMnths( wbN, sheetN ):
    """
    Parameters
    ----------
    wbN : 　str  
        対象となるExcelファイル名　拡張子.xlsxをつけ、　''でくくる。
    sheetN : str
        対象となるsheet
    
    Returns
    -------
    None.

    """
       
    #import openpyxl
    #import datetime
    
    wb = openpyxl.load_workbook(wbN)
    sheet = wb[sheetN]         #.get_sheet_by_name(sheetN)
    max_row = sheet.max_row
        
    for row_num in range(2, max_row + 1):     #先頭行をスキップ
        
        Group = sheet.cell(row=row_num, column=2).value #Group
        birthday = sheet.cell(row=row_num, column=7).value #生年月日
        lstAIDate = sheet.cell(row=row_num, column=14).value #最終授精日
        bDate = sheet.cell(row=row_num, column=18).value #基準日
        
        if birthday != None: #分娩日があるなら
            if Group <= 4 : #待機、授精待ち、妊娠鑑定予定
                openDays = bDate - birthday
                openDays = openDays.days
                # 空胎日数 = 基準日 - 生年月日
                openMonths = openDays/30
                #空胎月数
                
            elif Group == 5: #鑑定+なら
                openDays = lstAIDate - birthday 
                openDays = openDays.days
                #空胎日数 = 最終授精日 - 生年月日
                openMonths = openDays/30
                #空胎月数
            elif Group == 8:
                openDays = bDate - birthday
                openDays = openDays.days
                # 空胎日数 = 基準日 - 生年月日
                openMonths = openDays/30
                #空胎月数
            else: 
                openMonths = None
                
        else: #分娩日不明なら
            openMonths = None

        sheet.cell(row=row_num, column=17).value = openMonths
            #空胎日数を入力　T列
        
    wb.save(wbN)

#fpyopenMnths_s###############################################################
"""
fpyopenMnths_s : 空胎月数を計算する
sheet version
ｖ1.0
2023/7/6
@author: jicc
"""

#! python3
def fpyopenMnths_s( sheet ):
    """
    Parameters
    ----------
    sheet : worksheet.worksheet.Worksheet
         worksheet object　　　　　　
         シート名:"yyyymmddHeifer00"
    
    Returns
    -------
    sheet : worksheet.worksheet.Worksheet
         worksheet object　　　　　　
         シート名:"yyyymmddHeifer00"

    """
            
    for row_num in range(2, sheet.max_row + 1):     #先頭行をスキップ
        
        Group = sheet.cell(row=row_num, column=2).value #Group
        birthday = sheet.cell(row=row_num, column=7).value #生年月日
        lstAIDate = sheet.cell(row=row_num, column=14).value #最終授精日
        bDate = sheet.cell(row=row_num, column=18).value #基準日
        
        if birthday != None: #分娩日があるなら
            if Group <= 4 : #待機、授精待ち、妊娠鑑定予定
                openDays = bDate - birthday
                openDays = openDays.days
                # 空胎日数 = 基準日 - 生年月日
                openMonths = openDays/30
                #空胎月数
                
            elif Group == 5: #鑑定+なら
                openDays = lstAIDate - birthday 
                openDays = openDays.days
                #空胎日数 = 最終授精日 - 生年月日
                openMonths = openDays/30
                #空胎月数
            elif Group == 8:
                openDays = bDate - birthday
                openDays = openDays.days
                # 空胎日数 = 基準日 - 生年月日
                openMonths = openDays/30
                #空胎月数
            else: 
                openMonths = None
                
        else: #分娩日不明なら
            openMonths = None

        sheet.cell(row=row_num, column=17).value = openMonths
            #空胎日数を入力　T列
        
    return sheet

#fpyinput_PT##################################################################
"""
fpyinput_PT : 繁殖データに、AI台帳から最終の鑑定結果を入力する
Cow版
Umotion仕様：Group4 -> Group5の場合、空胎日数を書き換える
    
ｖ1.2
2021/9/12
@author: jicc
"""

#! python3
# MH_GList.xlsx のデータを取得して、　MH_MQ.xlsx に書き込む

def fpyinput_PT(wbAIＮ, sheetAIN, wbRPDN, sRPDN):
    """
    繁殖データに、AI台帳から最終の鑑定結果を入力する
    Cow版
    Umotion仕様：Group4 -> Group5の場合、空胎月数を書き換える

    Parameters
    ----------
    wbAIＮ : str 
        AIdata workbook name
        "MH_AI_.xlsx"
    sheetAIN : str 
        AI data sheet name
        "yyyy"
    wbRPDN : str 
        workbook name
        "MH_RPDu.xlsx"
    sRPDN : str 
        Cows' data sheet name
        "yyyymmddCow01"

    Returns
    -------
    None.

    """
    

    #import openpyxl
    #from datetime import date
    
    
    wbAI = openpyxl.load_workbook(wbAIN)        #MH_AI.xlsx
    sheetAI = wbAI[sheetAIN]      #.get_sheet_by_name(sheetAIN)

    wbRPD = openpyxl.load_workbook(wbRPDN)      #MH_RPD.xlsx
    sheetRPD = wbRPD[sRPDN]         #.get_sheet_by_name(sRPDN)


       #2行目からスタート
    for i in range(2, sheetRPD.max_row + 1):     #先頭行をスキップ
        
        #繁殖ファイルのGroupを取得
        Group = sheetRPD.cell(row=i, column=2).value   #Group
        
        if Group == 4:
            cowidNoRPD = sheetRPD.cell(row=i, column=6).value
            #繁殖ファイルの個体識別番号
            lstAIDtate = sheetRPD.cell(row=i, column=14).value
            #繁殖ファイルの最終授精日
            clvingDate = sheetRPD.cell(row=i, column=9).value
            #繁殖ファイルの分娩日
                        #AIyear = datetime.year(lstAIDtate)
            #最終授精日の年数を取得　yyyy
            AIyear = str(lstAIDtate.year)
            #授精年を　str型'yyyy'で取得
            
            sheetAI = wbAI[AIyear]        #.get_sheet_by_name(AIyear)
            #参照用ｗｂAIのsheet'yyyy'をアクティブにする
            #sheetAIN いらないかも　2021/6/20
            
            for j in range(2, sheetAI.max_row + 1 ):    #sheetAI.max_row + 1
                    
                cowidNoAI = sheetAI.cell(row=j, column=2).value
                #AI台帳の耳標番号
                lstAI_date = sheetAI.cell(row=j, column=10).value
                #AI台帳の最終授精日
                if (cowidNoRPD == cowidNoAI) and (lstAIDtate == lstAI_date):
                    
                    PT = sheetAI.cell(row=j, column=16).value
                    #AI台帳のPT（鑑定結果、+, -, ?, NONE）を取得
                    if PT == '+':
                        expDayofnxtclving = sheetAI.cell(row=j, column=18).value
                        #分娩予定日
                        sheetRPD.cell(row=i, column=2).value = 5
                        #Groupを　5　に変更
                        sheetRPD.cell(row=i, column=3).value = '妊娠鑑定+'
                        #Stage　を　'妊娠鑑定+'に変更
                        sheetRPD.cell(row=i, column=16).value = expDayofnxtclving
                        #分娩予定日を入力
                        openDays = lstAIDtate - clvingDate #空胎日数
                        sheetRPD.cell(row=i, column=17).value = openDays
                        #空胎日数を入力
                        break
                    
                    elif PT == '-':
                        sheetRPD.cell(row=i, column=2).value = 3
                        #Groupを　3　に変更
                        sheetRPD.cell(row=i, column=3).value = 'AI待ち 授精済'
                        #Stage　を　'AI待ち 授精済'に変更
                        
                        #bDate = sheetRPD.cell(row=i, column=18).value
                        #openDays = bDate - clvingDate #空胎日数
                        #sheetRPD.cell(row=i, column=17).value = openDays
                        #空胎日数を変更 2021/6/16追加
                        break
                    
                    else:   #None, ?, -?, +? など　何もしない
                        break
                else:   #何もしない
                    continue
        else:    #Group != 4 何もしない
            continue
        
    wbRPD.save(wbRPDN)

#fpyinput_PT_s##################################################################
"""
fpyinput_PT_s : 繁殖データに、AI台帳から最終の鑑定結果を入力する
Cow版
Umotion仕様：Group4 -> Group5の場合、空胎日数を書き換える
sheet version    
ｖ1.0
2023/7/16
@author: jicc
"""

#! python3
# MH_GList.xlsx のデータを取得して、　MH_MQ.xlsx に書き込む

def fpyinput_PT_s(wbAI, sheetRPD):
    """
    繁殖データに、AI台帳から最終の鑑定結果を入力する

    Parameters
    ----------
    wbAI : workbook.workbook.workbook
         workbook object
         "MH_AI_.xlsx"
    sheetRPD : worksheet.worksheet.Worksheet
         worksheet object　　　　　　
         "yyyymmddHeifer01"

    Returns
    -------
    sheetRPD : worksheet.worksheet.Worksheet
         worksheet object　

    """
    
       #2行目からスタート
    for i in range(2, sheetRPD.max_row + 1):     #先頭行をスキップ
        
        #繁殖ファイルのGroupを取得
        Group = sheetRPD.cell(row=i, column=2).value   #Group
        
        if Group == 4:
            cowidNoRPD = sheetRPD.cell(row=i, column=6).value
            #繁殖ファイルの個体識別番号
            lstAIDtate = sheetRPD.cell(row=i, column=14).value
            #繁殖ファイルの最終授精日
            clvingDate = sheetRPD.cell(row=i, column=9).value
            #繁殖ファイルの分娩日
                        #AIyear = datetime.year(lstAIDtate)
            #最終授精日の年数を取得　yyyy
            AIyear = str(lstAIDtate.year)
            #授精年を　str型'yyyy'で取得
            
            sheetAI = wbAI[AIyear]        #.get_sheet_by_name(AIyear)
            #参照用ｗｂAIのsheet'yyyy'をアクティブにする
            #sheetAIN いらないかも　2021/6/20
            
            for j in range(2, sheetAI.max_row + 1 ):    #sheetAI.max_row + 1
                    
                cowidNoAI = sheetAI.cell(row=j, column=2).value
                #AI台帳の耳標番号
                lstAI_date = sheetAI.cell(row=j, column=10).value
                #AI台帳の最終授精日
                if (cowidNoRPD == cowidNoAI) and (lstAIDtate == lstAI_date):
                    
                    PT = sheetAI.cell(row=j, column=16).value
                    #AI台帳のPT（鑑定結果、+, -, ?, NONE）を取得
                    if PT == '+':
                        expDayofnxtclving = sheetAI.cell(row=j, column=18).value
                        #分娩予定日
                        sheetRPD.cell(row=i, column=2).value = 5
                        #Groupを　5　に変更
                        sheetRPD.cell(row=i, column=3).value = '妊娠鑑定+'
                        #Stage　を　'妊娠鑑定+'に変更
                        sheetRPD.cell(row=i, column=16).value = expDayofnxtclving
                        #分娩予定日を入力
                        openDays = lstAIDtate - clvingDate #空胎日数
                        sheetRPD.cell(row=i, column=17).value = openDays
                        #空胎日数を入力
                        break
                    
                    elif PT == '-':
                        sheetRPD.cell(row=i, column=2).value = 3
                        #Groupを　3　に変更
                        sheetRPD.cell(row=i, column=3).value = 'AI待ち 授精済'
                        #Stage　を　'AI待ち 授精済'に変更
                        
                        #bDate = sheetRPD.cell(row=i, column=18).value
                        #openDays = bDate - clvingDate #空胎日数
                        #sheetRPD.cell(row=i, column=17).value = openDays
                        #空胎日数を変更 2021/6/16追加
                        break
                    
                    else:   #None, ?, -?, +? など　何もしない
                        break
                else:   #何もしない
                    continue
        else:    #Group != 4 何もしない
            continue
        
    return sheetRPD
    
#fpyinput_PTH############################################################
"""
fpyinput_PTH : 繁殖データに、AI台帳から最終の鑑定結果を入力する
Heifer版
Umotion仕様：Group4 -> Group5の場合、空胎月数を書き換える
    
ｖ1.0
2021/10/9
@author: jicc
"""
#! python3
# MH_GList.xlsx のデータを取得して、　MH_MQ.xlsx に書き込む

def fpyinput_PTH(wbAIＮ, sheetAIN, wbRPDN, sRPDN):
    """
    繁殖データに、AI台帳から最終の鑑定結果を入力する
    Heifer版
    Umotion仕様：Group4 -> Group5の場合、空胎月数を書き換える

    Parameters
    ----------
    wbAIＮ : str 
        AIdata workbook name
        "MH_AI_.xlsx"
    sheetAIN : str 
        AI data sheet name
        "yyyy"
    wbRPDN : str 
        workbook name
        "MH_RPDu.xlsx"
    sRPDN : str 
        Heifers' data sheet name
        "yyyymmddHeifer01"

    Returns
    -------
    None.

    """
    

    #import openpyxl
    #from datetime import date
    
    
    wbAI = openpyxl.load_workbook(wbAIN)        #MH_AI.xlsx
    sheetAI = wbAI[sheetAIN]      #.get_sheet_by_name(sheetAIN)

    wbRPD = openpyxl.load_workbook(wbRPDN)      #MH_RPDu.xlsx
    sheetRPD = wbRPD[sRPDN]         #.get_sheet_by_name(sRPDN)


       #2行目からスタート
    for i in range(2, sheetRPD.max_row + 1):     #先頭行をスキップ
        
        #繁殖ファイルのGroupを取得
        Group = sheetRPD.cell(row=i, column=2).value   #Group
        
        if Group == 4:
            cowidNoRPD = sheetRPD.cell(row=i, column=6).value
            #繁殖ファイルの個体識別番号
            birthday = sheetRPD.cell(row=i, column=7).value
            #繁殖ファイルの生年月日
            lstAIDtate = sheetRPD.cell(row=i, column=14).value
            #繁殖ファイルの最終授精日
            
            #clvingDate = sheetRPD.cell(row=i, column=9).value
            #繁殖ファイルの分娩日
            #AIyear = datetime.year(lstAIDtate)
            #最終授精日の年数を取得　yyyy
             
            AIyear = str(lstAIDtate.year)
            #授精年を　str型'yyyy'で取得
            
            sheetAI = wbAI[AIyear]        #.get_sheet_by_name(AIyear)
            #参照用ｗｂAIのsheet'yyyy'をアクティブにする
            #sheetAIN いらないかも　2021/6/20
            
            for j in range(2, sheetAI.max_row + 1 ):    #sheetAI.max_row + 1
                    
                cowidNoAI = sheetAI.cell(row=j, column=2).value
                #AI台帳の耳標番号
                lstAI_date = sheetAI.cell(row=j, column=10).value
                #AI台帳の最終授精日
                if (cowidNoRPD == cowidNoAI) and (lstAIDtate == lstAI_date):
                    
                    PT = sheetAI.cell(row=j, column=16).value
                    #AI台帳のPT（鑑定結果、+, -, ?, NONE）を取得
                    if PT == '+':
                        expDayofnxtclving = sheetAI.cell(row=j, column=18).value
                        #分娩予定日
                        sheetRPD.cell(row=i, column=2).value = 5
                        #Groupを　5　に変更
                        sheetRPD.cell(row=i, column=3).value = '妊娠鑑定+'
                        #Stage　を　'妊娠鑑定+'に変更
                        sheetRPD.cell(row=i, column=16).value = expDayofnxtclving
                        #分娩予定日を入力
                        openDays = lstAIDtate - birthday 
                        #空胎日数=最終授精日 - 生年月日
                        openDays = openDays.days
                        openMonths = openDays/30
                        #空胎月数
                        sheetRPD.cell(row=i, column=17).value = openMonths
                        #空胎月数を入力
                        break
                    
                    elif PT == '-':
                        sheetRPD.cell(row=i, column=2).value = 3
                        #Groupを　3　に変更
                        sheetRPD.cell(row=i, column=3).value = 'AI待ち 授精済'
                        #Stage　を　'AI待ち 授精済'に変更
                        
                        #bDate = sheetRPD.cell(row=i, column=18).value
                        #openDays = bDate - clvingDate #空胎日数
                        #sheetRPD.cell(row=i, column=17).value = openDays
                        #空胎日数を変更 2021/6/16追加
                        break
                    
                    else:   #None, ?, -?, +? など　何もしない
                        break
                else:   #何もしない
                    continue
        else:    #Group != 4 何もしない
            continue
        
    wbRPD.save(wbRPDN)

#fpyinput_PTH_s############################################################
"""
fpyinput_PTH_s : 繁殖データに、AI台帳から最終の鑑定結果を入力する
Heifer版
Umotion仕様：Group4 -> Group5の場合、空胎月数を書き換える
sheet version
    
ｖ1.0
2022/7/7
@author: jicc
"""
#! python3
# MH_GList.xlsx のデータを取得して、　MH_MQ.xlsx に書き込む

def fpyinput_PTH_s( wbAI, sheetRPD):
    """
    繁殖データに、AI台帳から最終の鑑定結果を入力する
    Heifer版
    Umotion仕様：Group4 -> Group5の場合、空胎月数を書き換える
    sheet version
    *) 授精年によってsheet "yyyy" を変更するため、　引数1をworkbook object wbAIとしている。
    AIsheet を1枚で統一する場合は、引数１をworksheet objectに変更する必要あり
    2023/7/7
    Parameters
    ----------
    wbAI : workbook.workbook.workbook
         workbook object
         "MH_AI_.xlsx"
    sheetRPD : worksheet.worksheet.Worksheet
         worksheet object　　　　　　
         "yyyymmddHeifer01"

    Returns
    -------
    sheetRPD : worksheet.worksheet.Worksheet
         worksheet object　

    """
       #2行目からスタート
    for i in range(2, sheetRPD.max_row + 1):     #先頭行をスキップ
        
        #繁殖ファイルのGroupを取得
        Group = sheetRPD.cell(row=i, column=2).value   #Group
        
        if Group == 4:
            cowidNoRPD = sheetRPD.cell(row=i, column=6).value
            #繁殖ファイルの個体識別番号
            birthday = sheetRPD.cell(row=i, column=7).value
            #繁殖ファイルの生年月日
            lstAIDtate = sheetRPD.cell(row=i, column=14).value
            #繁殖ファイルの最終授精日
            
            #clvingDate = sheetRPD.cell(row=i, column=9).value
            #繁殖ファイルの分娩日
            #AIyear = datetime.year(lstAIDtate)
            #最終授精日の年数を取得　yyyy
             
            AIyear = str(lstAIDtate.year)   
            #授精年を　str型'yyyy'で取得
            
            sheetAI = wbAI[AIyear]        #*)
            #参照用ｗｂAIのsheet'yyyy'をアクティブにする
            #sheetAIN いらないかも　2021/6/20
            
            for j in range(2, sheetAI.max_row + 1 ):    #sheetAI.max_row + 1
                    
                cowidNoAI = sheetAI.cell(row=j, column=2).value
                #AI台帳の耳標番号
                lstAI_date = sheetAI.cell(row=j, column=10).value
                #AI台帳の最終授精日
                if (cowidNoRPD == cowidNoAI) and (lstAIDtate == lstAI_date):
                    
                    PT = sheetAI.cell(row=j, column=16).value
                    #AI台帳のPT（鑑定結果、+, -, ?, NONE）を取得
                    if PT == '+':
                        expDayofnxtclving = sheetAI.cell(row=j, column=18).value
                        #分娩予定日
                        sheetRPD.cell(row=i, column=2).value = 5
                        #Groupを　5　に変更
                        sheetRPD.cell(row=i, column=3).value = '妊娠鑑定+'
                        #Stage　を　'妊娠鑑定+'に変更
                        sheetRPD.cell(row=i, column=16).value = expDayofnxtclving
                        #分娩予定日を入力
                        openDays = lstAIDtate - birthday 
                        #空胎日数=最終授精日 - 生年月日
                        openDays = openDays.days
                        openMonths = openDays/30
                        #空胎月数
                        sheetRPD.cell(row=i, column=17).value = openMonths
                        #空胎月数を入力
                        break
                    
                    elif PT == '-':
                        sheetRPD.cell(row=i, column=2).value = 3
                        #Groupを　3　に変更
                        sheetRPD.cell(row=i, column=3).value = 'AI待ち 授精済'
                        #Stage　を　'AI待ち 授精済'に変更
                        
                        #bDate = sheetRPD.cell(row=i, column=18).value
                        #openDays = bDate - clvingDate #空胎日数
                        #sheetRPD.cell(row=i, column=17).value = openDays
                        #空胎日数を変更 2021/6/16追加
                        break
                    
                    else:   #None, ?, -?, +? など　何もしない
                        break
                else:   #何もしない
                    continue
        else:    #Group != 4 何もしない
            continue
        
    return sheetRPD

#fpyAgeinDays#############################################################
"""
fpyAgeinDays: 	日齢を計算する
未経産用 yyyymmddHeifer00 日齢を計算入力する
 v1.0
2021/12/9
@author: jicc
"""

#! python3
def fpyAgeinDays( wbN, sheetN ):
    """
    Parameters
    ----------
    wbN : 　str  
        対象となるExcelファイル名　拡張子.xlsxをつけ、　''でくくる。
    sheetN : str
        対象となるsheet
    
    Returns
    -------
    None.

    """
       
    #import openpyxl
    #import datetime
    
    wb = openpyxl.load_workbook(wbN)
    sheet = wb[sheetN]         #.get_sheet_by_name(sheetN)
    max_row = sheet.max_row
        
    for row_num in range(2, max_row + 1):     #先頭行をスキップ
        
        bDate = sheet.cell(row=row_num, column=18).value #基準日
        birthday = sheet.cell(row=row_num, column=7).value  #生年月日
        ageindays = bDate - birthday  #日齢
        ageindays = ageindays.days
                
        sheet.cell(row=row_num, column=9).value = ageindays
            #日齢を入力
        
    wb.save(wbN)

#fpyAgeinDays_s#############################################################
"""
fpyAgeinDays_s: 	日齢を計算する
未経産用 yyyymmddHeifer00 日齢を計算入力する
sheet version
 v1.0
2021/12/9
@author: jicc
"""

#! python3
def fpyAgeinDays_s( sheet ):
    """
    Parameters
    ----------
    sheet : worksheet.worksheet.Worksheet
         worksheet object　　　　　　
         シート名:"yyyymmddHeifer00"
    
    Returns
    -------
    sheet : worksheet.worksheet.Worksheet
         worksheet object　　　　　　
         シート名:"yyyymmddHeifer00"

    """
       
    for row_num in range(2, sheet.max_row + 1):     #先頭行をスキップ
        
        bDate = sheet.cell(row=row_num, column=18).value #基準日
        birthday = sheet.cell(row=row_num, column=7).value  #生年月日
        ageindays = bDate - birthday  #日齢
        ageindays = ageindays.days
                
        sheet.cell(row=row_num, column=9).value = ageindays
            #日齢を入力
        
    return sheet

#fpyAgeinMnths#############################################################
"""
fpyAgeinMnths: 	月齢を計算する
未経産用 yyyymmddHeifer00 分娩後日数列に月齢を入力する
 v1.0
2021/10/3
@author: jicc
"""

#! python3
def fpyAgeinMnths( wbN, sheetN ):
    """
    Parameters
    ----------
    wbN : 　str  
        対象となるExcelファイル名　拡張子.xlsxをつけ、　''でくくる。
    sheetN : str
        対象となるsheet
    
    Returns
    -------
    None.

    """
       
    #import openpyxl
    #import datetime
    
    wb = openpyxl.load_workbook(wbN)
    sheet = wb[sheetN]         #.get_sheet_by_name(sheetN)
    max_row = sheet.max_row
        
    for row_num in range(2, max_row + 1):     #先頭行をスキップ
        
        bDate = sheet.cell(row=row_num, column=18).value #基準日
        birthday = sheet.cell(row=row_num, column=7).value  #生年月日
        daysfrmBirth = bDate - birthday  #日齢
        daysfrmBirth = daysfrmBirth.days
        ageinmonths = daysfrmBirth/30
        
        sheet.cell(row=row_num, column=10).value = ageinmonths
            #月齢を入力
        
    wb.save(wbN)

#fpyAgeinMnths_s#############################################################
"""
fpyAgeinMnths_s: 	月齢を計算する
未経産用 yyyymmddHeifer00 分娩後日数列に月齢を入力する
sheet version
 v1.0
2023/7/5
@author: jicc
"""

#! python3
def fpyAgeinMnths_s( sheet ):
    """
    Parameters
    ----------
   sheet : worksheet.worksheet.Worksheet
        worksheet object　　　　　　
        シート名:"yyyymmddHeifer00"
    
    Returns
    -------
    sheet : worksheet.worksheet.Worksheet
         worksheet object　　　　　　
         シート名:"yyyymmddHeifer00"

    """
            
    for row_num in range(2, sheet.max_row + 1):     #先頭行をスキップ
        
        bDate = sheet.cell(row=row_num, column=18).value #基準日
        birthday = sheet.cell(row=row_num, column=7).value  #生年月日
        daysfrmBirth = bDate - birthday  #日齢
        daysfrmBirth = daysfrmBirth.days
        ageinmonths = daysfrmBirth/30
        
        sheet.cell(row=row_num, column=10).value = ageinmonths
            #月齢を入力
        
    return sheet

#fpyopenDays###############################################################
"""
fpyopenDays : 空胎日数を計算する
過去の日付でUmotion経産牛一覧.csvを取得した時、grouping後に空胎日数を再計算する。
 分娩日　あるなし、　Groupによる場合分けに変更した。
umotion式空胎日数　
 鑑定待ち空胎日数　=　基準日　-　分娩日とする
ｖ3.0
2021/7/7
@author: jicc
"""

#! python3
def fpyopenDays( wbN, sheetN ):
    """
    Parameters
    ----------
    wbN : 　str  
        対象となるExcelファイル名　拡張子.xlsxをつけ、　''でくくる。
    sheetN : str
        対象となるsheet
    
    Returns
    -------
    None.

    """
       
    #import openpyxl
    #import datetime
    
    wb = openpyxl.load_workbook(wbN)
    sheet = wb[sheetN]         #.get_sheet_by_name(sheetN)
    max_row = sheet.max_row
        
    for row_num in range(2, max_row + 1):     #先頭行をスキップ
        
        Group = sheet.cell(row=row_num, column=2).value #Group
        clvingDate = sheet.cell(row=row_num, column=9).value #分娩日
        lstAIDate = sheet.cell(row=row_num, column=14).value #最終授精日
        bDate = sheet.cell(row=row_num, column=18).value #基準日
        
        if clvingDate != None: #分娩日があるなら
            if Group <= 4: #待機、授精待ち、鑑定待ちなら
                openDays = bDate - clvingDate
                # 空胎日数 = 基準日 - 分娩日  timedelta (datetime - datetime)
                openDays = openDays.days
                #int
                
            elif Group >= 5: #鑑定+なら
                openDays = lstAIDate - clvingDate 
                #空胎日数 = 最終授精日 - 分娩日  timedelta (datetime - datetime)
                openDays = openDays.days
                #int
            else:
                continue
                
        else: #分娩日不明なら
            openDays = None
        #sheet.cell(row=row_num, column=19).value =sheet.cell(row=row_num, column=17).value 
		#既存のデータバックアップ
        sheet.cell(row=row_num, column=17).value = openDays
            #空胎日数を再計算
        
    wb.save(wbN)

#fpyopenDays_s###############################################################
"""
fpyopenDays_s : 空胎日数を計算する
過去の日付でUmotion経産牛一覧.csvを取得した時、grouping後に空胎日数を再計算する。
 分娩日　あるなし、　Groupによる場合分けに変更した。
umotion式空胎日数　
 鑑定待ち空胎日数　=　基準日　-　分娩日とする
sheet version
ｖ3.0
2023/7/16
@author: jicc
"""

#! python3
def fpyopenDays_s( sheet ):
    """
    Parameters
    ----------
    sheet : worksheet.worksheet.Worksheet
         worksheet object　　　　　　
         シート名: "yyyymmddCow00"
    
    Returns
    -------
    sheet : worksheet.worksheet.Worksheet
         worksheet object　　　　　　
         シート名: "yyyymmddCow00"

    """
         
    for row_num in range(2, sheet.max_row + 1):     #先頭行をスキップ
        
        Group = sheet.cell(row=row_num, column=2).value #Group
        clvingDate = sheet.cell(row=row_num, column=9).value #分娩日
        lstAIDate = sheet.cell(row=row_num, column=14).value #最終授精日
        bDate = sheet.cell(row=row_num, column=18).value #基準日
        
        if clvingDate != None: #分娩日があるなら
            if Group <= 4: #待機、授精待ち、鑑定待ちなら
                openDays = bDate - clvingDate
                # 空胎日数 = 基準日 - 分娩日  timedelta (datetime - datetime)
                openDays = openDays.days
                #int
                
            elif Group >= 5: #鑑定+なら
                openDays = lstAIDate - clvingDate 
                #空胎日数 = 最終授精日 - 分娩日  timedelta (datetime - datetime)
                openDays = openDays.days
                #int
                
        else: #分娩日不明なら
            openDays = None
        #sheet.cell(row=row_num, column=19).value =sheet.cell(row=row_num, column=17).value 
		#既存のデータバックアップ
        sheet.cell(row=row_num, column=17).value = openDays
            #空胎日数を再計算
        
    return sheet
    
#fpyDaysfrmlstAI###########################################################
"""
fpyDaysfrmlstAI: 	授精後日数を計算する
for cows and heifers
v1.0
2021/12/13
@author: jicc
"""

#! python3
def fpyDaysfrmlstAI( wbN, sheetN ):
    """
    Parameters
    ----------
    wbN : 　str  
        対象となるExcelファイル名　拡張子.xlsxをつけ、　''でくくる。
    sheetN : str
        対象となるsheet
    
    Returns
    -------
    None.

    """
       
    #import openpyxl
    #import datetime
    
    wb = openpyxl.load_workbook(wbN)
    sheet = wb[sheetN]         #.get_sheet_by_name(sheetN)
    max_row = sheet.max_row
        
    for row_num in range(2, max_row + 1):     #先頭行をスキップ
        
        bDate = sheet.cell(row=row_num, column=18).value #基準日
        lstAIDate = sheet.cell(row=row_num, column=14).value  #最終授精日
        if(lstAIDate != None):
            dysfrmlstAI = bDate - lstAIDate  #授精後日数
            dysfrmlstAI = dysfrmlstAI.days
                
            sheet.cell(row=row_num, column=15).value = dysfrmlstAI
            #授精後日数を入力
        
    wb.save(wbN)

#fpyDaysfrmlstAI_s###########################################################
"""
fpyDaysfrmlstAI_s: 	授精後日数を計算する
for cows and heifers
sheet version
v1.0
2023/7/6
@author: jicc
"""

#! python3
def fpyDaysfrmlstAI_s( sheet ):
    """
    Parameters
    ----------
    sheet : worksheet.worksheet.Worksheet
         worksheet object　　　　　　
         シート名:"yyyymmddHeifer00", "yyyymmddCow00"
    
    Returns
    -------
    sheet : worksheet.worksheet.Worksheet
         worksheet object　　　　　　
         シート名:"yyyymmddHeifer00", "yyyymmddCow00"

    """
           
    for row_num in range(2, sheet.max_row + 1):     #先頭行をスキップ
        
        bDate = sheet.cell(row=row_num, column=18).value #基準日
        lstAIDate = sheet.cell(row=row_num, column=14).value  #最終授精日
        if(lstAIDate != None):
            dysfrmlstAI = bDate - lstAIDate  #授精後日数
            dysfrmlstAI = dysfrmlstAI.days
                
            sheet.cell(row=row_num, column=15).value = dysfrmlstAI
            #授精後日数を入力
        
    return sheet

#fpyDaysfrmlstclving#########################################################
"""
fpyDaysfrmlstclving: 	分娩後日数を計算する
for cows
v1.0
2023/7/16
@author: jicc
"""

#! python3
def fpyDaysfrmlstclving( wbN, sheetN ):
    """
    Parameters
    ----------
    wbN : 　str  
        対象となるExcelファイル名　拡張子.xlsxをつけ、　''でくくる。
    sheetN : str
        対象となるsheet
    
    Returns
    -------
    None.

    """
       
    import openpyxl
    #import datetime
    
    wb = openpyxl.load_workbook(wbN)
    sheet = wb[sheetN]         #.get_sheet_by_name(sheetN)
            
    for row_num in range(2, sheet.max_row + 1):     #先頭行をスキップ
        
        bDate = sheet.cell(row=row_num, column=18).value #基準日
        lstclvingDate = sheet.cell(row=row_num, column=9).value  #分娩日
        if(lstclvingDate != None):
            dysfrmlstclving = bDate - lstclvingDate  #分娩後日数
            dysfrmlstclving = dysfrmlstclving.days
                
            sheet.cell(row=row_num, column=10).value = dysfrmlstclving
            #授精後日数を入力
        
    wb.save(wbN)

#fpyDaysfrmlstclving_s#########################################################
"""
fpyDaysfrmlstclving_s: 	分娩後日数を計算する
for cows
sheet version
v1.0
2023/7/16
@author: jicc
"""

#! python3
def fpyDaysfrmlstclving_s( sheet ):
    """
    Parameters
    ----------
    sheet : worksheet.worksheet.Worksheet
         worksheet object　　　　　　
         シート名: "yyyymmddCow00"
    
    Returns
    -------
    sheet : worksheet.worksheet.Worksheet
         worksheet object　　　　　　
         シート名: "yyyymmddCow00"

    """
             
    for row_num in range(2, sheet.max_row + 1):     #先頭行をスキップ
        
        bDate = sheet.cell(row=row_num, column=18).value #基準日
        lstclvingDate = sheet.cell(row=row_num, column=9).value  #分娩日
        if(lstclvingDate != None):
            dysfrmlstclving = bDate - lstclvingDate  #分娩後日数
            dysfrmlstclving = dysfrmlstclving.days
                
            sheet.cell(row=row_num, column=10).value = dysfrmlstclving
            #授精後日数を入力
        
    return sheet

#fpymh_rpdumanualHeifer00###################################################
"""
    fpymh_rpdumanualHeifer00 :
    makes a reproductive data list in heifers 
    from Umotion data yyyymmddHeiferorg #1
    v 1.1 
    add 2 parameters VWPm  and VWPM
    2023/7/17
    @author: inoue
    
"""
def fpymh_rpdumanualHeifer00(wbN, sorgN, sheetN, bDate, VWPm, VWPM):
    """
    makes a reproductive data list in heifers 
    from Umotion data yyyymmddHeiferorg #1
    
    Parameters
    ----------
    wbN : ワークブック名
        "MH_RPDu.xlsx"
    sorgN : データ参照シート
        "yyyymmddHeiferorg"
    sheetN : 未経産牛データシート
        "yyyymmddHeifer00"
    bDate : base date 基準日
        "yyyy/mm/dd"
    VWPm : int
           授精待機開始月齢 the age in months, a Farm waits to inseminate.
    VWPM : int
           授精開始月齢 the age in months, a Farm start to inseminate.

    Returns
    -------
    None.

    """
    #import mh_rpdu
    #import openpyxl
    #import fmstls
    
    wb = openpyxl.load_workbook(wbN)  #MH_RPDu.xlsx
    
    wb.create_sheet(title=sheetN, index=0)
    #make a new sheet 'yyyymmddHeifer00'
    sheet = wb[sheetN] #yyyymmddHeifer00
    sorg = wb[sorgN]
    
    #列名だけのsheet sheetNをさくせいする
    #make a sheet sheetN columns'name only    
    sheet = fpyNewSheetHeifer_s(sheet)
    print("2." + wbN + " に　sheet " + sheetN + " を作成しました。")
    
    #input data to sheet 'yyyymmdd'
    #sheet 'yyyymmddHeifer00'にデータ入力する
    sheet = fpyDF_RPDHeifer_s(sorg, sheet, bDate)
    print("3. Sheet " + sheetN + " に　データを入力しました。")
    
    #make idNo(column 6) 9digits to 10digits'strings
    #個体識別番号(column 6)を10桁文字列に統一する
    sheet = fmstls.fpyidNo_9to10_s( sheet, 6 )
    print("4. 個体識別番号を10桁文字列に統一しました。")
    
    #input the age in months to sheet'yyyymmddHeifer00' column 10
    #月齢を　sheet'yyyymmddHeifer00' column '月齢' (column 10) に入力する
    sheet = fpyAgeinMnths_s( sheet )
    print("5. sheet" + sheetN + " の 月齢を計算、入力しました。")
    
    # groupe all heifers with the code 'Group' and rewrite column 'Group' and 'Stage'
    # 育成牛全頭を　code'Group'に従ってグループ分けし、column 'Group' and 'Stage'を書き換える。
    #VWPm = 12   #授精待機開始月齢 
                #the age in months, a Farm waits to inseminate.
    #VWPM = 13   #授精開始月齢     
                #the age in months, a Farm start to inseminate.
    sheet = fpygrouopingH_s(sheet, VWPm, VWPM)
    print("6. sheet " + sheetN +  " をグルーピングしました")
   
    wb.save(wbN)

#fpymh_rpdumanualHeifer01###################################################
"""
    fpymh_rpdumanualHeifer01 :
    makes a reproductive data list in heifers 
    from Umotion data yyyymmddHeiferorg #2
    v 1.0
    2023/6/29
    @author: inoue
    
"""
def fpymh_rpdumanualHeifer01(wbN, sheetN, sheetN_ ):
    """
    makes a reproductive data list in heifers 
    from Umotion data yyyymmddHeiferorg #2

    Parameters
    ----------
    wbN : str 
        workbook name
        "MH_RPDu.xlsx"
    sheetN : str 
        Heifers' data sheet name
        "yyyymmddHeifer00"
    sheetN_ : str
        new sheet title
        "yyyymmddHeifer01"
   
    Returns
    -------
    None.

    """
 
    #import mh_rpdu
    #import openpyxl
    #import fmstls
    
    wb = openpyxl.load_workbook(wbN)  #MH_RPDu.xlsx
    sheet = wb[sheetN] #yyyymmddHeifer00
    #wbAI = openpyxl.load_workbook(wbAIN)  #MH_AI_.xlsx
    #sAI = wb[sAIN]  #yyyy
    
    #初回授精月数を計算する
    #
    sheet = fpymnthsoffstAI_s( sheet )
    print("9. sheet" + sheetN + " の 初回授精月齢を計算、入力しました。")
    
    #空胎月数を計算する
    #
    sheet = fpyopenMnths_s( sheet )
    print("10. sheet" + sheetN + " の 空胎月数を計算、入力しました。")
    
    #日齢を再計算する
    #
    sheet = fpyAgeinDays_s( sheet )
    print("11. sheet" + sheetN + " の 日齢を再計算しました。")
    
    #授精後日数を再計算する
    #
    sheet = fpyDaysfrmlstAI_s( sheet )
    print("12. sheet" + sheetN + " の 授精後日数を再計算しました。")
    
    #鑑定結果を書き込むシートをコピーして作成する
    #
    sheet_ = fmstls.fpysheet_copy_s( wb, sheet, sheetN_ )
    print("13. sheet" + sheetN + "をコピーして、" + sheetN_ + "を作成しました。")
    #sheet_を先頭に移動する
    wb.move_sheet( sheet_, offset=-wb.index(sheet_) )
    #sheet_ 先頭のシートをactiveにする
    wb.active = 0
    
    wb.save(wbN)

#fpymh_rpdumanualHeifer02###################################################
"""
    fpymh_rpdumanualHeifer02 :
    makes a reproductive data list in heifers 
    from Umotion data yyyymmddHeiferorg #3 
    v 1.0
    2023/6/29
    @author: inoue
    
"""
def fpymh_rpdumanualHeifer02(wbN, sheetN, wbAIN):
    """
    makes a reproductive data list in heifers 
    from Umotion data yyyymmddHeiferorg #3 

    Parameters
    ----------
    wbN : str 
        workbook name
        "MH_RPDu.xlsx"
    sheetN : str 
        Heifers' data sheet name
        "yyyymmddHeifer01"
    wbAIN : str 
        AIdata workbook name
        "MH_AI_.xlsx"
        
    #sAIN : str 
    #    AI data sheet name
    #    "yyyy"

    Returns
    -------
    None.

    """
 
    #import mh_rpdu
    #import openpyxl
    #import fmstls
    
    wb = openpyxl.load_workbook(wbN)  #MH_RPDu.xlsx
    sheet = wb[sheetN] #yyyymmddHeifer01
    wbAI = openpyxl.load_workbook(wbAIN)  #MH_AI_.xlsx
    #sAI = wb[sAIN]  #yyyy
    
    #鑑定結果を書き換える。
    #
    fpyinput_PTH_s(wbAI, sheet)    
    #print(" sheet" + sheetN + "に鑑定結果を入力しました。")
    
    wb.save(wbN)

#fpymh_rpdumanualCow00#######################################################
"""
    fpymh_rpdumanualCow00 :
    makes a reproductive data list in cows
    from Umotion data yyyymmddCoworg #1
    v 1.0
    2023/7/16
    @author: inoue
    
"""
def fpymh_rpdumanualCow00(wbN, sorgN, sheetN, bDate, VWP):
    """
    makes a reproductive data list in Cows 
    from Umotion data yyyymmddCoworg #1
    
    Parameters
    ----------
    wbN : ワークブック名
        "MH_RPDu.xlsx"
    sorgN : データ参照シート
        "yyyymmddCoworg"
    sheetN : 経産牛データシート
        "yyyymmddCow00"
    bDate : base date 基準日
        "yyyy/mm/dd"
    VWP : int
     volantary waiting period
     50
     
    Returns
    -------
    None.

    """


    #import mh_rpdu
    #import openpyxl
    #import fmstls
    
    wb = openpyxl.load_workbook(wbN)  #MH_RPDu.xlsx
    
    wb.create_sheet(title=sheetN, index=0)
    #make a new sheet 'yyyymmddCow00'
    sheet = wb[sheetN] #yyyymmddCow00
    sorg = wb[sorgN]
    
    #列名だけのsheet sheetNを作成する
    #make a sheet sheetN columns'name only    
    sheet = fpyNewSheetCow_s(sheet)
    print("2." + wbN + " に　sheet " + sheetN + " を作成しました。")
    
    #input data to sheet 'yyyymmdd'
    #sheet 'yyyymmddCow00'にデータ入力する
    sheet = fpyDF_RPDCow_s(sorg, sheet, bDate)
    print("3. Sheet " + sheetN + " に　データを入力しました。")
    
    #make idNo(column 6) 9digits to 10digits'strings
    #個体識別番号(column 6)を10桁文字列に統一する
    sheet = fmstls.fpyidNo_9to10_s( sheet, 6 )
    print("4. 個体識別番号を10桁文字列に統一しました。")
    
    #input days from each last calving to sheet'yyyymmddCow00' column 10
    #分娩後日数を　sheet'yyyymmddCow00' column '分娩後日数' (column 10) に入力する
    sheet = fpyDaysfrmlstclving_s( sheet )
    print("5. sheet" + sheetN + " の 分娩後日数を計算、入力しました。")
    
    # groupe all Cows with the code 'Group' and rewrite column 'Group' and 'Stage'
    # 経産牛全頭を　code'Group'に従ってグループ分けし、column 'Group' and 'Stage'を書き換える。
    sheet = fpygrouoping_s(sheet, VWP)
    print("6. sheet " + sheetN +  " をグルーピングしました")
    
    wb.save(wbN)

#fpymh_rpdumanualCow01#######################################################
"""
    fpymh_rpdumanualCow01 :
    makes a reproductive data list in Cows 
    from Umotion data yyyymmddCoworg #2
    v 1.0
    2023/6/29
    @author: inoue
    
"""
def fpymh_rpdumanualCow01(wbN, sheetN, sheetN_ ):
    """
    makes a reproductive data list in Cows 
    from Umotion data yyyymmddCoworg #2

    Parameters
    ----------
    wbN : str 
        workbook name
        "MH_RPDu.xlsx"
    sheetN : str 
        Cows' data sheet name
        "yyyymmddCow00"
    sheetN_ : str
        new sheet title
        "yyyymmddCow01"
   
    Returns
    -------
    None.

    """
 
    #import mh_rpdu
    #import openpyxl
    #import fmstls
    
    wb = openpyxl.load_workbook(wbN)  #MH_RPDu.xlsx
    sheet = wb[sheetN] #yyyymmddCow00
    #wbAI = openpyxl.load_workbook(wbAIN)  #MH_AI_.xlsx
    #sAI = wb[sAIN]  #yyyy
    
    #初回授精日数を計算する
    #
    sheet = fpydysoffstAI_s( sheet )
    print("9. sheet" + sheetN + " の 初回授精日数を計算、入力しました。")
    
    #授精後日数を再計算する
    #
    sheet = fpyDaysfrmlstAI_s( sheet )
    print("10. sheet" + sheetN + " の 授精後日数を再計算しました。")
    
    #空胎日数を計算する
    #
    sheet = fpyopenDays_s( sheet )
    print("11. sheet" + sheetN + " の 空胎日数を計算、入力しました。")
    
    #鑑定結果を書き込むシートをコピーして作成する
    #
    sheet_ = fmstls.fpysheet_copy_s( wb, sheet, sheetN_ )
    print("12. sheet" + sheetN + "をコピーして、" + sheetN_ + "を作成しました。")
    #sheet_を先頭に移動する
    wb.move_sheet( sheet_, offset=-wb.index(sheet_) )
    #sheet_ 先頭のシートをactiveにする
    wb.active = 0
    
    wb.save(wbN)
    
#fpymh_rpdumanualCow02#######################################################
"""
    fpymh_rpdumanualCow02 :
    makes a reproductive data list in cows 
    from Umotion data yyyymmddCoworg #3 
    v 1.0
    2023/6/29
    @author: inoue
    
"""
def fpymh_rpdumanualCow02(wbN, sheetN, wbAIN ):
    """
    makes a reproductive data list in Cows 
    from Umotion data yyyymmddCoworg #3 

    Parameters
    ----------
    wbN : str 
        workbook name
        "MH_RPDu.xlsx"
    sheetN : str 
        Cows' data sheet name
        "yyyymmddCow01"
    wbAIN : str 
        AIdata workbook name
        "MH_AI_.xlsx"
    
    Returns
    -------
    None.

    """
 
    #import mh_rpdu
    import openpyxl
    #import fmstls
    
    wb = openpyxl.load_workbook(wbN)  #MH_RPDu.xlsx
    sheet = wb[sheetN] #yyyymmddCow00
    wbAI = openpyxl.load_workbook(wbAIN)  #MH_AI_.xlsx
    #sAI = wb[sAIN]  #yyyy
    
    #鑑定結果を書き換える。
    #
    fpyinput_PT_s(wbAI, sheet)    
    print("13. sheet" + sheetN + "に鑑定結果を入力しました。")
    
    wb.save(wbN)    

#fpymh_rpduCowManual#########################################################    
"""
fpymh_rpduCowManual:                        マニュアル
ｖ1.0
2021/9/26
@author: jicc
"""
def fpymh_rpduCowManual():
    
    print('-----mh_rpdu Manual Cow------------------------------------------------------v1.0-------')
    print('1.<>Umotion 経産牛一覧plusyyyymmdd.csv を移行、sheet\"yyyymmddCoworg\"とする')
    print('  sheet\"yyyymmdd_Coworg\"の場合は、授精日、分娩日を確認修正して、_をとる')
    print(' ')
    print('2.fpyNewSheetCow(wbN, sheetN)')
    print('  データ移行用の列名だけのsheet\'yyyymmddCow00\'を作成する')
    print('  PS> python ps_fpynewsheetcow_args.py wbN yyyymmddCow00  #wbN: MH_RPDu.xlsx')
    print(' ')
    print('3.sheet yyyymmddCow00 にデータ入力する')
    print('  PS> python ps_fpydf_rpdcow_args.py wbN yyyymmddCoworg yyyymmddCow00 yyyy/mm/dd')
    print(' ')
    print('4. 9桁耳標を10桁にし、文字列として再入力する')
    print('  PS> python ps_fpyidno_9to10_args.py wbN sheetN col  #col:6')
    print(' ')
    print('5. Group のcodeに従って、グループ分けし、Group列,Stage列を書き換える。')
    print('  PS> python ps_fpygrouping_args.py wbN yyyymmddCow00 VWP  #VWP:50')
    print(' ')
    print('6. <>B列　Group 昇順でソートする')
    print(' ')
    print('7. <>シートをコピー、7繁殖対象外の個体をこのシートに残し、シートyyyymmddCow00から削除する。')
    print(' sheet名:yyyymmddCowout')
    print(' ')
    print('8. 初回授精日数を計算する')
    print('  PS> python ps_fpydysoffstai_args.py wbN yyyymmddCow00')
    print(' ')
    print('9. <>sheet\'yyyymmddCow00\'をコピーし、sheet\'yyyymmddCow01\'を作成する')
    print(' ')
    print('10. 鑑定結果の書き換え')
    print('  PS> python ps_fpyinput_pt_args.py wbAIN yyyy wbRPDN yyyymmddCow01')
    print('  #wbAIN:MH_AI_.xlsx #wbRPDN:MH_RPDu.xlsx')
    print(' ')
    print('11. ..R\MH_RPDu\MH_yyyymmddCow00.csv(utf-8)として、csv保存する。')
    print(' ')
    print('12. ..R\MH_RPDu\MH_yyyymmddCow01.csv(utf-8)として、csv保存する。')
    print('--tools-----------------------------------------------------------------------------------')
    print('空胎日数を計算入力する')
    print('PS> python ps_fpyopendays_args.py wbN yyyymmddCow00')
    print('授精後日数を計算入力する')
    print('PS> python ps_fpydaysfrmlstai_args.py wbN yyyymmdd???00')
    print('---------------------------------------------------------------2021/12/13 by jicc---------')

#fpymh_rpduCowManual00#########################################################    
"""
fpymh_rpduCowManual00:                        マニュアル
ｖ1.0
2023/7/14
@author: jicc
"""
def fpymh_rpduCowManual00():
    
    print('-----mh_rpdu Manual Cow------------------------------------------------------v1.0-------')
    print('1.<>Umotion 経産牛一覧plusyyyymmdd.csv を移行、sheet\"yyyymmddCoworg\"とする')
    print('  sheet\"yyyymmdd_Coworg\"の場合は、授精日、分娩日を確認修正して、_をとる')
    print('*by hand work')
    print(' ')
    print(' ')
    print('#1 fpymh_rpdumanualCow00(wbN, sorgN, sheetN, bDate, VWP)')
    print('  makes a reproductive data list in cows') 
    print(' from Umotion data yyyymmddCoworg #1')
    print('  PS> python ps_fpyrpdumanualcow00_args.py wbN sorgN sheetN bDate VWP')
    print('  wbN:MH_RPDu.xlsx, sorgN:yyyymmddCoworg, sheetN:yyyymmddCow00 ')
    print('  bDate:yyyy/mm/dd, VWP:50 ')
    print(' ')
    print('  2.fpyNewSheetCow(wbN, sheetN)')
    print('  データ移行用の列名だけのsheet\'yyyymmddCow00\'を作成する')
    print('  PS> python ps_fpynewsheetcow_args.py wbN yyyymmddCow00  #wbN: MH_RPDu.xlsx')
    print(' ')
    print('  3.sheet yyyymmddCow00 にデータ入力する')
    print('  PS> python ps_fpydf_rpdcow_args.py wbN yyyymmddCoworg yyyymmddCow00 yyyy/mm/dd')
    print(' ')
    print('  4. 9桁耳標を10桁にし、文字列として再入力する')
    print('  PS> python ps_fpyidno_9to10_args.py wbN sheetN col  #col:6')
    print(' ')
    print('  5. 分娩後日数を再計算する')
    print('  PS> python ps_fpydaysfrmlstclving_args.py wbN yyyymmddCow00') 
    print(' ')
    print('  6. Group のcodeに従って、グループ分けし、Group列,Stage列を書き換える。')
    print('  PS> python ps_fpygrouping_args.py wbN yyyymmddCow00 VWP  #VWP:50')
    print(' ')
    print(' ')
    print('7. <>B列　Group 昇順でソートする')
    print('*by hand work')
    print(' ')
    print('8. <>シートをコピー、7繁殖対象外の個体をこのシートに残し、シートyyyymmddCow00から削除する。')
    print(' sheet名:yyyymmddCowout')
    print('*by hand work')
    print(' ')
    print('#2 fpymh_rpdumanualCow01(wbN, sheetN, sheetN_)')
    print('  makes a reproductive data list in cows') 
    print(' from Umotion data yyyymmddCoworg #2')
    print('  PS> python ps_fpyrpdumanualcow01_args.py wbN sheetN sheetN_')
    print('  wbN:MH_RPDu.xlsx, sheetN:yyyymmddCow00, sheetN_:yyyymmddCow01 ')
    print(' ')
    print('  9. 初回授精日数を計算する')
    print('  PS> python ps_fpydysoffstai_args.py wbN yyyymmddCow00')
    print(' ')
    print('  10.授精後日数を計算入力する')
    print('  PS> python ps_fpydaysfrmlstai_args.py wbN yyyymmdd???00')
    print(' ')
    print('  11.空胎日数を計算入力する')
    print('  PS> python ps_fpyopendays_args.py wbN yyyymmddCow00')
    print(' ')
    print('  12. <>sheet\'yyyymmddCow00\'をコピーし、sheet\'yyyymmddCow01\'を作成する')
    print(' ')
    print('#3 fpymh_rpdumanualCow02(wbN, sheetN, wbAIN)')
    print('  makes a reproductive data list in heifers') 
    print(' from Umotion data yyyymmddCoworg #3')
    print('  PS> python ps_fpyrpdumanualCow02_args.py wbN sheetN wbAIN')
    print('  wbN:MH_RPDu.xlsx, sheetN:yyyymmddCow01, wbAIN:MH_AI_.xlsx')
    print('  13. 鑑定結果の書き換え')
    print('     PS> python ps_fpyinput_pt_args.py wbAIN yyyy wbRPDN yyyymmddCow01')
    print('     #wbAIN:MH_AI_.xlsx #wbRPDN:MH_RPDu.xlsx')
    print(' ')
    print('14. ..RPDrpt\csv\MH_yyyymmddCow00.csv(utf-8)として、csv保存する。')
    print('*by hand work')
    print(' ')
    print('15. ..RPDrpt\csv\MH_yyyymmddCow01.csv(utf-8)として、csv保存する。')
    print('*by hand work')
    print('---------------------------------------------------------------2023/7/16 by jicc---------')
    
#fpymh_rpduHeiferManual#####################################################    
def fpymh_rpduHeiferManual():
    
    print('-----mh_rpdu Manual Heifer-----------------------------------------------------v1.02-------')
    print('1. <>Umotion 未経産牛一覧yyyymmdd.csv　を移行、sheetyyyymmddHeiferorgとする')
    print('  sheet\"yyyymmdd_Heiferorg\"の場合は、授精日を確認修正して、_をとる')
    print(' ')
    print('2.fpyNewSheetHeifer(wbN, sheetN)')
    print('  データ移行用の列名だけのsheet\'yyyymmddHeifer00\'を作成する')
    print('  PS> python ps_fpynewsheetheifer_args.py wbN yyyymmddHeifer00 wbN:MH_RPDu.xlsx')
    print(' ')
    print('3.sheet yyyymmddHeifer00 にデータ入力する')
    print('  PS> python ps_fpydf_rpdheifer_args.py wbN yyyymmddHeiferorg yyyymmddHeifer00 yyyy/mm/dd')
    print(' ')
    print('4. 9桁耳標を10桁にし、文字列として再入力する')
    print('  PS> python ps_fpyidno_9to10_args.py wbN sheetN col  ,#col:6') 
    print(' ')
    print('5. 未経産用 yyyymmddHeifer00 分娩後日数列に月齢を入力する')
    print('  PS> python ps_fpyageinmnths_args.py wbN yyyymmddHeifer00') 
    print(' ')
    print('6. Group のcodeに従って、グループ分けし、Group列,Stage列を書き換える。')
    print('  PS> python ps_fpygroupingh_args.py wbN yyyymmddHeifer00 VWPm VWPM')
    print('  #VWPm:12 #VWPM:13')
    print(' ')
    print('7.<> B列　Group 昇順でソートする')
    print(' ')
    print('8. <>シートをコピー、9預託の個体をこのシートに残し、シートyyyymmddHeifer00から削除する。')
    print(' sheet名:yyyymmddHeiferout')
    print(' ')
    print('9. 初回授精月数を計算する')
    print('  PS> python ps_fpymnthsoffstai_args.py wbN yyyymmddHeifer00')
    print(' ')
    print('10. 空胎月数を計算する')
    print('  PS> python ps_fpyopenmnths_args.py wbN yyyymmddHeifer00')
    print(' ')
    print('11. <>sheet\'yyyymmddHeifer00\'をコピーし、sheet\'yyyymmddHeifer01\'を作成する')
    print(' ')
    print('12. 鑑定結果の書き換え')
    print('  PS> python ps_fpyinput_pth_args.py wbAIN yyyy wbRPDN yyyymmddHeifer01')
    print('  #wbAIN:MH_AI_.xlsx #wbRPDN:MH_RPDu.xlsx')
    print(' ')
    print('13. ..R\MH_RPDu\MH_yyyymmddHeifer00.csv(utf-8)として、csv保存する。')
    print(' ')
    print('14. ..R\MH_RPDu\MH_yyyymmddHeifer01.csv(utf-8)として、csv保存する。')
    print('--tools-----------------------------------------------------------------------------------')
    print('日齢を計算入力する')
    print('PS>python ps_fpyageindays_args.py wbN sheetN')
    print('授精後日数を計算入力する')
    print('PS> python ps_fpydaysfrmlstai_args.py wbN yyyymmdd???00')
    print('---------------------------------------------------------------2021/12/13 by jicc---------')
    
#fpymh_rpduHeiferManual00#####################################################
def fpymh_rpduHeiferManual00():
    
    print('-----mh_rpdu Manual Heifer00---------------------------------------------------v1.00-------')
    print('1. <>Umotion 未経産牛一覧yyyymmdd.csv　を移行、sheetyyyymmddHeiferorgとする')
    print('  sheet\"yyyymmdd_Heiferorg\"の場合は、授精日を確認修正して、_をとる')
    print('*by hand work')
    print(' ')
    print(' ')
    print('#1 fpymh_rpdumanualHeifer00(wbN, sorgN, sheetN, bDate)')
    print('  makes a reproductive data list in heifers') 
    print(' from Umotion data yyyymmddHeiferorg #1')
    print('  PS> python ps_fpyrpdumanualheifer00_args.py wbN sorgN sheetN bDate VWPm VWPM')
    print('  wbN:MH_RPDu.xlsx, sorgN:yyyymmddHeiferorg, sheetN:yyyymmddHeifer00 ')
    print('  bDate:yyyy/mm/dd, VWPm:13, VWPM:14 ')
    print(' ')
    print('  2.fpyNewSheetHeifer(wbN, sheetN)')
    print('  データ移行用の列名だけのsheet\'yyyymmddHeifer00\'を作成する')
    print('  PS> python ps_fpynewsheetheifer_args.py wbN yyyymmddHeifer00 wbN:MH_RPDu.xlsx')
    print(' ')
    print('  3.sheet yyyymmddHeifer00 にデータ入力する')
    print('  PS> python ps_fpydf_rpdheifer_args.py wbN yyyymmddHeiferorg yyyymmddHeifer00 yyyy/mm/dd')
    print(' ')
    print('  4. 9桁耳標を10桁にし、文字列として再入力する')
    print('  PS> python ps_fpyidno_9to10_args.py wbN sheetN col  ,#col:6') 
    print(' ')
    print('  5. 未経産用 yyyymmddHeifer00 分娩後日数列に月齢を入力する')
    print('  PS> python ps_fpyageinmnths_args.py wbN yyyymmddHeifer00') 
    print(' ')
    print('  6. Group のcodeに従って、グループ分けし、Group列,Stage列を書き換える。')
    print('  PS> python ps_fpygroupingh_args.py wbN yyyymmddHeifer00 VWPm VWPM')
    print('  #VWPm:12 #VWPM:13')
    print(' ')
    print(' ')
    print('7.<> B列　Group 昇順でソートする')
    print('*by hand work')
    print(' ')
    print('8. <>シートをコピー、9預託の個体をこのシートに残し、シートyyyymmddHeifer00から削除する。')
    print(' sheet名:yyyymmddHeiferout')
    print('*by hand work')
    print(' ')
    print(' ')
    print('#2 fpymh_rpdumanualHeifer01(wbN, sheetN, sheetN_)')
    print('  makes a reproductive data list in heifers') 
    print(' from Umotion data yyyymmddHeiferorg #2')
    print('  PS> python ps_fpyrpdumanualheifer01_args.py wbN sheetN sheetN_')
    print('  wbN:MH_RPDu.xlsx, sheetN:yyyymmddHeifer00, sheetN_:yyyymmddHeifer01 ')
    print(' ')
    print('  9. 初回授精月数を計算する')
    print('  PS> python ps_fpymnthsoffstai_args.py wbN yyyymmddHeifer00')
    print(' ')
    print('  10. 空胎月数を計算する')
    print('  PS> python ps_fpyopenmnths_args.py wbN yyyymmddHeifer00')
    print(' ')
    print('  11. 日齢を計算入力する')
    print('  PS>python ps_fpyageindays_args.py wbN sheetN')
    print('  12. 授精後日数を計算入力する')
    print('  PS> python ps_fpydaysfrmlstai_args.py wbN yyyymmdd???00')
    print('  13. <>sheet\'yyyymmddHeifer00\'をコピーし、sheet\'yyyymmddHeifer01\'を作成する')
    print(' ')
    print(' ')
    print('#3 fpymh_rpdumanualHeifer02(wbN, sheetN, wbAIN)')
    print('  makes a reproductive data list in heifers') 
    print(' from Umotion data yyyymmddHeiferorg #3')
    print('  PS> python ps_fpyrpdumanualheifer02_args.py wbN sheetN wbAIN')
    print('  wbN:MH_RPDu.xlsx, sheetN:yyyymmddHeifer01, wbAIN:MH_AI_.xlsx')
    print(' ')
    print('  14. 鑑定結果の書き換え')
    print('  PS> python ps_fpyinput_pth_args.py wbAIN yyyy wbRPDN yyyymmddHeifer01')
    print('  #wbAIN:MH_AI_.xlsx #wbRPDN:MH_RPDu.xlsx')
    print(' ')
    print(' ')
    print('15. ..RPDrpt\csv\MH_yyyymmddHeifer00.csv(utf-8)として、csv保存する。')
    print('*by hand work')
    print(' ')
    print('16. ..RPDrpt\csv\MH_yyyymmddHeifer01.csv(utf-8)として、csv保存する。')
    print('*by hand work')
    print(' ')   
    print('---------------------------------------------------------------2023/7/14 by jicc---------')