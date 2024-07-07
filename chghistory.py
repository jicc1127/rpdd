# -*- coding: utf-8 -*-
import os, re     
import tabula
import openpyxl
import csv
import shutil
import datetime
####################################################from fmstls.py##########
"""
fpyopenxl(wbN, sheetN):
    Excelfile wbN.xlsx　sheet sheetN Open 
    v1.00
    2022/1/5
    @author: jicc
"""
def fpyopenxl(wbN, sheetN):
    """
    Excelfile wbN.xlsx　sheet sheetN Open

    Parameters
    ----------
    wbN : str
        ExcelFile Name   ex.MH_CowHistory.xlsx
    sheetN : str
        sheet name

    Returns
    -------
    None.

    """
    
    #import openpyxl
    
    wb = openpyxl.load_workbook(wbN)
    sheet = wb[sheetN]
    return [wb, sheet]
"""
fpyopencsv_robj:
    csvfile Open for Reader object
    v1.00
    2022/1/5
    @author: jicc
"""
def fpyopencsv_robj(csvN):
    """
    csvfile Open for Reader object

    Parameters
    ----------
    csvN : str
        csvFile Name   ex.MH_??_History.csv

    Returns
    -------
    None.

    """
    #import csv
    
    #filename = csvN.split('.')
    #filename = filename[0]  #拡張子を削除したfilename
    
    filename_file = open(csvN)     #csvfile open
    filename_reader = csv.reader(filename_file)       #get Reader object
    
    
    return filename_reader

#fpyopencsv_rdata    
"""
fpyopencsv_rdata:
    csvfile Open for Reader data
    v1.00
    2022/1/5
    @author: jicc
"""
def fpyopencsv_rdata(csvN):
    """
    csvfile Open for Reader data

    Parameters
    ----------
    csvN : str
        csvFile Name   ex.MH_??_History.csv

    Returns
    -------
    filename_data : list's list 

    """
    #import csv
    
    #filename = csvN.split('.')
    #filename = filename[0]  #拡張子を削除したfilename
    
    filename_file = open(csvN)     #csvfile open
    filename_reader = csv.reader(filename_file)       #get Reader object
    filename_data = list(filename_reader)             #list's list
    
    return filename_data
    
"""
fpyopencsv_w:
    csvfile Open for Writer
    v1.00
    2022/1/7
    @author: jicc
"""
def fpyopencsv_w(csvN):
    """
    csvfile Open for Writer

    Parameters
    ----------
    csvN : str
        csvFile Name   ex.MH_??_History.csv

    Returns
    -------
    None.

    """
    #import csv
    
    output_file = open(csvN, 'w', newline='')       #csvfile open
    output_writer = csv.writer(output_file)       #get Reader object
     
    return output_writer

"""
fpygetCell_value: get value from the target Cell
v1.00
2022/2/4

@author: inoue
"""
def fpygetCell_value(sheet, r, col):
    """
    get value from the target Cell on an Excelsheet

    Parameters
    ----------
    sheet : worksheet
        sheetBLV
    r : int
        row
    col : int
        column

    Returns
    -------
    value

    """

    value = sheet.cell(row=r, column=col).value
    return value

"""
fpyinputCell_value: input value to the target Cell
v1.00
2022/2/4

@author: inoue
"""
def fpyinputCell_value(sheet, r, col, vl):
    """
    input value to the target Cell

    Parameters
    ----------
    sheet : worksheet
        sheetBLV
    r : int
        row
    col : int
        column
    vl : type of value
    
    Returns
    -------
    None.

    """

    sheet.cell(row=r, column=col).value = vl 

#fpyNewSheet################################################################
"""
fpyNewSheet : Excelbookに
sheet　'columns'と同じ sheet　'scolN'を作成する。
ｖ1.01
2022/5/3

@author: jicc

"""
def fpyNewSheet(wbN, sheetN, scolN, r):
    """
    Excelbookに sheet 'scolN' r行目の'columns'を1行目に配置した sheet'sheetN'を作成する。
    *sheet 'columns'(列名を記入したシート) を作成しておく
    Parameters
    ----------
    wbN : 　str          
        sheetを作成するワークブック
    sheetN : str　　　　　　シート名:"????" 
        作成するシート
    scolN : str         シート名: "columns"
        参照するシート
	r : int		r行目 作成するcolumn行
    Returns
    -------
    None.

    """
    #import openpyxl
    
    wb = openpyxl.load_workbook(wbN)
    #sheetN = wb[sheetN]
    wb.create_sheet(title=sheetN, index=0)
    sheet = wb[sheetN]
    scol = wb[scolN]
    
    maxcol = scol.max_column #sheet columnの最終列
    
    for i in range(1,maxcol+1):
        sheet.cell(row=r, column=i).value = scol.cell(row=1, column=i).value
    
     
    wb.save(wbN)

"""
fpychgSheetTitle      :change ExcelSheet's title
v1.0
2022/3/30

@author: inoue
"""
def fpychgSheetTitle(wbN, sheetN, sheetN1):
    """
    change the sheet's title

    Parameters
    ----------
    wbN : str
        Excelfile to check double data  '??_CowsHistory.xlsx'
    sheetN : str
        元のシート名  : 'KTFarm'
    sheetN1 : str
        変更名      : 'KTFarmorg' 

    Returns
    -------
    None.

    """
    #import chghistory
    wbobj = fpyopenxl(wbN, sheetN)
    wb = wbobj[0]
    sheet = wbobj[1]
    sheet.title = sheetN1
    wb.save(wbN)
 
####################################################from fmstls.py##########

#fpymkd_path##############################################################
"""
fpymkd_path : 
    make a directory  at current directory
    #カレントディレクトリに　path名のディレクトリが存在しなければ作成する
    v1.0
    2023/9/2
    @author: jicc
"""
def fpymkd_path( path ):
    '''
    make a directory  at current directory

    Parameters
    ----------
    path : str
        directory name 　ex. '.\csvhistory'
    Returns
    -------
    None.

    '''
    import os
    
    #try:
    
    if os.path.exists( path ):
        print( "既に存在するディレクトリです。")
        #continue
    else:
            os.makedirs( path )
        
    #except FileExistsError:
    #   print( "既に存在するディレクトリです。")

#fpymkxlsheet#############################################################
"""
fpymkxlsheet : make an ExcelSheet if it dose not exist
v1.0
2023/9/30
#* sheet.cell(row=r,..) -> sheet.cell(row=1,..)
#  scol.cell(row=1,..) -> scol.cell(row=r,..) に訂正　#*
v1.01
2024/3/1
@author: inoue
"""
def fpymkxlsheet(wbN, sheetN, scolN, r):
    """
    make an ExcelSheet if it dose not exist

    Parameters
    ----------
    wbN : str
        Excelfile to make a new sheet
    sheetN : str
        new sheet name  : 'ABFarmout'
    scolN : str         シート名: "columns"
        参照するシート
	r : int		r行目 作成するcolumn行
    
    Returns
    -------
    None.

    """
    #import openpyxl
    
    wb = openpyxl.load_workbook(wbN)
    #sheet = wbobj[1]
    #sheet.title = sheetN1
    snames = []
    snames = wb.sheetnames #get_sheet_names()
    print(snames)
    
    if sheetN not in snames:
        wb.create_sheet(title=sheetN, index=0)
        sheet = wb[sheetN]
        scol = wb[scolN]
        
        maxcol = scol.max_column #sheet columnの最終列
                
        for i in range(1,maxcol+1): 
            sheet.cell(row=1, column=i).value = scol.cell(row=r, column=i).value #*
        print("Sheet " +sheetN + " を作成しました。")
    else:
        print("Sheet " +sheetN + " exists")
     
    wb.save(wbN)

#fpymkxlsheet_#############################################################
"""
fpymkxlsheet_ : make an ExcelSheet if it dose not exist
    workbook version
v1.0
2024/2/18
#* sheet.cell(row=r,..) -> sheet.cell(row=1,..)
#  scol.cell(row=1,..) -> scol.cell(row=r,..) に訂正　#*
v1.01
2024/3/1

@author: inoue
"""
def fpymkxlsheet_(wb, sheetN, scolN, r):
    """
    make an ExcelSheet if it dose not exist
    workbook version
    Parameters
    ----------
    wb : workbook.workbook.Workbook
         workbook object
    sheetN : str
        new sheet name  : 'ABFarmout'
    scolN : str         シート名: "columns"
        参照するシート
	r : int		r行目 作成するcolumn行
    
    Returns
    -------
    sheet : worksheet object

    """
    #import openpyxl
    
    #wb = openpyxl.load_workbook(wbN)
    #sheet = wbobj[1]
    #sheet.title = sheetN1
    snames = []
    snames = wb.sheetnames #get_sheet_names()
    print(snames)
    
    if sheetN not in snames:
        wb.create_sheet(title=sheetN, index=0)
        sheet = wb[sheetN]
        scol = wb[scolN]
        
        maxcol = scol.max_column #sheet columnの最終列
                
        for i in range(1,maxcol+1):
            sheet.cell(row=1, column=i).value = scol.cell(row=r, column=i).value #*
        print("Sheet " +sheetN + " を作成しました。")
    else:
        sheet = wb[sheetN]
        print("Sheet " +sheetN + " exists")
     
    return sheet

#fpypdf_to_csv############################################################
"""
fpypdf_to_csv
    convert ****.pdf to ****.csv file
v1.00
2022/1/1
@author: inoue
"""

#import tabula

def fpypdf_to_csv(filename, Path):
    '''
    filename: string
    csv変換するpdf  filename

    Returns
    -------
    filename.csv

    '''
    filename_pdf = Path + "\\" +  filename + ".pdf"
    filename_csv = Path + "\\" + filename + ".csv"
    tabula.convert_into(filename_pdf, filename_csv, 
                    stream=True , output_format="csv", pages="all")


"""
fpySpdf_in_Dir  -  ディレクトリ内の特定の拡張子を持つファイルを見つけ
                  file名をプリントする
v1.00
2022/1/1
by jicc
"""
def fpySpdf_in_Dir(Ext, Path):
    
    #import os, re
    
    fs = os.listdir(Path) 
    #Pathに指定したフォルダー内の、ファイル名とフォルダー名のリストを返す
    regex_ext = re.compile(Ext)   #Regex(regular expression)オブジェクトを返す
    #print(regex_ext)
    
    for f in fs:
        #print(f)
        mo = regex_ext.search(f) #regx_extにマッチするとmatdhオブジェクトを返す
        if mo: #!=None
            print(f)

######################################################未使用
"""
fpySpdf_in_Dir_to_csv  -  ディレクトリ内の特定の拡張子(.pdf)を持つファイルを見つけ
                  csvfile に変換する
                  変換したpdffileを　フォルダーpdforgに移動する
v1.01
2022/1/11
by jicc

"""
def fpySpdf_in_Dir_to_csv(Ext, Path, bckPath):
    """
    Parameters
    ----------
    Ext : str
        拡張子　　　'\.pdf'　　
    Path : str
        path      '.\\' カレントディレクトリ
    bckPath : str
        file移動するフォルダーのpath

    Returns
    -------
    None.

    """
    
    #import os, re
    #import chghistory
    #import shutil
    fs = os.listdir(Path)
    regex_ext = re.compile(Ext)
    #print(regex_ext)
    
    for f in fs:
        #print(f)
        mo = regex_ext.search(f)
        if mo:
            print(f)
            filename = f.split('.')
            filename = filename[0]
            fpypdf_to_csv(filename, Path)
            #print(mo.group())
            filename_pdf = filename + '.pdf'
            shutil.move(filename_pdf, bckPath)
            
#fpyCowHistory##############################################################            
"""
fpyCowHistory
    牛の個体情報.csvから、CowHistory.csv(changehistory's list )を作成する
    'No'を行頭にいどうすることを中止、単純に個体データと異動データを結合するように変更
    個体識別番号　９桁->１０桁
    日付データ　yyyy.mm.dd -> yyyy/mm/dd の処置を追加
ｖ1.02
2022/1/9
@author: inoue
"""

def fpyCowHistory(csvorgN, csvoutN):
    '''
    牛の個体情報.csvから、CowHistory.csv(changehistory's list )を作成する

    Parameters

    ----------
    csvorgN : str
        もととなるcsvファイル        MH_???_yyyymmdd.csv
    csvoutN : str
        作成するcsvファイル　　　　　　MH_???_yyyymmddH.csv　

    Returns
    -------
    None.

    '''
    #import csv
    #import chghistory
    
    mhcow_file = open(csvorgN)     
    ################################################################
    #csvfile open , encoding="utf-8",  "shift-jis"
    #UnicodeDecodeError: 'cp932' ... のエラーのためencoding="utf-8"を追加でもダメ2022/1/11
    #PS> ps_fpyymd_csvtocowshistory_csv_args.py .csv .\ .\csvorg で実施の時のみ。
    #アナコンダ　インタラクティブシェルで行ったら問題なし。
    #今日試したらError出なかった。　なぜ？？　2022/1/12
    #################################################################
    mhcow_reader = csv.reader(mhcow_file)       #get Reader object
    mhcow_data = list(mhcow_reader)             #list's list
    
    cowhistory_header = mhcow_data[0]
    print(cowhistory_header)
    cowhistory_header = cowhistory_header \
        + ['No', '異動内容', '異動年月日', '飼養施設所在地', '氏名または名称'] #見出し行のリスト
     
    output_file = open(csvoutN, 'w', newline='')
    output_writer = csv.writer(output_file)
    output_writer.writerow(cowhistory_header)
    row_max = mhcow_reader.line_num  # =len(mhcow_data) リストの行数
    
    id_info = mhcow_data[1]  	
    #['個体識別番号', '出生の年月日', '雌雄の別', '母牛の個体識別番号', '種別']
    #1行目の絶対データ
    #print(id_info)
    #id_info_ = id_info[0]
    id_info[0] = fpycsvidNo_9to10( id_info[0] ) #idNo
    id_info[1] = fpydate_dottoslash( id_info[1]) #出生年月日 'yyyy.mm.dd' -> 'yyyy/mm/dd'
    id_info[3] = fpycsvidNo_9to10( id_info[3] ) #damidNo
    print(id_info)
    for row_num in range(5, row_max):
  
        history = mhcow_data[row_num] 	
        #['No', '異動内容', '異動年月日', '住所', '氏名または名称']
        #row_num行目の相対データ
        #print(history)
        history[2] = fpydate_dottoslash( history[2]) #異動年月日 'yyyy.mm.dd' -> 'yyyy/mm/dd'
        print(history)
        id_info_history = id_info + history #行データを結合
        output_writer.writerow(id_info_history)
    
    output_file.close()

#fpycsvlisttoxls############################################################            
"""
fpycsvlisttoxls: 
    csvfileのデータをexcelfileに移行する
    死亡のテーブルを回避する処置を加えた
    ｖ1.01
    2022/1/13
    #1) 3,9,12列の年月日をdatetimeに変換　を追加
    v1.02
    2023/9/26
    @author: jicc
    
"""
def fpycsvlisttoxls(csvN, wbN, sheetN):
    """
    csvfileのデータをexcelfileに移行する

    Parameters
    ----------
    csvN : str
        original csvfile  'MH_???_History.csv', '0123456789_yyyymmdd.csv'
    wbN : str
        Excelfile to move History data  'MH_CowsHistory.xlsx'
    sheetN : str
        sheet name to add data   'MHFarm' 

    Returns
    -------
    None.

    """
    import chghistory
    #import openpyxl
    
    wbobj = chghistory.fpyopenxl(wbN, sheetN)   #get Worksheet object
    wb = wbobj[0]
    sheet = wbobj[1]
    #wb = openpyxl.load_workbook(wbN)
    #sheet = wb[sheetN]
    max_row = sheet.max_row                     
    
    csvdata = chghistory.fpyopencsv_rdata(csvN)     #list's list of the csvfile
    ln = len(csvdata)       #the length of the list csvdata
    ln_ = len(csvdata[0])   #the number of the csvdata's list[0]
    k = max_row
    for i in range(1, ln):
        No = csvdata[i][5]  #死亡のとき、'異動内容「死亡」の' 以下のデータを削除のため
        #* "1", "2", "異動内容「死亡」の"　strlength <=2 で振り分け　2022/1/13 ｖ1.01
        No_ln = len(No)      
        while (No_ln<=2):       #* #No==99まで可能
            for j in range(0, ln_+1):
                if j== 0:   #input the LineNo
                    sheet.cell(row=max_row+i, column=j+1).value = k
                    #the first LineNo is k (row k+1)
                    #最初のLineNoは　k　(k+1行)に等しい
                    #print(k)
                    k = k + 1
                else:
                    print(csvdata[i][j-1])
                    sheet.cell(row=max_row+i, column=j+1).value = \
                        csvdata[i][j-1]
                        
                        #l = csvdata[i][j-1]
                        #print(l)
            break #*
            
    #sheet 3列 '出生の年月日' 'yyyy/mm/dd' -> datetimeに変換  #1)
    chghistory.fpyxlstrymdtodatetime_s( sheet, 3 )
    print('sheet 3列 \'出生の年月日\' \'yyyy/mm/dd\' -> datetimeに変換')
    
    #sheet 9列 '異動年月日' 'yyyy/mm/dd' -> datetimeに変換  #1)
    chghistory.fpyxlstrymdtodatetime_s( sheet, 9 )
    print('sheet 9列 \'異動年月日\' \'yyyy/mm/dd\' -> datetimeに変換')
    
    #sheet 12列 '検索年月日' 'yyyy/mm/dd' -> datetimeに変換 #1)
    chghistory.fpyxlstrymdtodatetime_s( sheet, 12 )
    print('sheet 12列 \'検索年月日\' \'yyyy/mm/dd\' -> datetimeに変換')
        
    wb.save(wbN)

#fpycsvidNo_9to10###########################################################    
"""
fpycsvidNo_9to10:
    idNo in a csvfile 9figures to 10figures
    v1.00
    2022/1/9
    @author: inoue
    
"""
def fpycsvidNo_9to10( idNo ):
    """
    idNo in a csvfile 9figures to 10figures

    Parameters
    ----------
    idNo : str
        idNo

    Returns
    -------
    None.

    """
    if len(idNo) == 9:
        idNo = '0' + idNo 
    else:
        idNo = idNo 
    
    return idNo 

#fpydate_dottoslash#########################################################
"""
fpydate_dottoslash:
    date in a csvfile 'yyyy.mm.dd' to 'yyyy/mm/dd'
    v1.02
    2022/7/6
    @author: inoue
    
"""
def fpydate_dottoslash( date ):
    """
    date in a csvfile 'yyyy.mm.dd' to 'yyyy/mm/dd'

    Parameters
    ----------
    date : str
        date

    Returns
    -------
    date  : datetime 

    """
    date = date.split('.')
    date = "/".join(date)   #*
    #date = datetime.datetime.strptime(date, '%Y/%m/%d')
    #date(str)をdatetimeに変換　ｖ1.01　2022/3/1
    #date = date.strftime('%Y/%m/%d')
    #'yyyy/mm/dd'に変換 v1.02 2022/7/6 #*の状態と同じ解消する2022/7/7
    return date
 
"""
fpyymd_csvtoCowsHistory_csv:
    フォルダー内の個体履歴org(csv)をCowsHistory.csvに変更する
    変更後orgcsvfile を　別フォルダー(./csvorg)に移動する
    v1.01
    2022/1/10
    @author: jicc
"""
def fpyymd_csvtoCowsHistory_csv(Ext, Path, bckPath):
    """
    Parameters
    ----------
    Ext : str
        拡張子　　　'\.csv'　　
    Path : str
        path      '.\\' カレントディレクトリ
    bckPath : str
        file移動するフォルダーのpath　'.\\csvorg' 

    Returns
    -------
    None.

    """
    
    #import os, re
    #import shutil
    #import chghistory
    fs = os.listdir(Path)
    regex_ext = re.compile(Ext)
    #print(regex_ext)
        
    for f in fs:
        #print(f)
        mo = regex_ext.search(f)
        if mo:
            print(f)
            f_ = f.split('.')
            csvoutN = f_[0] + 'H.csv'
            csvorgN = f
            fpyCowHistory(csvorgN, csvoutN)
            
            try:
                shutil.move(csvorgN, bckPath)
                #csvoriginalfile(csvodgN) を　フォルダーbckPathに移動
            except shutil.Error:
                print( csvorgN + ' already exists') 
     
            
#fpyHistory_csvto_xlsx            
"""
fpyHistory_csvto_xlsx:
    フォルダー内の???H.csv)を???CowsHistory.xlsxに移動する
    移動後???H.csvを　別フォルダー(./csvhistory)に移動する
    v1.02
    2022/7/5
    @author: jicc
"""
def fpyHistory_csvto_xlsx(Ext, Path, bckPath, wbN, sheetN):
    """
    フォルダー内の???H.csv)を???CowsHistory.xlsxに移動する
    移動後???H.csvを　別フォルダー(./csvhistory)に移動する
    add try~except v1.02 2022/7/5
    Parameters
    ----------
    Ext : str
        拡張子　　　'\.csv'　　
    Path : str
        path      '.\\' カレントディレクトリ
    bckPath : str
        file移動するフォルダーのpath
    wbN : str
        Excelfile to move History data  'AB_CowsHistory.xlsx'
    sheetN : str
        sheet name to add data   'ABFarm' 

    Returns
    -------
    None.

    """
    
    import os, re
    import shutil
    #import chghistory
    fs = os.listdir(Path)
    #Path(.\\ カレントディレクトリ)に含まれるファイル名とフォルダ名のリスト
    regex_ext = re.compile(Ext)
    #Ext の　Regex(regular expression)オブジェクト
    
    #print(regex_ext)
        
    for f in fs:
        #print(f)
        mo = regex_ext.search(f)
        if mo:
            print(f)
            fpycsvlisttoxls(f, wbN, sheetN)
            
            try:
                shutil.move(f, bckPath)
            except shutil.Error:
                print( f + ' already exists')
            #csvoriginalfile(csvodgN) を　フォルダーbckPathに移動
            #上書きできないので例外処理
            #[ ]f(1).csvなどファイル名を変えて保存するなどに変更か

#fpystrtodatetime##########################################################
"""
fpystrtodatetime : str'yyyy/mm/dd'をdatetime に変換する

v1.00
2022/3/1

@author: inoue
"""
def fpystrtodatetime( date ):
    """
    str'yyyy/mm/dd'をdatetime に変換する

    Parameters
    ----------
    date : str
       'yyyy/mm/dd'

    Returns
    -------
    date  : datetime

    """
    #import datetime
    date = datetime.datetime.strptime( date, '%Y/%m/%d')
    
    return date
#fpyxlstrymdtodatetime######################################################
"""
fpyxlstrymdtodatetime : Excel cell 'yyyy/mm/dd'をdatetimeに変換する

v1.01
2022/3/2

@author: inoue
"""
def fpyxlstrymdtodatetime(wbN, sheetN, col):
    """
    Excel cell 'yyyy/mm/dd'をdatetimeに変換する

    Parameters
    ----------
    wbN : str
        書き換えするExcelFile名   :??_CowsHistory.xlsx
    sheetN : str
        書き換えするシート名　　　　：??Farm
    col : int
        書き換えする列

    Returns
    -------
    None.

    """
    #import fmstls 
    #import chghistory
    #import datetime
    
    xl = []
    xl = fpyopenxl(wbN, sheetN)
    wb = xl[0] #workbook
    sheet = xl[1] #worksheet
    
    for i in range(2, sheet.max_row+1):
        
        date = fpygetCell_value(sheet, i, col)
        if type(date) == str: #date = 'str'の場合datetimtに変換1.01
        #if type(date) != datetime.datetime: #これではNoneセルでstopする
            date = fpystrtodatetime( date )
            fpyinputCell_value(sheet, i, col, date)
        else:
            continue
            
        
        
    wb.save(wbN)


#fpyxlstrymdtodatetime_s####################################################
"""
fpyxlstrymdtodatetime_s: Excel cell 'yyyy/mm/dd'をdatetimeに変換する
    sheet version

v1.00
2023/9/12

@author: inoue
"""
def fpyxlstrymdtodatetime_s( sheet, col ):
    """
    Excel cell 'yyyy/mm/dd'をdatetimeに変換する

    Parameters
    ----------
    sheet : worksheet.worksheet.Worksheet
         worksheet object
    col : int
        書き換えする列

    Returns
    -------
    sheet : worksheet.worksheet.Worksheet
         worksheet object

    """
        
    for i in range(2, sheet.max_row+1):
        
        date = fpygetCell_value(sheet, i, col)
        if type(date) == str: #date = 'str'の場合datetimtに変換1.01
        #if type(date) != datetime.datetime: #これではNoneセルでstopする
            date = fpystrtodatetime( date )
            fpyinputCell_value(sheet, i, col, date)
        else:
            continue
            
    return sheet
        
#fpyxllist_to_list#########################################################
"""
fpyxllist_to_list: 
    excelfileのリストを　lists'　list にする
    
    ｖ1.00
    2022/3/9
    @author: jicc
    
"""
def fpyxllist_to_list(wbN, sheetN, ncol):
    """
    excelfileのデータをlists'listにする

    Parameters
    ----------
    wbN : str
        Excelfile to move History data  '??_CowsHistory.xlsx'
    sheetN : str
        sheet name to add data   '??Farm' 
    ncol :  int
        number of columns
    Returns
    -------
    xllists : lists' list

    """
    #import chghistory
    #import openpyxl
    
    wbobj = fpyopenxl(wbN, sheetN)   #get Worksheet object
    #wb = wbobj[0]
    sheet = wbobj[1]
    #wb = openpyxl.load_workbook(wbN)
    #sheet = wb[sheetN]
    max_row = sheet.max_row
    # max_col = sheet.max_col
    #AttributeError: 'Worksheet' object has no attribute 'max_col'
    xllist = []
    xllists = []
    for i in range(2, max_row+1):  #タイトル行は飛ばす
        
        for j in range(1,ncol+1):
            coldata = sheet.cell(row=i, column=j).value
            xllist.append(coldata)
            
        xllists.append(xllist)
        xllist = []    
    return xllists

#fpyxllist_to_list_s#########################################################
"""
fpyxllist_to_list_s: 
    excelfileのリストを　lists'　list にする
    sheet version
    ｖ1.00
    2022/9/26
    @author: jicc
    
"""
def fpyxllist_to_list_s(sheet, ncol):
    """
    excelfileのデータをlists'listにする
    sheet version
    Parameters
    ----------
    sheet : worksheet.worksheet.Worksheet
         worksheet object 
    ncol :  int
        number of columns
    Returns
    -------
    xllists : lists' list

    """
    #import chghistory
    #import openpyxl
    
    xllist = []

    xllists = []
    for i in range(2, sheet.max_row+1):  #タイトル行は飛ばす
        
        for j in range(1,ncol+1):
            coldata = sheet.cell(row=i, column=j).value
            xllist.append(coldata)
            
        xllists.append(xllist)
        xllist = []    
    return xllists

#fpyxllsit_to_indlist######################################################
"""
fpyxllist_to_indlist:
    get an individual lists' list from excelfile's list
        
    ｖ1.00
    2022/7/12
    @author: jicc
    
"""
def fpyxllist_to_indlist(wbN, sheetN, ncol, idno):
    """
    get an individual lists' list from excelfile's list

    Parameters
    ----------
    wbN : str
        Excelfile to move History data  '??_CowsHistory.xlsx'
    sheetN : str
        sheet name to add data   '??Farm' 
    ncol :  int
        number of columns
    idno : str
        ex. "0123456789"
    Returns
    -------
    xllists : lists' list

    """
    #from jiccModule import chghistory
    #import chghistory
    #import openpyxl
    
    wbobj = fpyopenxl(wbN, sheetN)   #get Worksheet object
    #wb = wbobj[0]
    sheet = wbobj[1]
    #wb = openpyxl.load_workbook(wbN)
    #sheet = wb[sheetN]
    max_row = sheet.max_row
    # max_col = sheet.max_col
    #AttributeError: 'Worksheet' object has no attribute 'max_col'
    xllist = []
    xllists = []
    for i in range(2, max_row+1):   #タイトル行は飛ばす
        idno_ = fpygetCell_value(sheet, i, 2) 
        #excellist's idno column 2
        if idno_ == idno:
            for j in range(1,ncol+1):
                coldata = sheet.cell(row=i, column=j).value
                xllist.append(coldata)
            
            xllists.append(xllist)
            xllist = [] 
            
    return xllists

#fpyxllist_to_indlist_s######################################################
"""
fpyxllist_to_indlist_s:
    get an individual lists' list from excelfile's list
    arguments 'wbN, sheetN' -> 'sheet' worksheetobject version    
    ｖ1.00
    2022/7/17
    @author: jicc
    
"""
def fpyxllist_to_indlist_s(sheet, ncol, idno):
    """
    get an individual lists' list from excelfile's list

    Parameters
    ----------
    sheet : worksheet.worksheet.Worksheet
         worksheet object
    ncol :  int
        number of columns
    idno : str
        ex. "0123456789"
    Returns
    -------
    xllists : lists' list

    """
    #import chghistory
  
    max_row = sheet.max_row

    xllist = []
    xllists = []
    for i in range(2, max_row+1):   #タイトル行は飛ばす
        idno_ = fpygetCell_value(sheet, i, 2) 
        #excellist's idno column 2
        if idno_ == idno:
            for j in range(1,ncol+1):
                coldata = sheet.cell(row=i, column=j).value
                xllist.append(coldata)
            
            xllists.append(xllist)
            xllist = [] 
            
    return xllists

#fpyxllist_to_indlist_s_######################################################
"""
fpyxllist_to_indlist_s_:
    get an individual lists' list from excelfile's list
    arguments 'wbN, sheetN' -> 'sheet' worksheetobject version
    add Excel's rowNo to xllist[0]  *)   
    ｖ1.00
    2022/10/17
    @author: jicc
    
"""
def fpyxllist_to_indlist_s_(sheet, ncol, idno):
    """
    get an individual lists' list from excelfile's list
    sheet version
    add Excel's rowNo to xllist[0] *)
    
    Parameters
    ----------
    sheet : worksheet.worksheet.Worksheet
         worksheet object
    ncol :  int
        number of columns
    idno : str
        ex. "0123456789"
    Returns
    -------
    xllists : lists' list

    """
    #import chghistory
  
    max_row = sheet.max_row

    xllist = []
    xllists = []
    for i in range(2, max_row+1):   #タイトル行は飛ばす
        idno_ = fpygetCell_value(sheet, i, 2) 
        #excellist's idno column 2
        if idno_ == idno:
            xllist.append(i) #xllist[0] : Excel's rowNo *)
            for j in range(1,ncol+1):
                coldata = sheet.cell(row=i, column=j).value
                xllist.append(coldata)
            
            xllists.append(xllist)
            xllist = [] 
            
    return xllists

#fpylst_to_indlst############################################################
"""
fpylst_to_indlst : 
    get an individual lists' list from Farm's lists' list
    list version
    v1.0
    2024/3/2
    @author: jicc
"""

def fpylst_to_indlst( xllists, idNo, index ):
    """
    get an individual lists' list from Farm's lists' list
    list version

    Parameters
    ----------
    xllists : list
        a list of Farm's calving data
    idNo : str
        '0123456789' 個体識別番号
    index : int
        the index of the element idNo
        リストの要素で、個体識別番号の入るインデックス
    Returns
    -------
    list : cow idNo's data

    """
    lxllists = len(xllists) #the length of list xllists
    
    idNos_data = [] # a list of cow idNo's (calving) data : default
    
    for i in range( 0,lxllists ):
        
        if xllists[i][index] == idNo:
            idNos_data.append(xllists[i])
        else:
            continue
        
    return idNos_data

#fpyaddclm_to_lsts_lst####################################################
"""
fpyaddclm_to_lsts_lst : 
   lists'listに最終カラムを追加する
   
   v1.0
   2022/3/28

@author: inoue
"""
def fpyaddclm_to_lsts_lst(xllists, colv):
    """
    lists'listに最終カラムを追加する

    Parameters
    ----------
    xllists : lists'list
        lists'list from Excelfile
    colv : int str None etc
        
    Returns
    -------
    最終列を追加した　lists'list 

    """
    
    lxll = len(xllists)
    for i in range(0, lxll):
        xllists[i].append(colv)
    return xllists

#fpydelclm_frm_lsts_lst#################################################
"""
fpydelclm_frm_lsts_lst : 
   lists'listのカラムを削除する
   
   v1.0
   2022/3/28

@author: inoue
"""
def fpydelclm_frm_lsts_lst(xllists, col):
    """
    lists'listのカラムを削除する

    Parameters
    ----------
    xllists : lists'list
        lists'list from ExcelFile
    col : int 
    削除する列番号   
    Returns
    -------
    列を削除した　lists'list 

    """
    
    lxll = len(xllists)
    for i in range(0, lxll):
        del xllists[i][col]
    return xllists

#fpyflag_dblrcd_1#######################################################
"""
fpyflag_dblrcd_1 : flag double record 1
   lists'listの重複リストに　1（重複）でチェックを入れる
   v1.01
   2022/4/3
   検索年月日追加のため, #*11->12 に変更
   v1.01
   2023/10/14
   *) 第2indexを変更
    v1.02
    2024/1/3
   #** olddata を削除対象の重複データとするように変更
   v1.03
   2024/1/11
   @author: inoue
   
"""
def fpyflag_dblrcd_1(xllists):
    """
    lists'listの重複リストに　1（重複）でチェックを入れる

    Parameters
    ----------
    xllists : lists'list
        lists'list from Excelfile

    Returns
    -------
    重複リストに "1"を追加した　lists'list 

    """
    
    lxll = len(xllists)
    #xldblrows = []
    for i in range(0, lxll):
        #print(xllists[i])
        #k=0
        
        for j in range(0, i+1):
            #range(0, i) とし、if j!= i節を削除可能かも　2024/1/12
            #print(xllists[j])
            if j!= i:
                if xllists[i][1:6] == xllists[j][1:6] and xllists[i][7:11] \
                == xllists[j][7:11]: #5 -> 6, 10 -> 11 に修正2024/1/3 *)
                #LinNo と No 以外が一致したら v1.01
                #clmn 13(list index 12) flg(0) 0->1とする
                    
                    #newdata を削除対象の重複データとする場合
                    #xllists[i][12] = 1          #*11->12 に変更
                    
                    #olddata を削除対象の重複データとする場合  #**
                    xllists[j][12] = 1

                else:
                    continue
            else:
                continue
            
    return xllists 

#fpyflag_dblrcd_1_###########################################################
"""
fpyflag_dblrcd_1_ : flag double record 1
   2つのlists'list　listorgとlisttmpを比較し、
   listtmpの重複リストに　1（重複）でチェックを入れる
   v1.0
   2022/7/15
   *)[1:]->[1:11] LineNo, 検索年月日をのぞいたリストの一致
   v1.01
   2023/10/5
   @author: jicc
   
"""
def fpyflag_dblrcd_1_(xllists, trs_inf):
    """
    2つのlists'list　listorgとlisttmpを比較し、
   listtmpの重複リストに　1（重複）でチェックを入れる

    Parameters
    ----------
    xllists : lists'list
        lists'list from Excelfile original list
    
    trs_inf : lists'list
        lists'list from web search data  

    Returns
    -------
    重複リストに "1"を追加した　lists'list listtmp 

    """
    lxll = len(xllists)
    ltrs = len(trs_inf)
    #print("xllists")
    #print(xllists)
    #print("trs_inf")
    #print(trs_inf)
    for i in range(1, ltrs):    #columns' list skip
        for j in range(0, lxll):
            #print(xllists[j])
            if trs_inf[i][0:10] == xllists[j][1:11]:    #*)[1:]->[1:11] 
                trs_inf[i][11] = 1
            else:
                continue
            
    return trs_inf

#fpydel_dblrcd##############################################################
"""
fpydel_dblrcd : delete double record
   lists'listの重複リストか新リストのどちらかを削除する
   add argument coln v1.01 2022/7/16
   v1.01
   2022/7/16

@author: inoue
"""
def fpydel_dblrcd(xllists, coln, colv):
    """
    lists'listの重複リストか新リストのどちらかを削除する
    
    Parameters
    ----------
    xllists : lists'list
        lists'list from Excelfile
    coln : int
        column flag's number 
    colv : int
        0, 1

    Returns
    -------
    lists'list : xllists[?][coln](flag)==colv only　

    """
    
    lxll = len(xllists)
    xllists_ = []
    for i in range(0, lxll):
        if xllists[i][coln] == colv:
            xllists_.append(xllists[i])
        else:
            continue

    return xllists_ 

#fpylisttoxls############################################################
"""
fpylisttoxls: 
    listのデータをexcelfileに移行する
    ｖ2.0
    2022/7/28
    @author: jicc
    
"""
def fpylisttoxls(xllist, fstcol, wbN, sheetN):
    """
    listのデータをexcelfileに移行する
    開始行　sheet.max_row + 1
    開始列 fstcol

    Parameters
    ----------
    xllist : str
        list from original csvfile  'MH_???_History.csv'
        [[..], [..],..]
    fstcol : int
        first column number to input data
    wbN : str
        Excelfile to move History data  'MH_CowsHistory.xlsx'
    sheetN : str
        sheet name to add data   'MHFarm' 

    Returns
    -------
    None.

    """
    #from jiccModule import chghistory
    #import chghistory
    #import openpyxl
    
    wbobj = fpyopenxl(wbN, sheetN)   #get Worksheet object
    wb = wbobj[0]
    sheet = wbobj[1]
    #wb = openpyxl.load_workbook(wbN)
    #sheet = wb[sheetN]
    max_row = sheet.max_row                     
    rn = max_row + 1 #first row to input records
    ln = len(xllist)
           #the length of the list xllist
    if ln > 0: #リストに要素がない場合を排除 v1.01 2022/4/3
        ln_ = len(xllist[0])   #the number of the xllist's list[0]
        for i in range(0, ln):
            for j in range(0, ln_):
                sheet.cell(row=rn, column=j+fstcol).value = xllist[i][j]
            rn = rn + 1
            print('add a new transfer informatyon')
    else:
    	print(' xllist have no element!')
        
    wb.save(wbN)        

#fpylisttoxls_s###########################################################
"""
fpylisttoxls_s: 
    listのデータをexcelfileに移行する
    ｖ2.0
    2022/7/28
    
    @author: jicc
    
"""
def fpylisttoxls_s(xllist, fstcol, sheet):
    """
    listのデータをexcelfileに移行する
    開始行　sheet.max_row + 1
    開始列 fstcol
    arguments 'wbN, sheetN' -> 'sheet' worksheetobject version 
    
    Parameters
    ----------
    xllist : str
        list from original csvfile  'MH_???_History.csv'
        [....]
    fstcol : int
        first column number to input data
   sheet : worksheet.worksheet.Worksheet
        worksheet object

    Returns
    -------
    None.

    """
    #import chghistory
    #import openpyxl
    
    max_row = sheet.max_row                     
    rn = max_row + 1 #first row to input records
    ln = len(xllist)   #count the number of xllist's elements
    if ln > 0: #リストに要素がない場合を排除 v1.01 2022/4/3
        for i in range(0, ln):
            sheet.cell(row=rn, column=i+fstcol).value = xllist[i]
        rn = rn + 1
            #print('add a new transfer information')　#2022/12/3 削除 *)
    else:
    	print(' xllist have no element!')
        
    print('add  new data')  #2022/12/4 *) から変更
    
    #return sheet
    #wb.save(wbN)


#fpylisttoxls_s_############################################################
"""
fpylisttoxls_s_: 
    listのデータをexcelfileに移行する
    ｖ2.0
    2022/7/28
    @author: jicc
    #cowshistory 以外で使用していた場合を考えて fpylisttoxls_s_ として保存
    2023/10/13
    
"""
def fpylisttoxls_s_(xllist, fstcol, sheet):
    """
    listのデータをexcelfileに移行する
    開始行　sheet.max_row + 1
    開始列 fstcol
    arguments 'wbN, sheetN' -> 'sheet' worksheetobject version 
    
    Parameters
    ----------
    xllist : str
        list from original csvfile  'MH_???_History.csv'
        [[..], [..],..]
    fstcol : int
        first column number to input data
   sheet : worksheet.worksheet.Worksheet
        worksheet object

    Returns
    -------
    None.

    """
    #import chghistory
    #import openpyxl
    
    max_row = sheet.max_row                     
    rn = max_row + 1 #first row to input records
    ln = len(xllist)
           #the length of the list xllist
    if ln > 0: #リストに要素がない場合を排除 v1.01 2022/4/3
        ln_ = len(xllist[0])   #the number of the xllist's list[0]
        for i in range(0, ln):
            for j in range(0, ln_):
                sheet.cell(row=rn, column=j+fstcol).value = xllist[i][j]
            rn = rn + 1
            #print('add a new transfer informatyon')　#2022/12/3 削除 *)
    else:
    	print(' xllist have no element!')
        
    print('add  new data')  #2022/12/4 *) から変更
    
    #wb.save(wbN)

#fpylisttoxls_s_ow############################################################
"""
fpylisttoxls_s_: 
    sheet のlistのデータをexcelfile sheet に上書きする
    overwrite Excel sheet with a modified list 
    ｖ1.0
    2024/2/11
    @author: jicc
        
"""
def fpylisttoxls_s_ow(xllist, fstcol, sheet):
    """
    listのデータをexcelfileに移行する
    
    開始行　2
    開始列 fstcol
    
    Parameters
    ----------
    xllist : str
        list from original Excel sheet 'AB_cowslist/cowslist2024'
    fstcol : int
        first column number to input data
   sheet : worksheet.worksheet.Worksheet
        worksheet object

    Returns
    -------
    None.

    """
    #import chghistory
    #import openpyxl
    
    #max_row = sheet.max_row                     
    rn = 2 #title行を除く
    ln = len(xllist)
           #the length of the list xllist
    if ln > 0: #リストに要素がない場合を排除 v1.01 2022/4/3
        ln_ = len(xllist[0])   #the number of the xllist's list[0]
        for i in range(0, ln):
            for j in range(0, ln_):
                sheet.cell(row=rn, column=j+fstcol).value = xllist[i][j]
            rn = rn + 1
            #print('add a new transfer informatyon')　#2022/12/3 削除 *)
    else:
    	print(' xllist have no element!')


#fpychk_drecords#########################################################
"""
fpychk_drecords   :check doublue records
    重複データを別シートに抜き出す
    v1.0
    2022/3/30
    検索年月日追加のため #* 11->12 に変更
    v1.01
    2023/10/14
    @author: inoue
    
"""
def fpychk_drecords(wbN, sheetN):
    """
    check doublue records
    重複データを別シートに抜き出す
    Parameters
    ----------
    wbN : str
        Excelfile to check double data  '??_CowsHistory.xlsx'
    sheetN : str
        sheet name to check double data   '??Farm'
    searchdate : str
        'yyyy/mm/dd' 検索年月日

    Returns
    -------
    None.

    """
    import chghistory
    #wbobj = chghistory.fpyopenxl(wbN, sheetN)
    #wb = wbobj[0]
    #sheet = wbobj[1]
    
    #excelfileのデータをlists'listにする
    xllists = chghistory.fpyxllist_to_list(wbN,sheetN, 12)      #*
    #print("xllists")
    #print(xllists)
    #value"0"のカラムflagをすべてのリストに追加する
    xllists_0 = chghistory.fpyaddclm_to_lsts_lst(xllists, 0)
    #print("xllists_0")
    #print(xllists_0)
    #重複データのflagを0->1に変更する
    xllists_01 = fpyflag_dblrcd_1(xllists_0)
    #print("xllists_01")
    #print(xllists_01)
    #重複データのないlist
    xllists0 = chghistory.fpydel_dblrcd(xllists_01, 12, 0)      #*
    #print("xllists0")
    #print(xllists0)
  
    #重複していたデータのリスト
    xllists1 = chghistory.fpydel_dblrcd(xllists_01, 12, 1)      #*
    ##print(xllists1)
   
    xllists0 = chghistory.fpydelclm_frm_lsts_lst(xllists0, 12)  #*
    #print("xllists0")
    #print(xllists0)
    #col 'flag'の削除
    
    xllists1 = chghistory.fpydelclm_frm_lsts_lst(xllists1, 12)  #*
    #print("xllists1")
    #print(xllists1)
    #col 'flag'の削除
    
    
    #シート名の変更
    chghistory.fpychgSheetTitle(wbN, sheetN, sheetN + 'org')
    #振り分け用のシート　KTFarm　と　KTFarmout　を作成する。
    chghistory.fpyNewSheet(wbN, sheetN, 'columns', 1)
    chghistory.fpyNewSheet(wbN, sheetN + 'del', 'columns', 1)
    #データを振り分ける
    chghistory.fpylisttoxls( xllists0, 1, wbN, sheetN)
    chghistory.fpylisttoxls( xllists1, 1, wbN, sheetN + 'del')
    
#fpychk_drecords_#########################################################
"""
fpychk_drecords_   :check doublue records
    重複データを別シートに抜き出す
    v1.0
    2022/3/30
    検索年月日追加のため #* 11->12 に変更
    v1.01
    2023/10/14
    add a parameter searchdate 
    重複をのぞいた最終リストの検索年月日をすべてsearchdateにする　#**
    v1.02
    2024/1/6
    @author: inoue
    注)fpyflag_dblrcd_1(xllists)の変更で検索年月日の入力を回避したので、
    v1.01に戻った。　この関数は使用せず。　2024/1/11
    注2）CowsHistory_webscrsys/ps_fpychk_drecords__args.py で使用 '__'に注意。
      後で、これは要らないことが判明。ps_fpychk_drecords_args.pywo
    修正した。2024/1/14
"""
def fpychk_drecords_(wbN, sheetN, searchdate):
    """
    check doublue records
    重複データを別シートに抜き出す
    Parameters
    ----------
    wbN : str
        Excelfile to check double data  '??_CowsHistory.xlsx'
    sheetN : str
        sheet name to check double data   '??Farm'
    searchdate : str
        'yyyy/mm/dd' 検索年月日

    Returns
    -------
    None.

    """
    import chghistory
    #wbobj = chghistory.fpyopenxl(wbN, sheetN)
    #wb = wbobj[0]
    #sheet = wbobj[1]
    
    #excelfileのデータをlists'listにする
    xllists = chghistory.fpyxllist_to_list(wbN,sheetN, 12)      #*
    #print("xllists")
    #print(xllists)
    #value"0"のカラムflagをすべてのリストに追加する
    xllists_0 = chghistory.fpyaddclm_to_lsts_lst(xllists, 0)
    #print("xllists_0")
    #print(xllists_0)
    #重複データのflagを0->1に変更する
    xllists_01 = fpyflag_dblrcd_1(xllists_0)
    #print("xllists_01")
    #print(xllists_01)
    #重複データのないlist
    xllists0 = chghistory.fpydel_dblrcd(xllists_01, 12, 0)      #*
    #print("xllists0")
    #print(xllists0)
    
    if type(searchdate) == str: #date = 'str'の場合datetimtに変換 #**
        searchdate = chghistory.fpystrtodatetime( searchdate )
    lxllists0 = len(xllists0)
    for i in range(0,lxllists0):
        xllists0[i][11] = searchdate                            #**
    #** 検索年月日をすべてserchdateにする
        
    print("xllists0")
    print(xllists0)
    #重複していたデータのリスト
    xllists1 = chghistory.fpydel_dblrcd(xllists_01, 12, 1)      #*
    ##print(xllists1)
   
    xllists0 = chghistory.fpydelclm_frm_lsts_lst(xllists0, 12)  #*
    #print("xllists0")
    #print(xllists0)
    #col 'flag'の削除
    
    xllists1 = chghistory.fpydelclm_frm_lsts_lst(xllists1, 12)  #*
    #print("xllists1")
    #print(xllists1)
    #col 'flag'の削除
    
    
    #シート名の変更
    chghistory.fpychgSheetTitle(wbN, sheetN, sheetN + 'org')
    #振り分け用のシート　KTFarm　と　KTFarmout　を作成する。
    chghistory.fpyNewSheet(wbN, sheetN, 'columns', 1)
    chghistory.fpyNewSheet(wbN, sheetN + 'del', 'columns', 1)
    #データを振り分ける
    chghistory.fpylisttoxls( xllists0, 1, wbN, sheetN)
    chghistory.fpylisttoxls( xllists1, 1, wbN, sheetN + 'del')

#fpyreplace_str#########################################################
"""
fpyreplace_str : replace str to another str
    v1.0
    2022/7/12
    @author: jicc
    
"""
def fpyreplace_str(text, txt0, txt1):
    '''
    replace str to another str

    Parameters
    ----------
    text : str
     ex. abc\u3000def
    txt0 : str
        ex. \u3000
    txt1 : str
        ex. ' '

    Returns
    -------
    txt
    'abc def'

    '''

    txt = text.replace(txt0, txt1)
    
    return txt

#fpydel_blanc###############################################################
"""
fpydel_blanc : delete a blanc character
    v1.0
    2022/9/13
    @author: jicc
    
"""
#import re
def fpydel_blanc(text):
    '''
    delete a blanc character

    Parameters
    ----------
    text : str
     ex. abc\u3000def
   
    Returns
    -------
    txt
    'abc def'

    '''

    text = re.sub(r"\s+", "", text)     #\s+ blanc regexp
    
    return text

#fpydel_quote################################################################
"""
fpydel_quote : delete a single or double quote 
    v1.0
    2022/9/13
    @author: jicc
    
"""
#import re
def fpydel_quote(text):
    '''
    delete a single or double quote 
    Parameters
    ----------
    text : str
     ex. 'abcdef',"abcdef"
   
    Returns
    -------
    text
    abcdef

    '''

    text = re.sub("\'|\"", "", text)     
    
    return text

#fpydel_hyphen###############################################################
"""
fpydel_hyphen : delete a hyphen
    v1.0
    2022/10/6
    @author: jicc
    
"""
#import re
def fpydel_hyphen(text):
    '''
    delete a hyphen

    Parameters
    ----------
    text : str
     ex. abc-def
    
    Returns
    -------
    text

    '''

    text = re.sub(r'[-]', "", text)     
    
    return text

#fpylstelemreplace_str#####################################################
"""
fpylstelemreplace_str : replace str to another str in a list's list
    v1.02
    2022/7/13
    @author: jicc
    
"""
def fpylstelemreplace_str(lst, elem, txt0, txt1):
    '''
    replace str to another str in a list's list

    Parameters
    ----------
    lst : list's list
     [[...], [...], ...]
     
    elem : int
        an element No of the target element to replace str 
    txt0 : str
        ex. \u3000
    txt1 : str
        ex. ' '

    Returns
    -------
    lst

    '''
    l = len(lst)
    for i in range(0, l):
        if lst[i][elem] == None:    #add if ~ else v1.01
            lst[i][elem] = '' #if lst[i][elem]...Noneとなる 
                              # lst[i][elem] -> '' に変更 v1.02
        else:
            lst[i][elem] = fpyreplace_str(lst[i][elem], txt0, txt1)

    return lst

#fpyselect_newrecords#######################################################
"""
fpyselect_newrecords   :select new records from transfer information
    異動情報から、新しいレコードを選択する
v1.0
2022/7/16

注) fpyselect_newrecords_s の修正、v1.01,1.02が反映されていない。2024/1/5

@author: inoue
"""
def fpyselect_newrecords(wbN, sheetN, ncol, idno):
    """
    select new records from transfer information
    異動情報から、新しいレコードを選択する
    Parameters
    ----------
    wbN : str
        Excelfile to check double data  'cowshistory.xlsx'
    sheetN : str
        sheet name of cowshistory   '??Farm'
    ncol :  int
        columns number of Excelfile's list
    idno : str
        ex. "0123456789"

    Returns
    -------
    lists' list　 [[title], [newrecords], [overlapped records]]

    """
    #import chghistory
    import nlbcs
    import time
    
    #excelfileのデータをlists'listにする
    xllists = fpyxllist_to_indlist(wbN, sheetN, ncol, idno)
    #氏名の全角空白'u\3000'を' 'に変換する
    xllists = fpylstelemreplace_str(xllists, 10, '\u3000', ' ')
   
    #個体識別情報検索画面のオープン
    driver = nlbcs.fpyopen_url("https://www.id.nlbc.go.jp/CattleSearch/search/agreement")
    nlbcs.fpyname_click(driver, "method:goSearch") 
    #個体識別番号 idno の情報を検索し、[[個体情報+異動情報], ...]
    #lists'list[[個体情報+異動情報], ...]を得る
    trs_inf = nlbcs.fpytrsinf_to_list(driver, idno)

    #lists' list [[title], [newrecords], [overlapped records]]
    trs_inf01 = [] #default
    
    #value"0"のカラムflagをすべてのリストに追加する
    trs_inf_0 = fpyaddclm_to_lsts_lst(trs_inf, 0)

    #xllistsにすでにあるlistのflagを0->1に変更する
    trs_inf_01 = fpyflag_dblrcd_1_(xllists, trs_inf_0)

    #list of new records
    trs_inf0 = fpydel_dblrcd(trs_inf_01, 10, 0)

    #list of overlapped records
    trs_inf1 = fpydel_dblrcd(trs_inf_01, 10, 1)

    #delete col 'flag'
    trs_inf0 = fpydelclm_frm_lsts_lst(trs_inf0, 10)

    #delete col 'flag'
    trs_inf1 = fpydelclm_frm_lsts_lst(trs_inf1, 10)
     
    trs_inf01.append(trs_inf0[0])  #[[title]]
    trs_inf01.append(trs_inf0[1:]) #[[title], [newrecords]]
    trs_inf01.append(trs_inf1)     #[[title], [newrecords], [overlapped records]] 
    print('trs_inf01')
    print(trs_inf01)
    
    time.sleep(3)
    nlbcs.fpydriver_quit(driver)
    
    return trs_inf01

#fpyselect_newrecords_s#####################################################
"""
fpyselect_newrecords_s   :select new records from transfer information
    異動情報から、新しいレコードを選択する
    arguments 'wbN, sheetN' -> 'sheet' worksheetobject and 
    add arguments 'driver' Webdriver object
    v1.0
    2022/7/17
    #*  change str yyyy/0m/0d to datetime yyyy/m/d
    v1.01
    2023/10/7
    #** コメント修正、追加
    v1.02
    2024/1/5
    @author: jicc
"""
def fpyselect_newrecords_s(driver, sheet, ncol, idno):
    """
    select new records from transfer information
    異動情報から、新しいレコードを選択する
    Parameters
    ----------
    driver : webdriver.chrome.webdriver.WebDriver
        WebDriver object of selenium.webdriver.chrome.webdriver module
    sheet : worksheet.worksheet.Worksheet
         worksheet object
    ncol :  int
        the number of columns of sheet(Excelfile's list)
    idno : str
        ex. "0123456789"

    Returns
    -------
    trs_inf01 : lists'list
    [[title], [[newrecord],..], [[overlapped record],..]] 

    """
    import chghistory
    import nlbcs
    #import time
    import fmstls
    
    #excelfileのデータをlists'listにする
    xllists = chghistory.fpyxllist_to_indlist_s(sheet, ncol, idno)
    #print("xllists")
    #print(xllists)
    xllists = chghistory.fpylstelemreplace_str(xllists, 10, '\u3000', ' ')
    #list 10(11番目)の"氏名または名称"の全角空白を半角空白に変換 
    #print('xllists')
    #print(xllists)    
    trs_inf = nlbcs.fpytrsinf_to_list(driver, idno)
    #lists' list [[title], [trsinf1],[trs_inf2],...]   #**
    #print('trs_inf')
    #print(trs_inf)
    
    l = len(trs_inf)
    for i in range(1,l): #*
            #出生の年月日 yyyy/0m/0d ->yyyy/m/d
            yyyy_mm_dd_0 = fmstls.fpyymd_0mtom_0dtod_(trs_inf[i][1])
            trs_inf[i][1] = yyyy_mm_dd_0
            #異動年月日 yyyy/0m/0d ->yyyy/m/d
            yyyy_mm_dd_1 = fmstls.fpyymd_0mtom_0dtod_(trs_inf[i][7])
            trs_inf[i][7] = yyyy_mm_dd_1
            #検索年月日 yyyy/0m/0d ->yyyy/m/d
            yyyy_mm_dd_2 = fmstls.fpyymd_0mtom_0dtod_(trs_inf[i][10])
            trs_inf[i][10] = yyyy_mm_dd_2                               #*
    
    #print('trs_inf')
    #print(trs_inf)
    
    trs_inf01 = [] #default
    #value"0"のカラムflagをすべてのリストに追加する
    #titleリストの最後にも0が追加されている
    trs_inf_0 = chghistory.fpyaddclm_to_lsts_lst(trs_inf, 0)
    
    #xllistsにすでにあるlistのflagを0->1に変更する
    trs_inf_01 = fpyflag_dblrcd_1_(xllists, trs_inf_0)
    
    #list of new records　[[title], [newrecord], ...]
    trs_inf0 = chghistory.fpydel_dblrcd(trs_inf_01, 11, 0)
    
    #list of overlapped records
    trs_inf1 = chghistory.fpydel_dblrcd(trs_inf_01, 11, 1)
    
    #delete col 'flag'
    trs_inf0 = chghistory.fpydelclm_frm_lsts_lst(trs_inf0, 11)
    
    #delete col 'flag'
    trs_inf1 = chghistory.fpydelclm_frm_lsts_lst(trs_inf1, 11)
        
    trs_inf01.append(trs_inf0[0])  #[[title]]
    trs_inf01.append(trs_inf0[1:]) 
    #[title]を除く   #**
    #[[title], [[newrecords],..]]
    trs_inf01.append(trs_inf1)     
    #[[title], [[newrecord],..], [[overlapped record],..]] 
    #print('trs_inf01')
    #print(trs_inf01)
    
    #time.sleep(3)
    #nlbcs.fpydriver_quit(driver)
    
    return trs_inf01
    #trs_inf[1] : newrecords, trs_inf[2] : overlapped records
    
#fpynewtrs_inf_to_list#####################################################
"""
fpynewtrs_inf_to_list:
    compare original taransfer information with new information and
    separate new recors and overlapped records
    v1.0
    2022/7/22
    @author: jicc
    
"""
def fpynewtrs_inf_to_list(wbN0, sheetN0, colidno0, wbN1, sheetN1, colidno1):
    """
    compare original taransfer information with new information and
    separate new recors and overlapped records

    Parameters
    ----------
    wbN0 : str
        Excelfile name of originaldata
        ex. "cowhistory.xlsx"
    sheetN0 : str
        sheet name
        ex. "MHFarm"
    colidno0 : int
        column number of 'idno0'(sheetN0 original data)
    
    wbN1 : str
        Excelfile name of new information
        ex. "??_cowslist.xlsx"
    sheetN1 : str
        sheet name
        ex. "cowslist"
    colidno1 : int
        column number of 'idno1' (sheetN1 new data)

    Returns
    -------
    trs_inf01 : lists'list
    [[newrecors'list], [overlappedrecords'list]]
    no title list

    """
    import nlbcs
    #import chghistory
    import time
    from selenium.common.exceptions import NoSuchElementException
    
    wb0 = fpyopenxl(wbN0, sheetN0)
    sheet0 = wb0[1]
    #max_row0 = sheet0.max_row
    
    wb1 = fpyopenxl(wbN1, sheetN1)
    sheet1 = wb1[1]
    max_row1 = sheet1.max_row
    
    trs_inf0 = [] #new records lists'list default
    trs_inf1 = [] #overlapped records lists'list default
    trs_inf01 = [] #all records which have searched

    driver = nlbcs.fpyopen_url("https://www.id.nlbc.go.jp/CattleSearch/search/agreement")
    nlbcs.fpyname_click(driver, "method:goSearch") 
    for row_num1 in range(2, max_row1 + 1):
        
        idno1 = fpygetCell_value(sheet1, row_num1, colidno1)
        
        try:
            tmp = fpyselect_newrecords_s(driver, sheet0, 11, idno1)
            #trs_inf0.append(tmp[0]) #columns list
            trs_inf0.append(tmp[1]) #new records list
            #trs_inf1.append(tmp[0]) #columns list
            trs_inf1.append(tmp[2]) #overlapped records list

        except NoSuchElementException:
             print("Error: " + idno1 + " not found")
                
    trs_inf01 = [trs_inf0, trs_inf1]
    #[[[newrecord],..], [[overlapped record],..]] 
    
    time.sleep(3)
    nlbcs.fpydriver_quit(driver)
    return trs_inf01

#fpynewtrs_inf_to_list_s#####################################################
"""
fpynewtrs_inf_to_list_s:
    compare original taransfer information with new information and
    separate new recors and overlapped records
    arguments 'wbN?, sheetN?' -> 'sheet?' worksheetobject version 
    v1.0
    2022/7/27
    #* change parameter 11 to 12 because of an added column 'seaching date'
    v1.01
    2023/10/7
    #** modified trs_inf0 and trs_inf1
    v2.0
    2023/10/13
    
    @author: jicc
    
"""
def fpynewtrs_inf_to_list_s(sheet0, colidno0, sheet1, colidno1):
    """
    compare original taransfer information with new information and
    separate new recors and overlapped records

    Parameters
    ----------
    sheet0 : worksheet.worksheet.Worksheet
         worksheet object
    colidno0 : int
        column number of 'idno0'(sheet0 original data ex.MHFarm)
    sheet1 : worksheet.worksheet.Worksheet
         worksheet object
    colidno1 : int
        column number of 'idno1' (sheet1 new data ex. "cowslist")

    Returns
    -------
    trs_inf01 : lists'list
    [[newrecords'list], [overlappedrecords'list]]
    no title list

    """
    import nlbcs
    #import chghistory
    import time
    from selenium.common.exceptions import NoSuchElementException
    
    max_row1 = sheet1.max_row
    
    trs_inf0 = [] #new records lists'list default
    trs_inf1 = [] #overlapped records lists'list default
    trs_inf01 = [] #all records which have searched

    driver = nlbcs.fpyopen_url("https://www.id.nlbc.go.jp/CattleSearch/search/agreement")
    nlbcs.fpyname_click(driver, "method:goSearch") 
    for row_num1 in range(2, max_row1 + 1):
        
        idno1 = fpygetCell_value(sheet1, row_num1, colidno1)
        
        try:
            tmp = fpyselect_newrecords_s(driver, sheet0, 12, idno1) #*
            #**
            #idno1の
            #lists' list [[title], [[newrecord],..], [[overlapped record],..]]
            ltmp1 = len(tmp[1]) #the number of elements in newrecords
            for i in range(0,ltmp1):
                trs_inf0.append(tmp[1][i]) #new records list
                #[[newrecord], [newrecord], ...]
            ltmp2 = len(tmp[2])
            for j in range(0,ltmp2):
                trs_inf1.append(tmp[2][j]) #overlapped records list
                #[[overlapped record], [overlapped record], ...]         #``
        except NoSuchElementException:
             print("Error: " + idno1 + " not found")
                
    trs_inf01 = [trs_inf0, trs_inf1]
    #[[[newrecord],..], [[overlapped record],..]] 
    
    time.sleep(3)
    nlbcs.fpydriver_quit(driver)
    return trs_inf01

#fpytrs_infs_to_xlsx########################################################
"""
fpytrs_infs_to_xlsx:
    search and save individual transfer informations to Excelfile
    v1.0
    2022/7/26
    @author: jicc
    
"""
def fpytrs_infs_to_xlsx(wbN0, sheetN0, wbN1, sheetN1, colidno1):
    """
    search and save individual transfer informations to Excelfile

    Parameters
    ----------
    wbN0 : str
        Excelfile name of originaldata
        ex. "AB_cowhistory.xlsx"
    sheetN0 : str
        sheet name
        ex. "ABFarm"
    wbN1 : str
        Excelfile name of new idno information
        ex. "AB_cowslist.xlsx"
    sheetN1 : str
        sheet name
        ex. "cowslist"
    colidno1 : int
        column number of 'idno1' (sheetN1 new data)

    Returns
    -------
    None.

    """
    import nlbcs
    #import chghistory
    import time
    from selenium.common.exceptions import NoSuchElementException
    
    wb0obj = fpyopenxl(wbN0, sheetN0) #[wb0, sheet0]
    wb0 = wb0obj[0] #ex. AB_cowshistory.xlsx
    sheet0 = wb0obj[1] #ex. ABFarm
    #max_row0 = sheet0.max_row
    
    wb1obj = fpyopenxl(wbN1, sheetN1) #[wb1, sheet1]
    #wb1 = wb1obj[0] #ex. AB_cowslist.xlsx
    sheet1 = wb1obj[1] #ex. cowslist
    max_row1 = sheet1.max_row 
    
    driver = nlbcs.fpyopen_url("https://www.id.nlbc.go.jp/CattleSearch/search/agreement")
    nlbcs.fpyname_click(driver, "method:goSearch") 
    for row_num1 in range(2, max_row1 + 1):
        
        idno1 = fpygetCell_value(sheet1, row_num1, colidno1)
        
        try:

            nlbcs.fpytrsinf_to_xlsx(driver, idno1, sheet0)            
            
        except NoSuchElementException:
             print("Error: " + idno1 + " not found")
    
    wb0.save(wbN0)
    time.sleep(3)
    nlbcs.fpydriver_quit(driver)
    
#fpynewtrs_infs_to_xlsx######################################################    
"""
fpynewtrs_infs_to_xlsx : 
    search individual transfer informations  
    select new transfer informations
    input and save Excelfile
    v1.0
    2022/7/28 @author: jicc
    
"""

def fpynewtrs_infs_to_xlsx(wbN0, sheetN0, colidno0, wbN1, sheetN1, colidno1):
    """
    search individual transfer informations  
    select new transfer informations
    input and save Excelfile    

    Parameters
    ----------
    wbN0 : str
        Excelfile name of originaldata
        ex. "AB_cowshistory.xlsx"
    sheetN0 : str
        sheet name
        ex. "ABFarm"
    colidno0 : int
        column number of 'idno0' (sheetN0 )
    wbN1 : str
        Excelfile name of new idno information
        ex. "AB_cowslist.xlsx"
    sheetN1 : str
        sheet name
        ex. "cowslist"
    colidno1 : int
        column number of 'idno1' (sheetN1 new data)

    Returns
    -------
    None.

    """

    #import chghistory
    #import fmstls
    
    wb0obj = fpyopenxl(wbN0, sheetN0)
    wb0 = wb0obj[0]
    sheet0 = wb0obj[1]
    #max_row0 = sheet0.max_row
    
    wb1obj = fpyopenxl(wbN1, sheetN1)
    sheet1 = wb1obj[1]
    #max_row1 = sheet1.max_row

    trs_inf01 = \
        fpynewtrs_inf_to_list_s(sheet0, colidno0, sheet1, colidno1)
        
    #print('trs_inf01')
    #print(trs_inf01)
        
    trs_inf0 = trs_inf01[0] #newrecords
        
    l0 = len(trs_inf0)
    if l0 > 0:
        for i in range(0, l0):

            fpylisttoxls_s(trs_inf0[i], 2, sheet0)
            
        wb0.save(wbN0)
        
#fpydel_d_idNo###############################################################
"""
fpydel_d_idNo:
    
    delete double idNos from idNos' list
    v1.0
    2023/6/2
    @author: inoue
    
"""
def fpydel_d_idNo( wbN, sheetN ):
    '''
    delete double idNos from idNos' list

    Parameters
    ----------
    wbN : str
        Excelfile name of cows'list
        ex. "AB_cowslist.xlsx"
    sheetN : str
        sheet name
        ex. "ABFarm"

    Returns
    -------
    None.

    '''
    #import openpyxl
    
    wb = openpyxl.load_workbook(wbN)
    sheet = wb[sheetN]
    max_row = sheet.max_row
    #print( sheet.max_row )
     
    for row_num in range( 2, max_row + 1 ):       
        
        #get the cowidNo of row row_num col 2
        cowidNo = sheet.cell(row=row_num, column=2).value
        for i in range( row_num+1, max_row + 1 ):
            cowidNo_ = sheet.cell(row=i, column=2).value
            if cowidNo == cowidNo_:
                sheet.delete_rows( i )
                max_row = max_row - 1
                print('delete row:'+ str(i))
                print('sheet.max_row:' + str(sheet.max_row))
                print('max_row:' + str(max_row))
                print('row_num:' + str(row_num))
                print('cowidNo' + cowidNo)
                print('cowidNo_' + cowidNo_)
            else:
                continue
        
    wb.save(wbN)

#fpyelems_lstfrmxls_lst######################################################
"""
fpyelems_lstfrmxls_lst : make an elements' list from Excel's list

    v1.0
    2023/10/1
    @author: inoue
"""
def fpyelems_lstfrmxls_lst(wbN, sheetN, coln):
    """
    make an elements' list from Excel's list

    Parameters
    ----------
    wbN : str
        Excelfile to check double data  'AB_cowshistory.xlsx'
    sheetN : str
        sheet name to check some column 'ABFarm'
    coln : int
        column's number to check elements

    Returns
    -------
    list

    """
    import chghistory
    wbobj = chghistory.fpyopenxl(wbN, sheetN)
    #wb = wbobj[0]
    sheet = wbobj[1]
    
    #max_row = sheet.max_row

    elements = []
    for i in range(2, sheet.max_row+1):   #タイトル行は飛ばす
        elem = fpygetCell_value(sheet, i, coln) 

        if elem not in elements:
            elements.append(elem)
        else:
            continue
            #print(elem + "is a element of the list elements.")
            
        
    return elements

#fpyelems_lstfrmxls_lst_s######################################################
"""
fpyelems_lstfrmxls_lst : make an elements' list from Excel's list
    sheet version
    v1.0
    2023/10/1
    @author: inoue
"""
def fpyelems_lstfrmxls_lst_s(sheet, coln):
    """
    make an elements' list from Excel's list

    Parameters
    ----------
    sheet : worksheet.worksheet.Worksheet
         worksheet object
    coln : int
        column's number to check elements

    Returns
    -------
    list

    """

    elements = []
    for i in range(2, sheet.max_row+1):   #タイトル行は飛ばす
        elem = fpygetCell_value(sheet, i, coln) 

        if elem not in elements:
            elements.append(elem)
        else:
            continue
            #print(elem + "is a element of the list elements.")
            
        
    return elements        
 
#fpyext_frmlsts_lst########################################################
"""
fpyext_frmlsts_lst:
    extract specific name's lists from lists' list
    ｖ1.00
    2023/10/16
    @author: jicc
    
"""
def fpyext_frmlsts_lst(lst, index, name):
    """
    extract specific name's lists from lists' list

    Parameters
    ----------
    lst : list
        individual transfer information 
    index : int
        index number of an element : Farm name
    name : str
        Farm name

    Returns
    -------
    list
    extracted list

    """
    llst = len(lst)
    lst_ = []
    for i in range(0,llst):
        
        if lst[i][index] == name:
            lst_.append(lst[i])
        else:
            continue
    
    return lst_

#fpyarr_frmlsts_lst###########################################################
"""
fpyarr_frmlsts_lst : arrange an individual and specific name's lists 
    個体の特定農場の異動情報を調整する
    最後の転出情報がかけていた場合の調整
    ex."転入" -> "搬入"の場合 : "搬入" -> "転出" とし、"住所", "氏名または名称"も変更し、
    xllixts_ の最後に追加する。
    注)xllists_ の中で、"転入" -> "転入" となって"転出"が欠けているいる場合は想定していない。
    ｖ1.0
    2023/12/16
    1) reason_to_transfer == "出生"　or "転入" ->
        reason_to_transfer == "出生" or reason_to_transfer == "転入"　に訂正
    2) ixllists__No < lxllists ->  ixllists_No < ixllistsNo に変更
    v1.01
    2023/12/24
    @author: jicc
    
"""

def fpyarr_frmlsts_lst( xllists, xllists_):
    """
    arrange an individual and specific name's lists 

    Parameters
    ----------
    xllists : list
        a list of an individual transfer information
    xlsists_ : list
        a list of an individual and specific name's lists 
     
    Returns
    -------
    xllists_ : list

    """
    lxllists = len(xllists)
    print("lxllists")
    print(lxllists)

    lxllists_ = len(xllists_)
    print("lxllists_")
    print(lxllists_)

    reason_to_transfer = xllists_[lxllists_-1][7] 
    #xllists_　最後の要素の"異動内容"
    print("reason_to_transfer")
    print(reason_to_transfer)
 
    if reason_to_transfer == "出生" or reason_to_transfer == "転入": # 1)
        
        xllistsNo = xllists[lxllists-1][6] # 2)
        #xllists　最後の要素の"No"
        ixllistsNo = int(xllistsNo) # 2*)
        
        xllists_No = xllists_[lxllists_-1][6]
        #xllists_　最後の要素の"No"
        ixllists_No = int(xllists_No)
        
        if ixllists_No < ixllistsNo: # 2)
            #xllists_の最後の要素のあとに、xllsitsの要素がある場合
            # "転出"が記載されていなくて、他所に"搬入", "転入"など異動している。
        
            tmp = xllists[ixllists_No]
            #xllistsで、xllists_最後の要素の次の要素
            print('tmp')
            print(tmp)
            tmp[7] = "転出"  #"転出"に変更
            tmp[9] = xllists_[lxllists_-1][9]   #"住所"を変更
            tmp[10] = xllists_[lxllists_-1][10] #"氏名または名称"を変更
            print('tmp')
            print(tmp)
        
            xllists_.append(tmp) #tmpをxllists_に追加して、転出した状態にする。
    
    print("arranged xllists_")
    print(xllists_)

    return xllists_

#fpyterms_in_farm############################################################
"""
fpyterms_in_farm:
    get a list 'term in farm'
    個体の牧場所属期間( a term in a farm)のリストを得る
    注) 異動情報に転出がなく、直接搬入など他所へ異動になっている場合の処理がない
    v1.0
    2023/12/9
    *)最後の転出情報がかけていた場合の調整
    ex."転入" -> "搬入"の場合 : "搬入" -> "転出" とし、"住所", "氏名または名称"も変更し、
    xllixts_ の最後に追加する。
    v1.1
    2023/12/16
    動き不十分のため、#**) 以下を全面書き換え
    v2.0
    2023/12/29
    by jicc
    
"""
def fpyterms_in_farm( wbN, sheetN, ncol, idno, name ):
    """
    get a list 'term in farm'

    Parameters
    ----------
    wbN : str
        Excelfile's name : AB_cowshistory.xlsx
    sheetN : str
        sheet name : ABFarm
    ncol : int
        number of columns 列数 : 12
    idno : str
        idNo 個体識別番号 : "0123456789"
    name : str
        氏名または名称　

    Returns
    -------
    terms_in_farm : list
        lists' list :[['出生'or'転入'の年月日, '転出','死亡'の年月日 or None],..] 

    """
    #from jiccModule import chghistory
    import chghistory
    #import openpyxl
    
    wbobj = chghistory.fpyopenxl(wbN, sheetN)   #get Worksheet object
    #wb = wbobj[0]
    sheet = wbobj[1]
    
    xllists = fpyxllist_to_indlist_s(sheet, ncol, idno)
    #個体識別番号 idNo の異動情報のリスト
    print("xllists")
    print(xllists)
    #lxllists = len(xllists)
    xllists.sort(key = lambda x:x[8]) #, reverse=True
    #lists' listを 異動年月日 昇順 でsort lambda関数を利用
       
    xllists_ = fpyext_frmlsts_lst(xllists, 10, name)
    #index 10 : "氏名または名称"
    #当該牧場の異動情報だけ抽出

    xllists_.sort(key = lambda x:x[8]) #, reverse=True
    #lists' listを 異動年月日 昇順 でsort lambda関数を利用
    
    print("xllists_")
    print(xllists_)
    
    xllists_ = chghistory.fpyarr_frmlsts_lst( xllists, xllists_)   # *)
    #最後の"転出"が欠けていた場合の調整
    print("arrxllists_")
    print(xllists_)
    
    #term_in_farm = []   #term_in_farm default
    #['出生'or'転入'の年月日, '転出'or'死亡'の年月日]　個体の牧場　1所属期間
    #terms_in_farm = []  #terms_in_farm default
    #　個体の牧場所属期間のすべてのリスト
    
    #**) v2.0
    lxllists_ = len(xllists_) #異動情報の要素数 
    
    #l_2 = lxllists_ // 2        #a divisor of lxllists_ by 2
    l_2_res = lxllists_ % 2     # a residural of lxllists_ by 2
    
    tmp = [] #異動年月日 昇順のリスト default
    terms_in_farm = []  #terms_in_farm default
    #　個体の牧場所属期間のすべてのリスト lists' list
    
    if l_2_res == 0: #異動情報の要素数が偶数 : 転出済の個体
        
        #異動年月日 昇順 のリスト tmp を作成
        for i in range(1,lxllists_+1):
            tmp.append(xllists_[0][8])
            del xllists_[0]
            
        #所属期間 ['出生'or'転入'の年月日, '転出'or'死亡'の年月日]
        #のリスト terms_in_farm(lists' list)を作成
        ltmp = len(tmp)
        lterms_in_farm = ltmp // 2
        for j in range(1,lterms_in_farm+1):
            terms_in_farm.append(tmp[0:2])  #term_in_farm
            del tmp[0:2]
        print('tmp')
        print(tmp)
        #print(idno)        
        print("terms_in_farm")
        print(terms_in_farm)
        
    elif l_2_res == 1:  #異動情報の要素数が奇数 : 所属している個体(転出していない)
    
        #異動年月日 昇順 のリスト tmp を作成
        for i in range(1,lxllists_+1):
            tmp.append(xllists_[0][8])
            del xllists_[0]
        
        #最後の転入以前の
        #所属期間 ['出生'or'転入'の年月日, '転出'or'死亡'の年月日]
        #のリスト terms_in_farm(lists' list)を作成
        ltmp = len(tmp)
        lterms_in_farm = ltmp // 2
        for j in range(1,lterms_in_farm+1):
            terms_in_farm.append(tmp[0:2])  #term_in_farm
            del tmp[0:2]
        
        print('tmp')
        print(tmp)
        print('terms_in_farm')
        print(terms_in_farm)
        
        #最後の転入(出生)の年月日だけ残っているtmpに、
        #None(転出していない) を加える        
        tmp.append(None)
        #最後の所属期間 ['出生'or'転入'の年月日, None] を加える。
        terms_in_farm.append(tmp[0:2])
        del tmp[0:2]
        print('tmp')
        print(tmp)
        #print(idno)        
        print("terms_in_farm")
        print(terms_in_farm)
 
    return terms_in_farm

#fpyterms_in_farm_############################################################
"""
fpyterms_in_farm_:
    get a list 'term in farm'
    a parameter is a list only
    v1.0
    2023/12/19
    動き不十分のため、全面書き換え
    v2.0
    2023/12/29
    by jicc
    
"""
def fpyterms_in_farm_( xllists_ ):
    """
    get a list 'term in farm'

    Parameters
    ----------
    xlsists_ : list
        an individual transfer informations' list at a specific farm 
        個体の特定の農場での異動情報のリスト
        注) リストは、異動年月日昇順であることが必要 必要なら　#*)を使用する
    Returns
    -------
    terms_in_farm : list
        lists' list :[['出生'or'転入'の年月日, '転出','死亡'の年月日 or None],..] 

    """
    #term_in_farm = [] default
    #['出生'or'転入'の年月日, '転出'or'死亡'の年月日]　個体の牧場　1所属期間
    
    #xllists_.sort(key = lambda x:x[8]) #, reverse=True  #*)
    #lists' listを 異動年月日 昇順 でsort lambda関数を利用
    
    lxllists_ = len(xllists_) #異動情報の要素数
    
    #l_2 = lxllists_ // 2        #a divisor of lxllists_ by 2
    l_2_res = lxllists_ % 2     # a residural of lxllists_ by 2
    
    tmp = [] #異動年月日 昇順のリスト default
    terms_in_farm = []  #terms_in_farm default
    #　個体の牧場所属期間のすべてのリスト lists' list
    
    if l_2_res == 0: #異動情報の要素数が偶数 : 転出済の個体
        
        #異動年月日 昇順 のリスト tmp を作成
        for i in range(1,lxllists_+1):
            tmp.append(xllists_[0][8])
            del xllists_[0]
            
        #所属期間 ['出生'or'転入'の年月日, '転出'or'死亡'の年月日]
        #のリスト terms_in_farm(lists' list)を作成
        ltmp = len(tmp)
        lterms_in_farm = ltmp // 2
        for j in range(1,lterms_in_farm+1):
            terms_in_farm.append(tmp[0:2])  #term_in_farm
            del tmp[0:2]
        print('tmp')
        print(tmp)
        #print(idno)        
        print("terms_in_farm")
        print(terms_in_farm)
        
    elif l_2_res == 1:  #異動情報の要素数が奇数 : 所属している個体(転出していない)
    
        #異動年月日 昇順 のリスト tmp を作成
        for i in range(1,lxllists_+1):
            tmp.append(xllists_[0][8])
            del xllists_[0]
        
        #最後の転入以前の
        #所属期間 ['出生'or'転入'の年月日, '転出'or'死亡'の年月日]
        #のリスト terms_in_farm(lists' list)を作成
        ltmp = len(tmp)
        lterms_in_farm = ltmp // 2
        for j in range(1,lterms_in_farm+1):
            terms_in_farm.append(tmp[0:2])  #term_in_farm
            del tmp[0:2]
        
        print('tmp')
        print(tmp)
        print('terms_in_farm')
        print(terms_in_farm)
        
        #最後の転入(出生)の年月日だけ残っているtmpに、
        #None(転出していない) を加える        
        tmp.append(None)
        #最後の所属期間 ['出生'or'転入'の年月日, None] を加える。
        terms_in_farm.append(tmp[0:2])
        del tmp[0:2]
        print('tmp')
        print(tmp)
        #print(idno)        
        print("terms_in_farm")
        print(terms_in_farm)
 
    return terms_in_farm

#fpybelong_or_not############################################################
"""
fpybelong_or_not : check a base date belongs to a period or not
    this function needs to import datetime
    v1.0
    20232/12/17
    @author: jicc
    
"""
def fpybelong_or_not( bdate, term ):
    '''
    check a base date belongs to a period or not

    Parameters
    ----------
    bdate : datetime.datetime
        base date
    term : list
        a belonging period:[in(datetime.datetime), out(datetime.datetime)] 

    Returns
    -------
    int belong_or_not : belong ==1, not == 0

    '''
    belong_or_not = None        #dfault
    
    if term[1] == None:
        if bdate >= term[0]:
            belong_or_not = 1
        elif bdate < term[0]:
            belong_or_not = 0
    
    else: 
    
        if bdate >= term[0] and bdate < term[1]:
        
            belong_or_not = 1
    
        elif bdate < term[0] or bdate >= term[1]:
        
            belong_or_not = 0
        
    
    return belong_or_not

#fpyind_belongornot#########################################################
"""
fpyind_belongornot : 
    an individual belongs to a farm or not at a base date
    this function needs to import datetime
    v1.0
    2023/12/19
    @author: jicc
    
"""
def fpyind_belongornot(bdate, terms):
    """
    
    an individual belongs to a farm or not at a base date

    Parameters
    ----------
    bdate : datetime.datetime
        base date
    terms : lists' list'
        belonging periods:
            [[in(datetime.datetime), out(datetime.datetime)] , [...], ]

    Returns
    -------
    int belong_or_not : belong ==1, not == 0

    """
    lterms = len(terms)
    for i in range(0, lterms):
        bn = fpybelong_or_not(bdate, terms[i])
        if bn == 1:
            break
        elif bn == 0:
            continue
        
    belongornot = bn
    
    return belongornot

#fpysep_outfrmin#############################################################
"""
fpysep_outfrmin
    separate move-out cows from move-in 
    異動情報のExcelfile: AB_cowshistory.xlsx の　sheet　ABFarmの情報を
    基準日における所属牛（転入牛move-in)と転出牛(move-out)の情報に分け、
    2枚のsheet ABFarmin, ABFarmout を作成する
    注) 使用前に２枚のsheet sheetN+'in'と sheetN+'out'を作成しておくこと
        chghistory.fpymkxlsheet(wbN, sheetN, scolN, r)
    v1.0
    2024/1/2
    当該牧場の異動情報だけの振り分けになっていたのを修正
    xllists_org を廃止
    v1.01
    2024/1/13
    @author: jicc
    
"""
def fpysep_outfrmin( wbN, sheetN, coln, ncol, index, name, bdate ):
    """
    separate move-out cows from move-in 

    Parameters
    ----------
    wbN : str
        Excelfile to check move-in or move-out data  
        'AB_cowshistory.xlsx'　対象のエクセルファイル名
    sheetN : str
        sheet name to separate move-out cows from move-in
        'ABFarm'　対象のエクセルシート名
    coln : int
        column's number of idNo 個体識別番号の入っている列番号
    ncol : int
        number of columns sheet ABFarm のリストの列数
    index : int
        index number of an element(Farm name)　
        リスト上の　'氏名または名称'のindex番号
    name : str
        Farm name '氏名または名称'
    bdate : str
        base date 基準日

    Returns
    -------
    None.

    """

    #import openpyxl
    #import chghistory
    #import datetime

    wb = openpyxl.load_workbook(wbN)
    sheet = wb[sheetN]
    sheetin = wb[sheetN + 'in']
    sheetout = wb[sheetN + 'out']
    
    #sheet 上の columns coln(個体識別番号)の要素のリストを作成
    idNos = fpyelems_lstfrmxls_lst_s(sheet, coln)
    print('idNos')
    print(idNos)

    lidNos = len(idNos)

    for i in range(0,lidNos):
        
        xllists = fpyxllist_to_indlist_s(sheet, ncol, idNos[i])
        #個体識別番号 idNo の異動情報のリスト
        print("xllists")
        print(xllists)
        lxllists = len(xllists)
        xllists.sort(key = lambda x:x[8]) #, reverse=True
        #lists' listを 異動年月日 昇順 でsort lambda関数を利用
        #No ([6])昇順でsortしたほうが良いかもしれない。2024/1/13
        
        xllists_ = fpyext_frmlsts_lst(xllists, index, name)
        #index 10 : "氏名または名称"
        #当該牧場の異動情報だけ抽出

        xllists_.sort(key = lambda x:x[8]) #, reverse=True
        #lists' listを 異動年月日 昇順 でsort lambda関数を利用
        
        print("xllists_")
        print(xllists_)
        
        xllists_ = fpyarr_frmlsts_lst( xllists, xllists_ )   # *)
        #最後の"転出"が欠けていた場合の調整
        print("xllists_")
        print(xllists_)
        
        terms_in_farm = fpyterms_in_farm_( xllists_ )
        #当該牧場にいた滞在期間
        print("terms_in_farm")
        print(terms_in_farm)
        
        #bdate = '2023/12/31'
        print(bdate)
        #基準日
        if type(bdate) == str: 
            bdate = datetime.datetime.strptime(bdate, '%Y/%m/%d')
            #datetimeに変換
        print('bdate')
        print(bdate)
        
        belongornot = fpyind_belongornot( bdate, terms_in_farm )
        #基準日にその農場に所属していたかどうか
        print('belongornot')
        print(belongornot)
        
        if belongornot == 0: #move-out
            for k in range(0, lxllists):
                fpylisttoxls_s(xllists[k], 1, sheetout)
                #wb.save('..\CD_cowshistory.xlsx')
        
        elif belongornot == 1: #move-in belonging
            for k in range(0, lxllists):
                fpylisttoxls_s(xllists[k], 1, sheetin)
                #wb.save('..\CD_cowshistory.xlsx')
                
    wb.save(wbN)

#fpychghistoryReference###################################################################
"""
fpychghistoryReference:         reference of chghistory's functions
ｖ2.0
2024/1/2
@author: jicc
"""
def fpychghistoryReference():
    
    print('-----chghistoryReference ---------------------------------------------------------v2.00------')
    print('**fpyopenxl(wbN, sheetN)')
    print('Excelfile wbN.xlsx　sheet sheetN Open ')
    print('.............................................................................................')
    print('**fpyopencsv_robj(csvN)')
    print('csvfile Open for Reader object')
    print('.............................................................................................')
    print('**fpyopencsv_rdata(csvN)')
    print('csvfile Open for Reader data')
    print('.............................................................................................')
    print('**fpyopencsv_w(csvN)')
    print('csvfile Open for Writer')
    print('.............................................................................................')
    print('**fpypdf_to_csv(filename, Path)')
    print('convert ****.pdf to ****.csv file')
    print('.............................................................................................')
    print('**fpySpdf_in_Dir_to_csv(Ext, Path)')
    print('ディレクトリ内の特定の拡張子(.pdf)を持つファイルを見つけcsvfile に変換する')
    print('.............................................................................................')
    print('**fpyCowHistory(csvorgN, csvoutN)')
    print('牛の個体情報.csvから、CowHistory.csv(changehistory\'s list )を作成する')
    print('.............................................................................................')
    print('**fpycsvlisttoxls(csvN, wbN, sheetN)')
    print('csvfileのデータをexcelfileに移行する')
    print('.............................................................................................')
    print('**fpycsvidNo_9to10( idNo )')
    print('idNo in a csvfile 9figures to 10figures')
    print('.............................................................................................')
    print('**fpydate_dottoslash( date )')
    print('date in a csvfile \'yyyy.mm.dd\' to \'yyyy/mm/dd\'')
    print('.............................................................................................')
    print('**fpyymd_csvtoCowsHistory_csv(Ext, Path, bckPath)')
    print('フォルダー内の個体履歴org(csv)をCowsHistory.csvに変更する')
    print('.............................................................................................')
    print('**fpyHistory_csvto_xlsx(Ext, Path, bckPath, wbN, sheetN)')
    print('フォルダー内の???H.csv)を???CowsHistory.xlsxに移動する')
    print('.............................................................................................')
    print('**fpystrtodatetime( date )')
    print('str\'yyyy/mm/dd\'をdatetime に変換する')
    print('.............................................................................................')
    print('**fpyxlstrymdtodatetime(wbN, sheetN, col)')
    print('Excel cell \'yyyy/mm/dd\'をdatetimeに変換する')
    print('.............................................................................................')
    print('**fpyxllist_to_list(wbN, sheetN, ncol)')
    print('excelfileのリストを　lists\'　list にする')
    print('.............................................................................................')
    print('**fpyxllist_to_list_s(sheet, ncol)')
    print('excelfileのリストを　lists\'　list にする')
    print('sheet version')
    print('.............................................................................................')
    print('**fpyxllist_to_indlist(wbN, sheetN, ncol, idno)')
    print('get an individual lists\' list from excelfile\'s list')
    print('.............................................................................................')
    print('get an individual lists\' list from excelfile\'s list')
    print('**fpyxllist_to_indlist_s(sheet, ncol, idno)')
    print('arguments \'wbN, sheetN\' -> \'sheet\' worksheetobject version ')
    print('.............................................................................................')
    print('get an individual lists\' list from excelfile\'s list')
    print('**fpyxllist_to_indlist_s_(sheet, ncol, idno)')
    print('arguments \'wbN, sheetN\' -> \'sheet\' worksheetobject version ')
    print('add Excel\'s rowNo to xllist[0]')
    print('.............................................................................................')
    print('get an individual lists\' list from Farm\'s lists\'list')
    print('**fpylst_to_indlst( xllists, idNo, index )')
    print('list version')
    print('.............................................................................................')
    print('**fpyaddclm_to_lsts_lst(xllists, colv)')
    print('lists\'listに最終カラムを追加する')
    print('.............................................................................................')
    print('**fpydelclm_frm_lsts_lst(xllists, col)')
    print('lists\'listのカラムを削除する')
    print('.............................................................................................')
    print('**fpyflag_dblrcd_1_(xllists, trs_inf)')
    print('2つのlists\'list　listorgとlisttmpを比較し、')
    print('2つのlists\'listtmpの重複リストに　1（重複）でチェックを入れる')
    print('.............................................................................................')
    print('**fpydel_dblrcd(xllists, ,coln, colv)')
    print('lists\'listの重複リストの一つを削除する')
    print('.............................................................................................')
    print('**fpylisttoxls(xllist, fstcol, wbN, sheetN)')
    print('listのデータをexcelfileに移行する')
    print('....................................................................................')
    print('**fpychgSheetTitle(wbN, sheetN, sheetN1)')
    print('change ExcelSheet\'s title')
    print('....................................................................................')
    print('**fpyreplace_str(text, txt0, txt1)')
    print('replace str to another str')
    print('....................................................................................')
    print('**fpydel_blanc(text)')
    print('delete a blanc character')
    print('....................................................................................')
    print('**fpydel_quote(text)')
    print('delete a single or double quote ')
    print('....................................................................................')
    print('**fpydel_hyphen(text)')
    print('delete a hyphen')
    print('....................................................................................')
    print('**fpylstelemreplace_str(lst, elem, txt0, txt1)')
    print('replace str to another str in a list\'s list')
    print('....................................................................................')
    print('**fpyselect_newrecords(wbN, sheetN, ncol, idno)')
    print('select new records from transfer information')
    print('異動情報から、新しいレコードを選択する')
    print('....................................................................................')
    print('**fpyselect_newrecords_s(driver, sheet, ncol, idno)')
    print('select new records from transfer information')
    print('異動情報から、新しいレコードを選択する')
    print('arguments \'wbN, sheetN\' -> \'sheet\' worksheetobject and ')
    print('add arguments \'driver\' Webdriver object')
    print('....................................................................................')
    print('**fpynewtrs_inf_to_list(wbN0, sheetN0, colidno0, wbN1, sheetN1, colidno1)')
    print('compare original taransfer information with new information and')
    print('separate new recors and overlapped records')
    print('....................................................................................')
    print('**fpytrs_infs_to_xlsx(wbN0, sheetN0, wbN1, sheetN1, colidno1)')
    print('search and save individual transfer informations to Excelfile')
    print('....................................................................................')
    print('**fpynewtrs_infs_to_xlsx(wbN0, sheetN0, colidno0, wbN1, sheetN1, colidno1)')
    print('search individual transfer informations')
    print('select new transfer informations')
    print('input and save Excelfile ')
    print('....................................................................................')
    print('**fpydel_d_idNo(wbN, sheetN)')
    print('delete double idNos from idNos\' list')
    print('....................................................................................')
    print('**fpymkd_path( path )')
    print('make a directory  at current directory')
    print('....................................................................................')
    print('**fpymkxlsheet(wbN, sheetN, scolN, r)')
    print('make an ExcelSheet if it dose not exist')
    print('....................................................................................')
    print('**fpymkxlsheet_(wb, sheetN, scolN, r)')
    print('make an ExcelSheet if it dose not exist')
    print('workbook version')
    print('....................................................................................')
    print('**fpyelems_lstfrmxls_lst(wbN, sheetN, coln)')
    print('  fpyelems_lstfrmxls_lst_s(sheet, coln)')
    print('make an elements\' list from Excel\'s list')
    print('....................................................................................')
    print('**fpyext_frmlsts_lst(lst, index, name)')
    print('extract specific name\'s lists from lists\' list')
    print('個体異動情報リストの、特定農場だけの異動情報を抽出する')
    print('....................................................................................')
    print('**fpyarr_frmlsts_lst( xllists, xllists_)')
    print('arrange an individual and specific name\'s lists ')
    print('個体の特定農場の異動情報を調整する,最後の転出情報がかけていた場合の調整')
    print('....................................................................................')
    print('**fpyterms_in_farm( wbN, sheetN, ncol, idno, name )')
    print('get a list \'term in farm\'')
    print('個体の牧場所属期間( a term in a farm)のリストを得る')
    print('注) 異動情報に転出がなく、直接搬入など他所へ異動になっている場合の処理がない')
    print('....................................................................................')
    print('**fpyterms_in_farm_( xllists_ )')
    print('get a list \'term in farm\'')
    print('個体の牧場所属期間( a term in a farm)のリストを得る')
    print('a paramete is a list only')
    print('....................................................................................')
    print('**fpyterms_in_farm_( xllists_ )')
    print('get a list \'term in farm\'')
    print('parameter is one list(xllists_) only version')
    print('....................................................................................')
    print('**fpybelong_or_not( bdate, term )')
    print('check a base date belongs to a period or not')
    print('基準日(bdate)に個体がその所属期間に属しているかどうか')
    print('....................................................................................')
    print('**fpyind_belongornot( bdate, terms )')
    print('an individual belongs to a farm or not at a base date')
    print('基準日(bdate)に個体がその農場に属しているかどうか')
    print('....................................................................................')
    print('**fpysep_outfrmin( wbN, sheetN, coln, ncol, index, name, bdate )')
    print('separate move-out cows from move-in ')
    print('異動情報のExcelfile: AB_cowshistory.xlsx の　sheet　ABFarmの情報を')
    print('基準日における所属牛（転入牛move-in)と転出牛(move-out)の情報に分け、')
    print('2枚のsheet ABFarmin, ABFarmout を作成する')
    print('----------------------------------------------------------2024/1/2　by jicc---------')
    
    
"""
fpyCowsHistoryManualfrmpdf:                        マニュアル
ｖ1.0
2022/1/11
@author: jicc
"""
def fpyCowsHistoryManualfrmpdf():
    
    print('-----CowsHistoryManual from pdffile--------------------------------------------v1.01-------')
    print(' ')
    print(' \"牛の個体情報検索サービス-個体識別番号の検索\"から個体の異動情報を検索し、')
    print('保存したpdffilesをcsvfileを介してExcelファイルにリスト化する。 ')
    print(' ')
    print('1.ディレクトリ内(..//CowsHistory)の特定の拡張子(.pdf)を持つファイルを見つけcsvfile に変換する')
    print('	MH_???_yyyymmdd.pdf -> ****.csv テーブル部分のデータ抽出')
    print('	MH_???_yyyymmdd.pdf -> ".\\pdforg\\"へ移行')
    print('   PS> python ps_fpyspdfindirtocsv_args.py Ext Path bckPath')
    print(' Ext: \.pdf, Path: .\\(カレントディレクトリ), bckPath: .\\pdforg')
    print(' ')
    print('2.フォルダー内の個体履歴org(csv)をCowsHistory.csvに変更する')
    print('変更後orgcsvfile を　別フォルダー(./csvorg)に移動する')
    print('MH_???_yyyymmdd.csv -> MH_???_yyyymmddH.csv')
    print('	MH_???_yyyymmdd.csv -> ".\\csvorg\\"へ移行')
    print('   PS> ps_fpyymd_csvtocowshistory_csv_args.py Ext Path bckPath')
    print(' Ext: \.csv, Path: .\\(カレントディレクトリ), bckPath: .\\csvorg')
    print(' ')
    print('3.フォルダー内の???H.csv)を???CowsHistory.xlsxに移動する')
    print('移動後???H.csvを　別フォルダー(./csvhistory)に移動する')
    print('	MH_???_yyyymmddH.csv -> ".\\csvhistory\\"へ移行')
    print('   PS> ps_fpyhistory_csvto_xlsx_args.py Ext Path bckPath wbN sheetN')
    print(' Ext: \.csv, Path: .\\(カレントディレクトリ), bckPath: .\\csvhistory')
    print(' wbN: ..\\KT_CowsHistory.xlsx, sheetN:KTFarm')
    print(' ')
    print('4.??_CowsHistory.xlsx\/??Farm の　str\"yyyy\/mm\/dd\"を')
    print('datetimeに変換する')
    print('   PS> ps_fpyxlstrymdtodatetime_args.py wbN sheetN　col')
    print(' wbN: ..\\KT_CowsHistory.xlsx, sheetN:KTFarm, col: 3 and 9')
    print('---------------------------------------------------------------2022/3/6 by jicc---------')
    
"""
fpyCowsHistoryManualfrmweb                        マニュアル
ｖ1.0
2022/7/11
@author: jicc
"""
def fpyCowsHistoryManualfrmweb():
    
    print('-----CowsHistoryManual from web site------------------------------------------v1.0-------')
    print(' ')
    print('\"牛の個体情報検索サービス-個体識別番号の検索\"から個体の異動情報を検索し、 ')
    print('Excelファイルにリスト化する。 ')
    print('web -> csvfile ->Excelfile ')
    print(' ')
    print('1. ABFarmの個体リスト(AB_cowslist.xlsx/ABFarm)から、個体識別番号(colum2 idno)によって、')
    print('個体情報+異動情報を検索し、リストにし、idno_ymd.csv fileに保存する')
    print('   PS> ps_fpyindtrsinf_to_csv_args.py wbN sheetN col')
    print(' wbN : AB_cowslist.xlsx, sheetN : cowslist col:2')
    print('牛の個体情報検索サービス-個体識別番号の検索')
    print(' url : https://www.id.nlbc.go.jp/CattleSearch/search/agreement')
    print(' ')
    print('2.フォルダー内のidno_ymd.csvをcowshistory.xlsx/ABFarmに移動する')
    print('移動後idno_ymd.csvを　別フォルダー(./csvhistory)に移動する')
    print('	idno_ymd.csv -> ".\\csvhistory\\"へ移動')
    print('   PS> python ps_fpyhistory_csvto_xlsx_args.py Ext Path bckPath wbN sheetN')
    print(' Ext: \.csv, Path: .\\(カレントディレクトリ), bckPath: .\\csvhistory')
    print(' wbN: (..\\)cowshistory.xlsx, sheetN:ABFarm')
    print(' ')
    print('---------------------------------------------------------------2023/9/26 by jicc---------')
    
    #以下を削除 2023/9/26
    #print('3.cowshistory.xlsx\/ABFarm の　str\"yyyy\/mm\/dd\"を')
    #print('datetimeに変換する')
    #print('   PS> ps_fpyxlstrymdtodatetime_args.py wbN sheetN　col')
    #print(' wbN: ..\\KT_CowsHistory.xlsx, sheetN:KTFarm, col: 3 and 9')
    
"""
fpyCowsHistoryTools:                        tools
ｖ1.0
2022/7/29
v1.01
2024/1/7
@author: jicc
"""
def fpyCowsHistoryTools():
    
    print('-----CowsHistoryTools---------------------------------------------------------v2.01-------')
    print('牛の個体情報検索サービス 個体識別番号の検索から個体の異動情報を検索し、')
    print('Excelファイルにリスト化するための　Tool集')
    print(' ')
    print('#fpytrs_infs_to_xlsx(wbN0, sheetN0, wbN1, sheetN1, colidno1)')
    print('個体リスト AB_cowslist/cowslistのidnoから個体異動情報を検索する')
    print('個体情報リスト cowshistory.xlsx/ABFarmに新規または追加入力する')
    print('   PS> ps_fpytrs_infs_to_xlsx_args.py wbN0 sheetN0 wbN1 sheetN1 colidno1')
    print(' wbN0 : cowshistory.xlsx, sheetN0 : ABFarm, ')
    print(' wbN1 : AB_cowslist.xlsx, sheetN1 : cowslist, colidno1 : 2 (column number of idno1)')
    print(' ')
    print('#fpynewtrs_infs_to_xlsx(wbN0, sheetN0, colidno0, wbN1, sheetN1, colidno1)')
    print('個体リスト AB_cowslist/のidnoから個体異動情報を検索する')
    print('個体情報リスト cowshistory.xlsx/ABFarmにない新しい情報を抽出する')
    print(' cowshistory.xlsx/ABFarmに追加入力する')
    print('   PS> ps_fpynewtrs_infs_to_xlsx_args.py wbN0 sheetN0 colidno0 wbN1 sheetN1 colidno1')
    print('  wbN0 : cowshistory.xlsx, sheetN0 : ABFarm, colidno0 : 2 (column number fo idno0), ')
    print(' wbN1 : AB_cowslist.xlsx, sheetN1 : cowslist, colidno1 : 2 (column number fo idno1)')
    print(' ')
    print('#fpychk_drecords(wbN, sheetN, searchdate)')
    print('Excel個体情報リスト cowshistory/ABFarmの重複データをを削除する')
    print('   PS> ps_fpychk_drecords_args.py wbN sheetN')
    print(' wbN: ..\\AB_cowshistory.xlsx, sheetN:ABFarm')
    print(' ')
    print('#fpydel_d_idNo(wbN, sheetN)')
    print('個体リスト AB_cowslist/ABFarmのidnoの重複データをを削除する')
    print('   PS> ps_fpydel_d_idno_args.py wbN sheetN')
    print(' wbN: ..\\AB_cowslist.xlsx, sheetN:ABFarm')
    print(' ')
    print('#fpymkd_path( path )')
    print('make a directory  at current directory')
    print('カレントディレクトリに　path名のディレクトリが存在しなければ作成する')
    print('   PS> ps_fpymkd_path_args.py path')
    print(' path: .\\csvhistory, .\\bck etc')
    print('---------------------------------------------------------------2024/1/7 by jicc---------')    


"""
fpyCowsHistory_webscrsys:                        tools
ｖ1.0
2024/1/13

@author: jicc
"""
def fpyCowsHistory_webscrsys():
    
    print('-----CowsHistory_webscrsys---------------------------------------------------------v2.01-------')
    print('牛の個体情報検索サービス 個体識別番号の検索から個体の異動情報を検索し、')
    print('Excelファイルにリスト化するシステム')
    print(' ')
    print('#fpytrs_infs_to_xlsx(wbN0, sheetN0, wbN1, sheetN1, colidno1)')
    print('個体リスト AB_cowslist/cowslistのidnoから個体異動情報を検索する')
    print('個体情報リスト cowshistory.xlsx/ABFarmに新規または追加入力する')
    print('   PS> ps_fpytrs_infs_to_xlsx_args.py wbN0 sheetN0 wbN1 sheetN1 colidno1')
    print(' wbN0 : cowshistory.xlsx, sheetN0 : ABFarm, ')
    print(' wbN1 : AB_cowslist.xlsx, sheetN1 : cowslist, colidno1 : 2 (column number of idno1)')
    print(' ')
    print('#fpychk_drecords(wbN, sheetN)')
    print('Excel個体情報リスト cowshistory/ABFarmの重複データをを削除する')
    print('   PS> ps_fpychk_drecords_args.py wbN sheetN')
    print(' wbN: ..\\AB_cowshistory.xlsx, sheetN:ABFarm')
    print(' ')
    print('#fpysep_outfrmin( wbN, sheetN, coln, ncol, index, name, bdate )')
    print('   PS> python ps_fpysep_outfrmin_args.py wbN sheetN coln ncol index name bdate')
    print(' wbN : ..\\AB_cowshistory.xlsx, sheetN : ABFarm, coln : 2, ncol : 12,')
    print('  index : 10, name :  AB Farm, bdate : yyyy/mm/dd')
    print('separate move-out cows from move-in ')
    print('異動情報のExcelfile: AB_cowshistory.xlsx の　sheet　ABFarmの情報を')
    print('基準日における所属牛（転入牛move-in)と退出牛(転出牛move-out)の情報に分け、')
    print('2枚のsheet ABFarmin, ABFarmout を作成する')
    print('---------------------------------------------------------------2024/1/13 by jicc---------')  