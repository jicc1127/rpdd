# -*- coding: utf-8 -*-
import openpyxl
import csv
#import datetime

#fpyopenxl#
"""
fpyopenxl(wbN, sheetN):
    Excelfile wbN.xlsx　sheet sheetN Open 
    v1.00
    2022/1/5
    @author: jicc
    returnがリストでうまく使用できるか不明?2022/02/11
    list[]に出力すれば使用可だが、普通にopenするのとどう違うか疑問
    2022/2/17
"""
def fpyopenxl(wbN, sheetN):
    """
    Excelfile wbN.xlsx　sheet sheetN Open

    Parameters
    ----------
    wbN : str
        ExcelFile Name   ex.MH_CowHistory.xlsx
    sheetN : str
        sheet name

    Returns
    -------
    list : [wb, sheet]

    """
    
    #import openpyxl
    
    wb = openpyxl.load_workbook(wbN)
    sheet = wb[sheetN]
    return [wb, sheet]

#fpyopencsv_robj#
"""
fpyopencsv_robj:
    csvfile Open for Reader object
    v1.00
    2022/1/5
    @author: jicc
"""
def fpyopencsv_robj(csvN):
    """
    csvfile Open for Reader object

    Parameters
    ----------
    csvN : str
        csvFile Name   ex.MH_??_History.csv

    Returns
    -------
    None.

    """
    #import csv
    
    #filename = csvN.split('.')
    #filename = filename[0]  #拡張子を削除したfilename
    
    filename_file = open(csvN)     #csvfile open
    filename_reader = csv.reader(filename_file)       #get Reader object
    
    
    return filename_reader
    
    
#fpyopencsv_rdata#
"""
fpyopencsv_rdata:
    csvfile Open for Reader data
    v1.00
    2022/1/5
    @author: jicc
"""
def fpyopencsv_rdata(csvN):
    """
    csvfile Open for Reader data

    Parameters
    ----------
    csvN : str
        csvFile Name   ex.MH_??_History.csv

    Returns
    -------
    filename_data : list's list

    """
    #import csv
    
    #filename = csvN.split('.')
    #filename = filename[0]  #拡張子を削除したfilename
    
    filename_file = open(csvN)     #csvfile open
    filename_reader = csv.reader(filename_file)       #get Reader object
    filename_data = list(filename_reader)             #list's list
    
    return filename_data
    
"""
fpyopencsv_w:
    csvfile Open for Writer
    v1.00
    2022/1/7
    @author: jicc
"""
def fpyopencsv_w(csvN):
    """
    csvfile Open for Writer

    Parameters
    ----------
    csvN : str
        csvFile Name   ex.MH_??_History.csv

    Returns
    -------
    None.

    """
    #import csv
    
    output_file = open(csvN, 'w', newline='')       #csvfile open
    output_writer = csv.writer(output_file)       #get Reader object
     
    return output_writer


#fpygetCell_value#
"""
fpygetCell_value: get value from the target Cell
v1.00
2022/2/4

@author: inoue
"""
def fpygetCell_value(sheet, r, col):
    """
    get value from the target Cell on an Excelsheet

    Parameters
    ----------
    sheet : worksheet
        sheetBLV
    r : int
        row
    col : int
        column

    Returns
    -------
    value

    """

    value = sheet.cell(row=r, column=col).value
    return value


#fpyinputCell_value#
"""
fpyinputCell_value: input value to the target Cell
v1.00
2022/2/4

@author: inoue
"""
def fpyinputCell_value(sheet, r, col, vl):
    """
    input value to the target Cell

    Parameters
    ----------
    sheet : worksheet
        sheetBLV
    r : int
        row
    col : int
        column
    vl : type of value
    
    Returns
    -------
    None.

    """

    sheet.cell(row=r, column=col).value = vl    
 
#fpyifNone_inputCell_value#
"""
fpyifNone_inputCell_value:if Cellvalue is None,  input value to the Cell
v1.00
2022/2/5

@author: inoue
"""
def fpyifNone_inputCell_value(sheet, r, col, vl):
    """
    input value to the target Cell

    Parameters
    ----------
    sheet : worksheet
        sheetBLV
    r : int
        row
    col : int
        column
    vl : type of value
    
    Returns
    -------
    None.

    """
    Cellvalue = sheet.cell(row=r, column=col).value
    if Cellvalue == None:
        sheet.cell(row=r, column=col).value = vl    

#fpyidNo_9to10#
"""
fpyidNo_9to10 : ９～10桁耳標の数値を文字列として、
    9桁の耳標に1桁目に０を加え10桁とする
ｖ1.0
2021/4/29
@author: jicc
"""
def fpyidNo_9to10( wbN, sheetN, col ):
    """
    Parameters
    ----------
    wbN : 　str  
        対象となるExcelファイル名　拡張子.xlsxをつけ、　''でくくる。
    sheetN : str
        対象となるsheet
    col : int
        変更する10桁耳標の列

    Returns
    -------
    None.

    """
       
    #import openpyxl
    #import datetime
    
    wb = openpyxl.load_workbook(wbN)
    sheet = wb[sheetN]   #wb.get_sheet_by_name(sheetN)
    max_row = sheet.max_row
        
    for row_num in range(2, max_row + 1):     #先頭行をスキップ
        
        idNo = sheet.cell(row=row_num, column=col).value
        idNo = str(idNo)
        if len(idNo) == 9:
            sheet.cell(row=row_num, column=col).value = '0' + idNo 
        else:
            sheet.cell(row=row_num, column=col).value = idNo 
              
    
    wb.save(wbN)
    
#fpyidNo_9to10_s#
"""
fpyidNo_9to10_s : ９～10桁耳標の数値を文字列として、
    9桁の耳標に1桁目に０を加え10桁とする
    sheet version
ｖ1.0
2021/4/29
@author: jicc
"""
def fpyidNo_9to10_s( sheet, col ):
    """
    ９～10桁耳標の数値を文字列として、
    9桁の耳標に1桁目に０を加え10桁とする
    sheet version
    Parameters
    ----------
    sheet : worksheet.worksheet.Worksheet
         worksheet object　　　
    col : int
        変更する10桁耳標の列

    Returns
    -------
    sheet : worksheet.worksheet.Worksheet
         worksheet object

    """
       
    #import openpyxl
    #import datetime
    
    max_row = sheet.max_row
        
    for row_num in range(2, max_row + 1):     #先頭行をスキップ
        
        idNo = sheet.cell(row=row_num, column=col).value
        idNo = str(idNo)
        if len(idNo) == 9:
            sheet.cell(row=row_num, column=col).value = '0' + idNo 
        else:
            sheet.cell(row=row_num, column=col).value = idNo 
              
    
    return sheet
    
#fpyNewSheet#
"""
fpyNewSheet : Excelbookに
sheet　'columns'と同じ sheet　'scolN'を作成する。
ｖ1.01
2022/5/3

@author: jicc

"""
def fpyNewSheet(wbN, sheetN, scolN, r):
    """
    Excelbookに sheet 'scolN' r行目の'columns'を1行目に配置した sheet'sheetN'を作成する。
    *sheet 'columns'(列名を記入したシート) を作成しておく
    Parameters
    ----------
    wbN : 　str          
        sheetを作成するワークブック
    sheetN : str　　　　　　シート名:"????" 
        作成するシート
    scolN : str         シート名: "columns"
        参照するシート
	r : int		r行目 作成するcolumn行
    Returns
    -------
    None.

    """
    #import openpyxl
    
    wb = openpyxl.load_workbook(wbN)
    #sheetN = wb[sheetN]
    wb.create_sheet(title=sheetN, index=0)
    sheet = wb[sheetN]
    scol = wb[scolN]
    
    maxcol = scol.max_column #sheet columnの最終列
    
    for i in range(1,maxcol+1):
        sheet.cell(row=r, column=i).value = scol.cell(row=1, column=i).value
    
     
    wb.save(wbN)
    
#fpyNewSheet_w############################################################
"""
fpyNewSheet : Excelbookに
sheet　'columns'と同じ sheet　'scolN'を作成する。
workbook version
ｖ1.0
2024/3/29

@author: jicc

"""
def fpyNewSheet_w(wb, sheetN, scolN, r):
    """
    Excelbookに sheet 'scolN' r行目の'columns'を1行目に配置した sheet'sheetN'を作成する。
    *sheet 'columns'(列名を記入したシート) を作成しておく
    workbook version
    Parameters
    ----------
    wb : 　workbook.workbook.workbook          
        workbook objevt
    sheetN : str　　　　　　シート名:"????" 
        作成するシート
    scolN : str         シート名: "columns"
        参照するシート
	r : int		r行目 作成するcolumn行
    Returns
    -------
    sheet : worksheet.worksheet.worksheet
        worksheet object

    """
    
    wb.create_sheet(title=sheetN, index=0)
    sheet = wb[sheetN]
    scol = wb[scolN]
    
    maxcol = scol.max_column #sheet columnの最終列
    
    for i in range(1,maxcol+1):
        sheet.cell(row=r, column=i).value = scol.cell(row=1, column=i).value
    
     
    return sheet

#fpychgSheetTitle##########################################################
"""
fpychgSheetTitle      :change ExcelSheet's title
v1.0
2022/3/30

@author: inoue
"""
def fpychgSheetTitle(wbN, sheetN, sheetN1):
    """
    change the sheet's title

    Parameters
    ----------
    wbN : str
        Excelfile to check double data  '??_CowsHistory.xlsx'
    sheetN : str
        元のシート名  : 'KTFarm'
    sheetN1 : str
        変更名      : 'KTFarmorg' 

    Returns
    -------
    None.

    """
    wbobj = fpyopenxl(wbN, sheetN)
    wb = wbobj[0]
    sheet = wbobj[1]
    sheet.title = sheetN1
    wb.save(wbN)
    
#fpysheet_copy############################################################    
"""
fpysheet_copy : copy a worksheet with another sheetname
    ｖ1.0
2022/9/14
@author: jicc
"""

#! python3
def fpysheet_copy( wbN, sheetN, sheetN_ ):
    """
    
    copy a worksheet with another sheetname
    
    Parameters
    ----------
    wbN : 　str  
        対象となるExcelファイル名　ex. KT_RPDorg.xlsx
    sheetN : str
        対象となるsheet ex. yyyymmdd
    sheetN_ : str
        コピーしたsheet ex. yyyymmddout
 
    Returns
    -------
    None.

    """
    #import openpyxl
       
    wb = openpyxl.load_workbook(wbN)
    sheet = wb[sheetN]           
    
    sheet_ = wb.copy_worksheet(sheet)
    sheet_.title = sheetN_
    
    # シートを先頭へ移動
    wb.move_sheet(sheet_, offset=-wb.index(sheet_))
    
    # 先頭のシートを再度選択状態にする
    #wb.active = 0
        
    wb.save(wbN)

#fpysheet_copy_ws############################################################    
"""
fpysheet_copy : copy a worksheet with another sheetname
    ｖ1.0
    2022/9/14
    workbook sheet version
    v1.0
    2024/6/5
    @author: jicc
"""

#! python3
def fpysheet_copy_ws( wb, sheet, sheetN_ ):
    """
    
    copy a worksheet with another sheetname
    
    Parameters
    ----------
    wb : workbook.workbook.Workbook
         workbook object
    sheet : worksheet.worksheet.Worksheet
         worksheet object
    sheetN_ : str
        コピーしたsheetのsheet名 ex. yyyymmddout
 
    Returns
    -------
    sheet_ : worksheet.worksheet.Worksheet
         worksheet object

    """
    sheet_ = wb.copy_worksheet(sheet)
    sheet_.title = sheetN_
    
    # シートを先頭へ移動
    wb.move_sheet(sheet_, offset=-wb.index(sheet_))
    
    # 先頭のシートを再度選択状態にする
    #wb.active = 0
        
    return sheet_ 

    
#fpycol_blk_rowslist#########################################################
"""
fpycol_blk_rowslist : rows'list column data is blank
    ｖ1.0
2022/9/14
@author: jicc
"""

#! python3
def fpycol_blk_rows_list( wbN, sheetN, col ):
    """
    rows'list column data is blank
    
    Parameters
    ----------
    wbN : 　str  
        対象となるExcelファイル名　ex. KT_RPDorg.xlsx
    sheetN : str
        対象となるsheet ex. yyyymmdd
    col : int
        column number of lastAI_date
 
    Returns
    -------
    None.

    """
    #import openpyxl
    #import chghistory
    #import datetime
    #import kt_ai
    
    wb = openpyxl.load_workbook(wbN)
    sheet = wb[sheetN]           

    max_row = sheet.max_row
    rows_list = []
    for row_num in range(2, max_row + 1): #先頭行をスキップ
        clmn_data = sheet.cell(row=row_num, column=col).value
        if clmn_data == None:
            rows_list.append(row_num)
    
    return rows_list


#fpycol_blk_rowslist_s#########################################################
"""
fpycol_blk_rowslist_s : rows'list column data is blank
sheet version
    ｖ1.0
2022/9/14
@author: jicc
"""

#! python3
def fpycol_blk_rows_list_s( sheet, col ):
    """
    rows'list column data is blank
    sheet version
    
    Parameters
    ----------
    sheet : worksheet.worksheet.Worksheet
         worksheet object
    col : int
        column number of lastAI_date
 
    Returns
    -------
    None.

    """

    max_row = sheet.max_row
    rows_list = []
    for row_num in range(2, max_row + 1): #先頭行をスキップ
        clmn_data = sheet.cell(row=row_num, column=col).value
        if clmn_data == None:
            rows_list.append(row_num)
    
    return rows_list

#fpycol_nonblk_rowslist######################################################
"""
fpycol_nonblk_rowslist : rows'list column data is not blank
    ｖ1.0
2022/9/14
@author: jicc
"""

#! python3
def fpycol_nonblk_rows_list( wbN, sheetN, col ):
    """
    rows'list column data is not blank
    
    Parameters
    ----------
    wbN : 　str  
        対象となるExcelファイル名　ex. KT_RPDorg.xlsx
    sheetN : str
        対象となるsheet ex. yyyymmdd
    col : int
        column number of lastAI_date
 
    Returns
    -------
    None.

    """
    #import openpyxl
    #import chghistory
    #import datetime
    #import kt_ai
    
    wb = openpyxl.load_workbook(wbN)
    sheet = wb[sheetN]           

    max_row = sheet.max_row
    rows_list = []
    for row_num in range(2, max_row + 1): #先頭行をスキップ
        clmn_data = sheet.cell(row=row_num, column=col).value
        if clmn_data != None:
            rows_list.append(row_num)
    
    return rows_list

#fpycol_nonblk_rowslist_s######################################################
"""
fpycol_nonblk_rowslist_s : rows'list column data is not blank
sheet version
    ｖ1.0
2022/9/14
@author: jicc
"""

#! python3
def fpycol_nonblk_rows_list_s( sheet, col ):
    """
    rows'list column data is not blank
    sheet version
    
    Parameters
    ----------
    sheet : worksheet.worksheet.Worksheet
         worksheet object
    col : int
        column number of lastAI_date
 
    Returns
    -------
    None.

    """

    max_row = sheet.max_row
    rows_list = []
    for row_num in range(2, max_row + 1): #先頭行をスキップ
        clmn_data = sheet.cell(row=row_num, column=col).value
        if clmn_data != None:
            rows_list.append(row_num)
    
    return rows_list

#fpycol_cellv_s_rows_list#########################################################
"""
fpycol_cellv_s_rows_list : rows'list column data is a cell value
    ｖ1.0
2022/9/19
@author: jicc
"""

#! python3
def fpycol_cellv_s_rows_list( wbN, sheetN, col, cellv ):
    """
    rows'list column data is a cell value
    
    Parameters
    ----------
    wbN : 　str  
        対象となるExcelファイル名　ex. KT_RPDorg.xlsx
    sheetN : str
        対象となるsheet ex. yyyymmdd
    col : int
        column number of lastAI_date
    cellv : int, str, etc.
        cellvalue in column col 
 
    Returns
    -------
    None.

    """
    import openpyxl
    #import chghistory
    #import datetime
    #import kt_ai
    
    wb = openpyxl.load_workbook(wbN)
    sheet = wb[sheetN]           

    rows_list = []
    for row_num in range(2, sheet.max_row + 1): #先頭行をスキップ
        clmn_data = sheet.cell(row=row_num, column=col).value
        if clmn_data == cellv:
            rows_list.append(row_num)
    
    return rows_list

#fpycol_cellv_s_rows_list_s#########################################################
"""
fpycol_cellv_s_rows_list_s : rows'list column data is a cell value
    ｖ1.0
2022/9/19
@author: jicc
"""

#! python3
def fpycol_cellv_s_rows_list_s( sheet, col, cellv ):
    """
    rows'list column data is a cell value
    sheet version
    
    Parameters
    ----------
    sheet : worksheet.worksheet.Worksheet
         worksheet object
    col : int
        column number of lastAI_date
    cellv : int, str, etc.
        cellvalue in column col 
 
    Returns
    -------
    None.

    """
 
    rows_list = []
    for row_num in range(2, sheet.max_row + 1): #先頭行をスキップ
        clmn_data = sheet.cell(row=row_num, column=col).value
        if clmn_data == cellv:
            rows_list.append(row_num)
    
    return rows_list

#fpycol_cellvs_rows_list#########################################################
"""
fpycol_cellvs_rows_list : 
    rows'list column data is a element of a cell values'list
    cell values'list version
    ｖ1.0
    2022/9/19
    @author: jicc
"""

#! python3
def fpycol_cellvs_rows_list( wbN, sheetN, col, cellvs_lst ):
    """
    rows'list column data is a element of a cell values'list
    cell values'list version
    
    Parameters
    ----------
    wbN : 　str  
        対象となるExcelファイル名　ex. KT_RPDorg.xlsx
    sheetN : str
        対象となるsheet ex. yyyymmdd
    col : int
        column number of lastAI_date
    cellvs_lst : list
        cellvalues'list in column col ex. ['a','b','c'] 
 
    Returns
    -------
    None.

    """
    import openpyxl
     
    wb = openpyxl.load_workbook(wbN)
    sheet = wb[sheetN]           
    lcellvs_lst = len(cellvs_lst)
    
    rows_list = []
    for i in range(0, lcellvs_lst):
        for row_num in range(2, sheet.max_row + 1): #先頭行をスキップ
            clmn_data = sheet.cell(row=row_num, column=col).value
            if clmn_data == cellvs_lst[i]:
                rows_list.append(row_num)
    rows_list.sort()
    
    return rows_list

#fpycol_cellvs_rows_list_s#########################################################
"""
fpycol_cellvs_rows_list_s : 
    rows'list column data is a element of a cell values'list
    cell values'list version
    sheet version
    ｖ1.0
    2022/9/19
    @author: jicc
"""
#! python3
def fpycol_cellvs_rows_list_s( sheet, col, cellvs_lst ):
    """
    rows'list column data is a element of a cell values'list
    cell values'list version
    sheet version
    
    Parameters
    ----------
    sheet : worksheet.worksheet.Worksheet
         worksheet object
    col : int
        column number of lastAI_date
    cellvs_lst : list
        cellvalues'list in column col ex. ['a','b','c'] 
 
    Returns
    -------
    None.

    """
    lcellvs_lst = len(cellvs_lst)
    
    rows_list = []
    for i in range(0, lcellvs_lst):
        for row_num in range(2, sheet.max_row + 1): #先頭行をスキップ
            clmn_data = sheet.cell(row=row_num, column=col).value
            if clmn_data == cellvs_lst[i]:
                rows_list.append(row_num)
    rows_list.sort()
    
    return rows_list

#fpyxllst_rows_list#########################################################
"""
fpyxllst_rows_list : get rows'list from a list of a Excel sheet
    ｖ1.0
    2022/9/19
    @author: jicc
    
"""
#! python3
def fpyxllst_rows_list( wbN, sheetN ):
    """
    get rows'list from a list of a Excel sheet
    
    Parameters
    ----------
    wbN : 　str  
        対象となるExcelファイル名　ex. KT_RPDorg.xlsx
    sheetN : str
        対象となるsheet ex. yyyymmdd
  
    Returns
    -------
    rows_list : list

    """
    import openpyxl
    
    wb = openpyxl.load_workbook(wbN)
    sheet = wb[sheetN]           
     
    rows_list = []
    for row_num in range(2, sheet.max_row+1): #先頭行をスキップ

        rows_list.append(row_num)
        
    return rows_list

#fpyxllst_rows_list_s#########################################################
"""
fpyxllst_rows_list_s : get rows'list from a list of a Excel sheet
                       sheet version 
    ｖ1.0
    2022/9/19
    @author: jicc
    
"""
#! python3
def fpyxllst_rows_list_s( sheet ):
    """
    get rows'list from a list of a Excel sheet
    sheet version
    
    Parameters
    ----------
    sheet : worksheet.worksheet.Worksheet
         worksheet object
  
    Returns
    -------
    rows_list : list

    """
 
    rows_list = []
    for row_num in range(2, sheet.max_row+1): #先頭行をスキップ

        rows_list.append(row_num)
        
    return rows_list

#fpyrm_smelems_frm_list#####################################################
"""
fpyrm_smelems_frm_list : remove some elements from a list
    v1.0
    2022/9/19
    @author: inoue
    
"""
def fpyrm_smelems_frm_list(lst0, lst1):
    """
    remove some elements from a list

    Parameters
    ----------
    lst0 : list
        list to be removed some elements
    lst1 : list
        elements'list to remove
    Returns
    -------
    lst0 : list
        
    """
    llst1 = len(lst1)
    
    for i in range(0, llst1):
        
        lst0.remove(lst1[i])
        
    return lst0

#fpydelete_rows##############################################################
"""
fpydelete_rows : delete rows in rows_list
    ｖ1.0
2022/9/14
@author: jicc
"""

#! python3
def fpydelete_rows( wbN, sheetN, rows_list ):
    """
    delete rows in rows_list
    
    Parameters
    ----------
    wbN : 　str  
        対象となるExcelファイル名　ex. KT_RPDorg.xlsx
    sheetN : str
        対象となるsheet ex. yyyymmdd
    rows_list : list
        rows list to delete
 
    Returns
    -------
    None.

    """
    #import openpyxl
    #import chghistory
    #import datetime
    #import kt_ai
    
    wb = openpyxl.load_workbook(wbN)
    sheet = wb[sheetN]           

    for row_num in reversed(rows_list): #list 降順
        
        sheet.delete_rows(row_num)
    
    wb.save(wbN)
    
#fpydelete_rows_s##############################################################
"""
fpydelete_rows_s : delete rows in rows_list
sheet version
    ｖ1.0
2022/9/14
@author: jicc
"""

#! python3
def fpydelete_rows_s( sheet, rows_list ):
    """
    delete rows in rows_list
    sheet version
    Parameters
    ----------
    
    sheet : worksheet.worksheet.Worksheet
         worksheet object
    rows_list : list
        rows list to delete
 
    Returns
    -------
    sheet : worksheet.worksheet.Worksheet

    """

    for row_num in reversed(rows_list): #list 降順
        
        sheet.delete_rows(row_num)
    
    return sheet

#fpynumber_rows#############################################################
"""
fpynumber_rows : number rows in a column
    ｖ1.0
2022/9/16
@author: jicc
"""

#! python3
def fpynumber_rows( wbN, sheetN, col ):
    """
    number rows in a column
    
    Parameters
    ----------
    wbN : 　str  
        対象となるExcelファイル名　ex. KT_RPDorg.xlsx
    sheetN : str
        対象となるsheet ex. yyyymmdd
    col : int
        column number to number rows
 
    Returns
    -------
    None.

    """
    import openpyxl
   
    wb = openpyxl.load_workbook(wbN)
    sheet = wb[sheetN]           
    max_row = sheet.max_row
    
    for row_num in range(2, max_row + 1): #先頭行をスキップ
        
        sheet.cell(row=row_num, column=col).value = row_num - 1
    
    wb.save(wbN)

#fpynumber_rows_s#############################################################
"""
fpynumber_rows : 
    number rows in a column
    sheet version
    ｖ1.0
2022/10/2
@author: jicc
"""

#! python3
def fpynumber_rows_s( sheet, col ):
    """
    number rows in a column
    sheet version
    
    Parameters
    ----------
    sheet : worksheet.worksheet.Worksheet
         worksheet object
    col : int
        column number to number rows
 
    Returns
    -------
    None.

    """
        
    max_row = sheet.max_row
    
    for row_num in range(2, max_row + 1): #先頭行をスキップ
        
        sheet.cell(row=row_num, column=col).value = row_num - 1
   

#fpylstNo_to_rowNo###########################################################
"""
fpylstNo_to_rowNo : 
    get the row number of a list number element (list[0])
    2022/9/23
    v1.0
    @author: inoue
    
"""
def fpylstNo_to_rowNo(wbN, sheetN, col, lstNo):
    """
    get the row number of a list number element (list[0])

    Parameters
    ----------
    wbN : str
        ExcelFile Name   ex."KT_RPDorg.xlsx"
    sheetN : str
        sheet name       ex."yyyymmdd"
    col : int
       Excel list number's column  ex. 1 : AI_id
    lstNo: int
       listNo if a list   

    Returns
    -------
   row_num : int
       row number

    """

    #import fmstls
     
    
    wbobj = fpyopenxl(wbN, sheetN)   #get Worksheet object
    #wb = wbobj[0]
    sheet = wbobj[1]
       
    for row_num in range(2, sheet.max_row+1):
        line_id = fpygetCell_value(sheet, row_num, col)
        #get Excel's line_id No  column 1 (col)
        if line_id == lstNo:
            return row_num
            
            break
        else:
            continue

#fpylstNo_to_rowNo_s###########################################################
"""
fpylstNo_to_rowNo_s : 
    get the row number of a list number element (list[0])
    sheet version
    2022/9/23
    v1.0
    @author: inoue
    
"""
def fpylstNo_to_rowNo_s(sheet, col, lstNo):
    """
    get the row number of a list number element (list[0])

    Parameters
    ----------
    sheet : worksheet.worksheet.Worksheet
         worksheet object
    col : int
       Excel list number's column  ex. 1 : AI_id
    lstNo: int
       listNo if a list   

    Returns
    -------
   row_num : int
       row number

    """
    for row_num in range(2, sheet.max_row+1):
        line_id = fpygetCell_value(sheet, row_num, col)
        #get Excel's line_id No  column 1 (col)
        if line_id == lstNo:
            return row_num
            
            break
        else:
            continue

#fpy_0nton###################################################################
"""
fpy_0nton : change str '0n' to 'n'
    2023/8/12
    v1.0
    @author: jicc
    
"""
def fpy_0nton( mn ):
    """
    change str '0n' to 'n'

    Parameters
    ----------
    mn : str
       string of two figures

    Returns
    -------
    one figure or two figures

    """
    if mn[0] == '0':
        mn = mn[1]
    else:
        mn = mn
    
    return mn

#fpy_nto0n###################################################################
"""
fpy_nto0n : change str 'n' to '0n'
    2024/2/21
    v1.0
    @author: jicc
    
"""
def fpy_nto0n( st ):
    """
    change str 'n' to '0n'

    Parameters
    ----------
    st : str
       string of one or two figures

    Returns
    -------
    str : two figures mn or 0n

    """
    lst = len(st)
    if lst == 1:
        st = '0' + st
    else:
        st = st
    
    return st

#fpyymd_0mtom_0dtod_######################################################
"""
fpyymd_0mtom_0dtod_str : change str yyyy/0m/0d to datetime yyyy/m/d
    2023/10/7
    v1.0
    @author: inoue

"""
def fpyymd_0mtom_0dtod_( yyyy_mm_dd ):
    """
    change str yyyy/0m/0d to str yyyy/m/d

    Parameters
    ----------
    yyyy_mm_dd : str
        
    Returns
    -------
    str yyyy_mm_dd_

    """
    import datetime
    date = yyyy_mm_dd.split('/')
    date[1] = fpy_0nton( date[1] )
    date[2] = fpy_0nton( date[2] )
    yyyy_mm_dd_ = '/'.join(date)
    yyyy_mm_dd_ = datetime.datetime.strptime( yyyy_mm_dd_, '%Y/%m/%d')
    return yyyy_mm_dd_


#fpyymd_0mtom_0dtod#########################################################
"""
fpyymd_0mtom_0dtod : change str yyyy/0m/0d to datetime yyyy/m/d
    2023/8/12
    v1.0
    @author: inoue

"""
def fpyymd_0mtom_0dtod( wbN, sheetN, col ):
    """
    change str yyyy/0m/0d to datetime yyyy/m/d

    Parameters
    ----------
    wbN : str
        ExcelFile Name   ex."KT_cowslist.xlsx"
    sheetN : str
        sheet name       ex."tmp"
    col : int
       Excel list number's column  ex. 7 : birthday

    Returns
    -------
    None.

    """
    import openpyxl
    import datetime
    import fmstls
    
    wb = openpyxl.load_workbook(wbN)
    sheet = wb[sheetN]
    
    for row_num in range(2, sheet.max_row + 1): #先頭行をスキップ
        
        clmn_data = sheet.cell(row=row_num, column=col).value
        #birthday
        if type(clmn_data) == str:
            date = clmn_data.split('/')
            date[1] = fmstls.fpy_0nton( date[1] )
            date[2] = fmstls.fpy_0nton( date[2] )
            clmn_data = '/'.join(date)
            clmn_data = datetime.datetime.strptime( clmn_data, '%Y/%m/%d')
            sheet.cell(row=row_num, column=col).value = clmn_data
        else:
            continue
        
    wb.save(wbN)

#fpystrdate_to_yyyymmdd######################################################
"""
fpystrdate_to_yyyymmdd : 
    change str date yyyy/mm/dd to str yyyymmdd
    v1.0
    2024/2/25
    @author: inoue
    
"""
def fpystrdate_to_yyyymmdd( date ):
    """
    change str date yyyy/mm/dd to str yyyymmdd

    Parameters
    ----------
    date : str
        yyyy/mm/dd
    
    Returns
    -------
    str : yyyymmdd

    """
    strd = date.split('/')
    #strdate yyyy/mm/dd を '/' で分離
    #strd = [yyyy, mm, dd]
    print(strd)
    
    yyyy = strd[0] #year yyyy
    
    lmm = len(strd[1]) #month mm の文字数
    if lmm == 1:
        mm = '0' + strd[1] #add '0' first
    else: #lmm == 2: 
        mm = strd[1] #without change
        
    ldd = len(strd[2])
    if ldd == 1:
        dd = '0' + strd[2]
    else: #ldd == 2:
        dd = strd[2]
    
    yyyymmdd = yyyy + mm + dd
    
    return yyyymmdd

#fpyyyyymmdd_to_strdate######################################################
"""
fpyyyyymmdd_to_strdate : 
    change str yyyymmdd to str date yyyy/mm/dd
    v1.0
    2024/3/2
    @author: inoue
    注) mm,dd で 0m,0d と 0 が残ることに注意
"""
def fpyyyyymmdd_to_strdate( yyyymmdd ):
    """
    change str yyyymmdd to str date yyyy/mm/dd

    Parameters
    ----------
    
    str : yyyymmdd

    Returns
    -------
    strdate : str
        yyyy/mm/dd
    

    """
    
    strdate = yyyymmdd[0:4] + '/' + yyyymmdd[4:6] + '/' +  yyyymmdd[6:8]

    
    return strdate

######################################################################
def fpyfmstlsReference():
    
    print('-----fmstlsReference ---------------------------------------------------v1.05------')
    print('**fpyopenxl(wbN, sheetN)')
    print('Excelfile wbN.xlsx　sheet sheetN Open ')
    print('...................................................................................')
    print('**fpyopencsv_robj(csvN)')
    print('csvfile Open for Reader object')
    print('...................................................................................')
    print('**fpyopencsv_rdata(csvN)')
    print('csvfile Open for Reader data')
    print('..................................................................................')
    print('**fpyopencsv_w(csvN)')
    print('csvfile Open for Writer')
    print('....................................................................................')
    print('**fpygetCell_value(sheet, r, col)')
    print('Excelシート上のセルの値を取得する')
    print('get value from the target Cell on an Excelsheet')
    print('....................................................................................')
    print('**fpyinputCell_value(sheet, r, col, vl)')
    print('Excelシート上のセルに値を上書き入力する')
    print('input value to the target Cell on an Excelsheet')
    print('....................................................................................')
    print('**fpyifNone_inputCell_value(sheet, r, col, vl)')
    print('Excelシート上のセルに値がなければ、入力する')
    print('if Cellvalue is None,  input value to the Cell')
    print('....................................................................................')
    print('**fpyidNo_9to10(wbN, sheetN, col)')
    print('9桁耳標を10桁にし、文字列として再入力する')
    print(' wbN:workbooks_name,  sheetN:worksheets_name, col: columns_no')
    print('....................................................................................')
    print('**fpyNewSheet(wbN, sheetN, scolN, r)')
    print('Excelbookに sheet　\'columns\'r行と同じ sheet　\'scolN\'を作成する')
    print(' wbN:workbooks_name,  sheetN:worksheets_name, scolN: columns_sheets_name')
    print('....................................................................................')
    print('**fpyNewSheet_w(wb, sheetN, scolN, r)')
    print('Excelbookに sheet　\'columns\'r行と同じ sheet　\'scolN\'を作成する')
    print('workbook version')
    print(' wb:workbook object,  sheetN:worksheets_name, scolN: columns_sheets_name')
    print('....................................................................................')
    print('**fpychgSheetTitle(wbN, sheetN, sheetN1)')
    print('change ExcelSheet\'s title')
    print('....................................................................................')
    print('**fpysheet_copy( wbN, sheetN, sheetN_ )')
    print('copy a worksheet with another sheetname')
    print('....................................................................................')
    print('**fpysheet_copy_ws( wb, sheet, sheetN_ )')
    print('copy a worksheet with another sheetname')
    print('workbook sheet version')
    print('....................................................................................')
    print('**fpycol_blk_rows_list( wbN, sheetN, col )')
    print('rows\'list column data is blank')
    print('....................................................................................')
    print('**fpycol_blk_rows_list_s( sheet, col )')
    print('rows\'list column data is blank')
    print('sheet version')
    print('....................................................................................')
    print('**fpycol_nonblk_rows_list( wbN, sheetN, col )')
    print('rows\'list column data is not blank')
    print('....................................................................................')
    print('**fpycol_nonblk_rows_list_s( sheet, col )')
    print('rows\'list column data is not blank')
    print('sheet version')
    print('....................................................................................')
    print('**fpycol_cellv_s_rows_list( wbN, sheetN, col, cellv )')
    print('rows\'list column data is a cell value')
    print('....................................................................................')
    print('**fpycol_cellv_s_rows_list_s( sheet, col, cellv )')
    print('rows\'list column data is a cell value')
    print('sheet version')
    print('....................................................................................')
    print('**fpycol_cellvs_rows_list( wbN, sheetN, col, cellvs_lst )')
    print('rows\'list column data is cell values\'list')
    print('cell values\'list version')
    print('....................................................................................')
    print('**fpycol_cellvs_rows_list_s( sheet, col, cellvs_lst )')
    print('rows\'list column data is cell values\'list')
    print('cell values\'list version')
    print('sheet version')
    print('....................................................................................')
    print('**fpyxllst_rows_list( wbN, sheetN )')
    print('get rows\'list from a list of a Excel sheet')
    print('....................................................................................')
    print('**fpyxllst_rows_list_s( sheet )')
    print('get rows\'list from a list of a Excel sheet')
    print('sheet version')
    print('....................................................................................')
    print('**fpyrm_smelems_frm_list(lst0, lst1)')
    print('remove some elements from a list')
    print('....................................................................................')
    print('**fpydelete_rows( wbN, sheetN, rows_list )')
    print('delete rows in rows_list')
    print('....................................................................................')
    print('**fpydelete_rows_s(  sheet, rows_list )')
    print('delete rows in rows_list')
    print('sheet version')
    print('need to save workbook')
    print('....................................................................................')
    print('**fpynumber_rows( wbN, sheetN, col )')
    print('number rows in a column')
    print('....................................................................................')
    print('**fpynumber_rows_s( sheet, col )')
    print('number rows in a column')
    print('sheet version')
    print('....................................................................................')
    print('**fpylstNo_to_rowNo(wbN, sheetN, col, lstNo)')
    print('get the row number of a list number element (list[0])')
    print('....................................................................................')
    print('**fpylstNo_to_rowNo_s(sheet, col, lstNo)')
    print('get the row number of a list number element (list[0])')
    print('sheet version')
    print('....................................................................................')
    print('**fpy_0nton( mn )')
    print('change str \'0n\' to \'n\'')
    print('....................................................................................')
    print('**fpy_nto0n( st )')
    print('change str \'n\' to \'0n\'')
    print('....................................................................................')
    print('**fpyymd_0mtom_0dtod_str ( yyyy_mm_dd )')
    print('change str yyyy/0m/0d to str yyyy/m/d')
    print('....................................................................................')
    print('**fpyymd_0mtom_0dtod ( wbN, sheetN, col )')
    print('change str yyyy/0m/0d to datetime yyyy/m/d')
    print('....................................................................................')
    print('**fpystrdate_to_yyyymmdd( date )')
    print('change str date yyyy/mm/dd to str yyyymmdd')
    print('....................................................................................')
    print('**fpyyyyymmdd_to_strdate( yyyymmdd )')
    print('change str yyyymmdd to str date yyyy/mm/dd')
    print('----------------------------------------------------------2024/6/5 by jicc---------')
    