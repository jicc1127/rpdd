# -*- coding: utf-8 -*-
"""
Tools for a Farm's reproductive data operation
    v1.0
    2024/3/2
    by jicc

"""
import openpyxl
import chghistory
#import datetime
import fmstls
import mh_rpdu

#fpymkst_Cow_Heifer00#######################################################
"""
fpymkst_Cow_Heifer00 : 
    make two sheets for cows and heifers at base date
    v1.0
    2024/2/28
    @author: inoue
    
"""
def fpymkst_Cow_Heifer00( wbN, scolN, colnc, colnh, bdate):
    """
    make two sheets for cows and heifers at base date

    Parameters
    ----------
    wbN : str
        Excelbook's name  : '.\\MH_rpdd.xlsx'
    scolN : str
        sheet's name : 'columns'
    colnc : int
        row no of sheet columns for Cow : 1 
    colnh : int
        row no of sheet columns for Heifer :3
    bdate : str
        base date : 'yyyy/mm/dd'

    Returns
    -------
    None.

    """
    #import openpyxl
    #import chghistory
    #import datetime
    #import fmstls

    wb = openpyxl.load_workbook(wbN) 
    #scol = wb[scolN]
    
    bdate_ = fmstls.fpystrdate_to_yyyymmdd( bdate )
    
    scowN = bdate_ + 'Cow00'
    sheiferN = bdate_ + 'Heifer00'
    
    chghistory.fpymkxlsheet_(wb, scowN, scolN, colnc)
    
    chghistory.fpymkxlsheet_(wb, sheiferN, scolN, colnh)
    
        
    wb.save(wbN) 
    
#fpychgelmv_of_lsts_lst#######################################################
"""
fpychgelmv_of_lsts_lst : 
    change a element value of a lists'list
    v1.0
    2024/3/3
    @author: jicc
"""
def fpychgelmv_of_lsts_lst( lst, index, chgv ):
    """
    change a element value of a lists'list

    Parameters
    ----------
    lst : list
        cowslistyyyymmdd
        基準日(yyyymmdd)のcowslist の検索年月日を基準日yyyy/mm/dd に変更する
    index : int
        list index to change the value
    chgv : the value to change
        yyyy/mm/dd (datetime)
    Returns
    -------
    lst : list
        検索年月日を基準日に変更したcowslist    
    """
    llst = len(lst)
    for i in range(0, llst):
        lst[i][index]=chgv
        
    return lst

#fpysrt_into_Cow_Heiferlst################################################## 
"""
fpysrt_into_Cow_Heiferlst :
    sort cowslist into two lists Heifer and Cow
    v1.0
    2024/3/3
    @author: jicc
    
"""
def fpysrt_into_Cow_Heiferlst(wbN0, sheetN0, ncol0, wbN1, sheetN1, ncol1):
    """
    sort cowslist into two lists Heifer and Cow

    Parameters
    ----------
    wbN0 : str
        Excelfile name : 'AB_cowslist.xlsx'
    sheetN0 : int
        sheet name : 'cowslistyyyymmdd'
    ncol0 : int
        the number of columns of sheet cowslist*'s list 
    wbN1 : str
        Excelfile name : 'AB_calving.xlsx'
    sheetN1 : str
        sheet name : 'calving'
    ncol1 :int
        the number of columns of sheet calving's list

    Returns
    -------
    list : [[Heifer], [Cow]]

    """
    #import openpyxl
    import chghistory
    #import datetime
    import fmstls
    import rpdd
    
    #make a list of cows and heifers at base date : sheet cowslistyyyymmdd
    xlcowslist = chghistory.fpyxllist_to_list(wbN0, sheetN0, ncol0)
    lxlcowslist = len(xlcowslist) # the length of the xlcowslist
    
    #make a list of Farm's calving data
    xlcalving = chghistory.fpyxllist_to_list(wbN1, sheetN1, ncol1)
    
    #get a basedate yyyy/mm/dd from a cowslist's sheet sheetN0 'cowslistyyyymmdd'
    strdate = fmstls.fpyyyyymmdd_to_strdate( sheetN0[-8:] ) 
    #'yyyymmdd'->'yyyy/mm/dd'
    print(strdate)
    #change str(strdate) to datetime(date)
    date = fmstls.fpyymd_0mtom_0dtod_ ( strdate )
    print(type(date))
    print(date)
    #xlcowslist[?][19]
    # (search date 検索年月日)-> (base date 基準日)に変更する
    rpdd.fpychgelmv_of_lsts_lst( xlcowslist, 19, date )
    #for j in range(0, lxlcowslist):
    #    print(xlcowslist[j][19])
    #    j +=1
    
    Heifer = [] #heifers'list defaoult
    Cow = []    #cows'list defaoult
    cowslists = [] # cowslist [Heifer, Cow] default
    for i in range(0,lxlcowslist):
        #個体識別番号を取得
        idNo = xlcowslist[i][1] 
        #　1: 個体識別番号のリスト上index
        idNo_scalving = chghistory.fpylst_to_indlst( xlcalving, idNo, 1 )
        #print(idNo_scalving)             #　1: 個体識別番号のリスト上index
        #idNo の間違いがあると、分娩データなしとなる場合に注意
        lidNo_scalving = len(idNo_scalving)
        if lidNo_scalving == 0:         #noncalving
            Heifer.append(xlcowslist[i])
        else:
            Cow.append(xlcowslist[i])   #calving
        
    cowslists.append(Heifer)
    cowslists.append(Cow)
    
    return cowslists  #[Heifer, Cow]

#fpyinput_cwlstd_to_Cow_Heifer00#############################################
"""
fpyinput_cwlstd_to_Cow_Heifer00 :
    input cowslists'data to xls'two sheets 
    yyyymmddHeifer00 and yyyymmddCow00
    v1.0
    2024/3/3
    add one parameter enofetn(element number of column eartagNo at cowslists)
    v1.01
    2024/6/4
    @author: jicc
　 **fpysrt_into_Cow_Heiferlst の　cowslists が必要    
"""
def fpyinput_cwlstd_to_Cow_Heifer00( cowslists, fstcol, wbN, s0N, s1N, enofetn ):
    """
    input cowslists'data to xls'two sheets 
    yyyymmddCow00 and yyyymmddHeifer00

    Parameters
    ----------
    cowslists : lists' list
        list[[Heifer],[Cow]]
    fstcol : int
       first column number to input data
    wbN : str
        Excelfile name : 'AB_rpdd.xlsx'
    s0N : str
        sheet name : 'yyyymmddHeifer00'
    s1N : str
        sheet name : 'yyyymmddCow00'
    enofetn : int                             *)v1.01
        element number of an eartagNo at cowslists
    Returns
    -------
    None.

    """
    wb = openpyxl.load_workbook(wbN)
    s0 = wb[s0N]
    s1 = wb[s1N]
    
    Heifer = cowslists[0]
    lH = len(Heifer)
    Cow = cowslists[1]
    lC = len(Cow)
    
    for i in range(0,lH):
        #LineNo
        fmstls.fpyinputCell_value(s0, i+2, 1, i+1)
        #diNo 個体識別番号
        fmstls.fpyinputCell_value(s0, i+2, 6, Heifer[i][1])
        #cowcode 牛ｺｰﾄﾞ
        fmstls.fpyinputCell_value(s0, i+2, 4, Heifer[i][enofetn])
        #birthday 生年月日
        fmstls.fpyinputCell_value(s0, i+2, 7, Heifer[i][6])
        #base date 基準日
        fmstls.fpyinputCell_value(s0, i+2, 18, Heifer[i][19])
        
    for j in range(0,lC):
        #LineNo
        fmstls.fpyinputCell_value(s1, j+2, 1, j+1)
        #diNo 個体識別番号
        fmstls.fpyinputCell_value(s1, j+2, 6, Cow[j][1])
        #cowcode 牛ｺｰﾄﾞ
        fmstls.fpyinputCell_value(s1, j+2, 4, Cow[j][enofetn])
        #birthday 生年月日
        fmstls.fpyinputCell_value(s1, j+2, 7, Cow[j][6])
        #base date 基準日
        fmstls.fpyinputCell_value(s1, j+2, 18, Cow[j][19])
        
    wb.save(wbN)

#fpyext_idNo_s_calvinglist####################################################
"""
fpyext_idNo_s_calvinglist : 
    extract an individual calving list of two elements,
    calving_date and parity
    v1.0
    2024/3/14
    @author: jicc
    
"""
def fpyext_idNo_s_calvinglist( idNo, wbN, sN, idNo_coln, clvd_coln, prty_coln ):
    """
    extract an individual calving list of two elements,
    calving_date and parity

    Parameters
    ----------
    idNo : str
        cowidNo 個体識別番号 
    wbN : str
        Excelfile's name : 'AB_calving.xlsx'
    sN : str
        sheet name : 'calving'
    idNo_coln : int
        the column's number of cowidNo at sheet sN 
        sheet sN上の　個体識別番号のある列数
    clvd_coln : int
        the column's number of calving_date at sheet sN
    prty_coln : int
        the column's number of parity at sheet sN

    Returns
    -------
    list : ind_calving_list : [[calving_date, parity],...]

    """
    wb = openpyxl.load_workbook(wbN)
    s = wb[sN]
    
    ind_calving_list = [] #individual calving list  default
                            #[[calving_date, parity],...]
    clvd_prty = []  #[calving_date, parity] default
    
    for i in range(2,s.max_row+1):
        
        idNo_ = fmstls.fpygetCell_value(s, i, idNo_coln)
        
        if idNo_ == idNo:
            #calving_date
            calving_date = fmstls.fpygetCell_value(s, i, clvd_coln)
            clvd_prty.append(calving_date)
            #parity
            parity = fmstls.fpygetCell_value(s, i, prty_coln)
            clvd_prty.append(parity)
            
            ind_calving_list.append(clvd_prty)
            clvd_prty = []
        else:
            continue
        
    ind_calving_list.sort(key = lambda x:x[0]) #, reverse=True
    #lists' listを 分娩日 昇順 でsort lambda関数を利用
    
    return ind_calving_list

#fpyext_idNo_s_calvinglist_s##################################################
"""
fpyext_idNo_s_calvinglist : 
    extract an individual calving list of two elements,
    calving_date and parity
    v1.0
    2024/3/15
    @author: jicc
    
"""
def fpyext_idNo_s_calvinglist_s( idNo, sheet, idNo_coln, clvd_coln, prty_coln ):
    """
    extract an individual calving list of two elements,
    calving_date and parity
    sheet version
    Parameters
    ----------
    idNo : str
        cowidNo 個体識別番号 
    sheet : worksheet.worksheet.Worksheet
         worksheet object
    idNo_coln : int
        the column's number of cowidNo at sheet sN 
        sheet sN上の　個体識別番号のある列数
    clvd_coln : int
        the column's number of calving_date at sheet sN
    prty_coln : int
        the column's number of parity at sheet sN

    Returns
    -------
    list : ind_calving_list : [[calving_date, parity],...]

    """
    #wb = openpyxl.load_workbook(wbN)
    #s = wb[sN]
    
    ind_calving_list = [] #individual calving list  default
                            #[[calving_date, parity],...]
    clvd_prty = []  #[calving_date, parity] default
    
    for i in range(2,sheet.max_row+1):
        
        idNo_ = fmstls.fpygetCell_value(sheet, i, idNo_coln)
        
        if idNo_ == idNo:
            #calving_date
            calving_date = fmstls.fpygetCell_value(sheet, i, clvd_coln)
            clvd_prty.append(calving_date)
            #parity
            parity = fmstls.fpygetCell_value(sheet, i, prty_coln)
            clvd_prty.append(parity)
            
            ind_calving_list.append(clvd_prty)
            clvd_prty = []
        else:
            continue
        
    ind_calving_list.sort(key = lambda x:x[0]) #, reverse=True
    #lists' listを 分娩日 昇順 でsort lambda関数を利用
    
    return ind_calving_list

#fpysel_clvingdate#########################################################
"""
fpysel_clvingdate : 
    select the latest calving date at a base date
    基準日における最新の分娩日を選択する
    v1.0
    2024/3/15
    @author: jicc
    
"""
def fpysel_clvdate( bdate, ind_calving_list ):
    """
    select the latest calving date at a base date
    基準日における最新の分娩日を選択する

    Parameters
    ----------
    bdate : datetime.datetime
        base date
    ind_calving_list : list's list
        idNo's calvinglist  [[calving_date, parity],...]
        個体(idNo)の分娩、産次リスト
    Returns
    -------
    list 
    a list of the latest calving_date and parity at basedate
    [calving_date, parity]

    """
    latest_calving = [] 
    # a list of the latest calving_date and parity at basedate : default
    # 基準日直近の分娩データ（分娩日、産次数)　のリスト : デフォルト　 　                   
    lind_calving_list = len(ind_calving_list)
    #the length of list ind_calving_list
    for i in range(lind_calving_list-1,-1,-1):
        #operate from the last calving to the first calving
        #最終分娩から逆順に作業する
        #listを分娩日降順にソートして最初から作業する方法もあり。
        clvdate = ind_calving_list[i][0] #calving date
        prty = ind_calving_list[i][1]    #parity
        
        if clvdate <= bdate :   #分娩日 <= 基準日ならそれが直近の分娩
            calving_date = clvdate   #the latest calving date 直近の分娩日
            parity = prty            #its parity その産次数
            break
        elif clvdate > bdate:                   
            if i>0:         # 分娩日 > 基準日 なら次の分娩との比較へ
                continue
            elif i==0:      #i=0なら次の分娩がないので基準日以降の初産分娩
                calving_date = ''
                parity = 0
        else:
            calving_date = clvdate   #Nan data etc.
            parity = prty 
            
    #try:    
    latest_calving.append(calving_date)
    latest_calving.append(parity)
    print(latest_calving)
    #except UnboundLocalError:
        #print('local variable \'calving_date\' referenced before assignment')
    
    return latest_calving

#fpyinput_clvdt_into_ymdcow##################################################
"""
fpyinput_clvdt_into_ymdcow : 
    input calving data at base date into sheet yyyymmddCow00
    v1.0
    2024/3/16
    @author: jicc
    
"""
def fpyinput_clvdt_into_ymdcow(wb0N, s0N, wb1N, s1N):
    """
    input calving data at base date into sheet yyyymmddCow00

    Parameters
    ----------
    wb0N : str
        Excelfile's name : 'AB_rpdd.xlsx'
    s0N : str
        sheet name : 'yyyymmddCow00'
    wb1N : str
        Excelfile's name : 'AB_calving.xlsx'
    s1N : str
        sheet name : 'calving'

    Returns
    -------
    None.

    """
    wb0 = openpyxl.load_workbook(wb0N)
    s0 = wb0[s0N]
    
    wb1 = openpyxl.load_workbook(wb1N)
    s1 = wb1[s1N]
    
    for i in range(2,s0.max_row+1):
        
        bdate = fmstls.fpygetCell_value(s0, i, 18)   #基準日
        idNo = fmstls.fpygetCell_value(s0, i, 6)     #個体識別番号
        
        ind_calving_list =  fpyext_idNo_s_calvinglist_s( idNo, s1, 2, 6, 7 )
        
        latest_calving = fpysel_clvdate( bdate, ind_calving_list )
        
        fmstls.fpyinputCell_value(s0, i, 9, latest_calving[0]) #calving_date
        fmstls.fpyinputCell_value(s0, i, 8, latest_calving[1]) #parity
        
    wb0.save(wb0N)

#fpyinput_clvdt_into_ymdcow_s##################################################
"""
fpyinput_clvdt_into_ymdcow_s : 
    input calving data at base date into sheet yyyymmddCow00
    sheet version
    v1.0
    2024/6/5
    @author: jicc
    
"""
def fpyinput_clvdt_into_ymdcow_s(s0, s1):
    """
    input calving data at base date into sheet yyyymmddCow00
    sheet version
    Parameters
    ----------
    s0 : worksheet.worksheet.Worksheet
         worksheet object  #yyyymmddCow00
    s1 : worksheet.worksheet.Worksheet
         worksheet object  #calving

    Returns
    -------
    sheet
    s0
    """
    for i in range(2,s0.max_row+1):
        
        bdate = fmstls.fpygetCell_value(s0, i, 18)   #基準日
        idNo = fmstls.fpygetCell_value(s0, i, 6)     #個体識別番号
        
        ind_calving_list =  fpyext_idNo_s_calvinglist_s( idNo, s1, 2, 6, 7 )
        
        latest_calving = fpysel_clvdate( bdate, ind_calving_list )
        
        fmstls.fpyinputCell_value(s0, i, 9, latest_calving[0]) #calving_date
        fmstls.fpyinputCell_value(s0, i, 8, latest_calving[1]) #parity
        
    return s0

#fpyheifers_in_cow_to_heifer##################################################
"""
fpyheifers_in_cow_to_heifer :
    transfer heifers from sheet yyyymmddCow00 to yyyymmddHeifer00
    基準日以降に初産分娩した未経産牛を、
    sheet yyyymmddCow00 から yyyymmddHeifer00 へ移動する
    移動前のyyyymmddCow00をyyyymmddCow00bckとして残すようにしてある。
    ｖ1.0
    2024/3/17
    #* 基準日以降の初産分娩した牛がいない場合の処置を追加
    ｖ1.01
    2024/4/13
    @author: jicc
    
"""
def fpyheifers_in_cow_to_heifer(wbN, s0N, s1N):
    """
    transfer heifers from sheet yyyymmddCow00 to yyyymmddHeifer00
    基準日以降に初産分娩した未経産牛を、
    sheet yyyymmddCow00 から yyyymmddHeifer00 へ移動する
    移動前のyyyymmddCow00をyyyymmddCow00bckとして残すようにしてある。
    2024/3/17
    #* 基準日以降の初産分娩した牛がいない場合の処置を追加
    ｖ1.01
    2024/4/13
    Parameters
    ----------
    wbN : str
        Excelbook's name  : '.\\MH_rpdd.xlsx'
    s0N : str
        sheet name : 'yyyymmddCow00'
    s1N : str
        sheet name : 'yyyymmddHeifer00'

    Returns
    -------
    None.

    """
    wb = openpyxl.load_workbook(wbN) 
    s0 = wb[s0N] #sheet yyyymmddCow00
    s1 = wb[s1N] #sheet yyyymmeeHeifer00
    
    #make lists'list from an list of sheet yyyymmddCow00
    xlcow00list = chghistory.fpyxllist_to_list_s(s0, 18) #sheet 列数 18
    #the length of the list xlcow00list
    lxlcow00list = len(xlcow00list)
    #print(lxlcow00list)
    #print(xlcow00list[0])
    #print(xlcow00list[1])
    #print(xlcow00list[2])

    cow = []            # cows in sheet yyyymmddCow00 
    heifer_in_cow = []  # heifers in sheet yyyymmddCow00

    for i in range(0,lxlcow00list):
        parity = xlcow00list[i][7]
        if parity > 0:
            cow.append(xlcow00list[i])
        elif parity == 0:
            heifer_in_cow.append(xlcow00list[i])
            
    print('heifer_in_cow')
    print(heifer_in_cow)
    
    if len(heifer_in_cow) == 0:     #*
        print(' heifer_in_cow have no element!')
    else:
        #transfer heifers from sheet yyyymmddCow00 to yyyymmddHeifer00
        chghistory.fpylisttoxls_s_(heifer_in_cow, 1, s1)
    
        #change the sheet title from s0N to s0N+'bck'
        s0.title = s0N + 'bck'
        #wb.remove(s0) # or del web[s0]  #sheet s0 を削除する場合
        #wb.remove_sheet(s0) is a deprecated function
    
        s2 = chghistory.fpymkxlsheet_(wb, s0N, 'columns', 1)
    
        chghistory.fpylisttoxls_s_(cow, 1, s2)
    
    wb.save(wbN)

#fpyheifers_in_cow_to_heifer_ws################################################
"""
fpyheifers_in_cow_to_heifer_ws :
    transfer heifers from sheet yyyymmddCow00 to yyyymmddHeifer00
    基準日以降に初産分娩した未経産牛を、
    sheet yyyymmddCow00 から yyyymmddHeifer00 へ移動する
    移動前のyyyymmddCow00をyyyymmddCow00bckとして残すようにしてある。
    v1.0
    2024/3/17
    #* 基準日以降の初産分娩した牛がいない場合の処置を追加
    ｖ1.01
    2024/4/13
    workkbook sheet version
    v1.0
    2024/6/5
    @author: jicc
    
"""
def fpyheifers_in_cow_to_heifer_ws(wb, s0cN, s0c, s0hN, s0h):
    """
    transfer heifers from sheet yyyymmddCow00 to yyyymmddHeifer00
    基準日以降に初産分娩した未経産牛を、
    sheet yyyymmddCow00 から yyyymmddHeifer00 へ移動する
    workkbook sheet version
    
    Parameters
    ----------
    wb : workbook.workbook.Workbook
         workbook object : 'AB_rpdd.xlsx'
    s0cN : str
        sheet name : 'yyyymmddCow00'
    s0c : worksheet.worksheet.Worksheet
         worksheet object : 'sheet yyyymmddCow00'
    s0hN : str
        sheet name : 'yyyymmddHeifer00'
    s0h : worksheet.worksheet.Worksheet
         worksheet object 'sheet yyyymmddHeifer00'

    Returns
    -------
    worksheet.worksheet.Worksheet
    s0c_ , s0c, s0h 
    """
    #make lists'list from an list of sheet yyyymmddCow00
    xlcow00list = chghistory.fpyxllist_to_list_s(s0c, 18) #sheet 列数 18
    #the length of the list xlcow00list
    lxlcow00list = len(xlcow00list)
    #print(lxlcow00list)
    #print(xlcow00list[0])
    #print(xlcow00list[1])
    #print(xlcow00list[2])

    cow = []            # cows in sheet yyyymmddCow00 
    heifer_in_cow = []  # heifers in sheet yyyymmddCow00

    for i in range(0,lxlcow00list):
        parity = xlcow00list[i][7]
        if parity > 0:
            cow.append(xlcow00list[i])
        elif parity == 0:
            heifer_in_cow.append(xlcow00list[i])
            
    print('heifer_in_cow')
    print(heifer_in_cow)
    
    if len(heifer_in_cow) == 0:     #*
        print(' heifer_in_cow have no element!')
    else:
        #transfer heifers from sheet yyyymmddCow00 to yyyymmddHeifer00
        chghistory.fpylisttoxls_s_(heifer_in_cow, 1, s0h)
    
        #change the sheet title from s0N to s0N+'bck'
    s0c.title = s0cN + 'bck'
        #wb.remove(s0) # or del web[s0]  #sheet s0c を削除する場合
        #wb.remove_sheet(s0) is a deprecated function
    s0c_ = chghistory.fpymkxlsheet_(wb, s0cN, 'columns', 1)
    chghistory.fpylisttoxls_s_(cow, 1, s0c_)
    
    return s0c_ , s0c, s0h  

#fpydaysfrmcalving###########################################################
"""
fpydaysfrmcalving :
    calculate days from calving at sheet yyyymmddCow00
    v1.0
    2024/3/17
    @author: jicc
"""
def fpydaysfrmcalving(wbN, sheetN, col_bd, col_clv, col_dsfrmclv):
    """
    calculate days from calving at sheet yyyymmddCow00

    Parameters
    ----------
    wbN : str
        Excelbook's name  : '.\\MH_rpdd.xlsx'
    sheetN : str
        sheet name : 'yyyymmddCow00'
    col_bd : int
        column's number of base_date 基準日
    col_clv : int
        column's number of calving_date 分娩日
    col_dsfrmclv : int
        column's number of daysfrmcalving

    Returns
    -------
    None.

    """
    wb = openpyxl.load_workbook(wbN) 
    sheet = wb[sheetN] #sheet yyyymmddCow00

    for i in range(2,sheet.max_row+1):
        base_date = fmstls.fpygetCell_value(sheet, i, col_bd)
        calving_date = fmstls.fpygetCell_value(sheet, i, col_clv)
        daysfrmcalving = base_date - calving_date
        
        fmstls.fpyinputCell_value(sheet, i, col_dsfrmclv, daysfrmcalving.days)
        
    wb.save(wbN)

#fpydaysfrmcalving_s###########################################################
"""
fpydaysfrmcalving_s :
    calculate days from calving at sheet yyyymmddCow00
    v1.0
    2024/3/17
    sheet version
    v1.0
    2024/6/5
    @author: jicc
"""
def fpydaysfrmcalving_s(sheet, col_bd, col_clv, col_dsfrmclv):
    """
    calculate days from calving at sheet yyyymmddCow00

    Parameters
    ----------
    wbN : str
        Excelbook's name  : '.\\MH_rpdd.xlsx'
    sheet : worksheet.worksheet.Worksheet
         worksheet object : 'sheet yyyymmddCow00'
    col_bd : int
        column's number of base_date 基準日
    col_clv : int
        column's number of calving_date 分娩日
    col_dsfrmclv : int
        column's number of daysfrmcalving

    Returns
    -------
    sheet : worksheet.worksheet.Worksheet

    """
    for i in range(2,sheet.max_row+1):
        base_date = fmstls.fpygetCell_value(sheet, i, col_bd)
        calving_date = fmstls.fpygetCell_value(sheet, i, col_clv)
        daysfrmcalving = base_date - calving_date
        
        fmstls.fpyinputCell_value(sheet, i, col_dsfrmclv, daysfrmcalving.days)
        
    return sheet

#fpyext_idNo_clvd_s_AIlist####################################################
"""
fpyext_idNo_clvd_s_AIlist :
    extract an individual AI list after the latest calving at base date
    個体の、基準日直近の分娩以後のAIlistを抽出する
    ｖ1.0
    2014/3/18
    @author: jicc
    
"""
def fpyext_idNo_clvd_s_AIlist(idNo, clvd, bdate, wbN, sN, idNo_coln, clvd_coln,
                              lstAI_coln, AIt_coln, PT_coln, eDofnc_coln):
    """
    extract an individual AI list after the latest calving at base date
    個体の、基準日直近の分娩以後のAIlistを抽出する

    Parameters
    ----------
    idNo : str
        cowidNo 個体識別番号 
    clvd : datetime.datetime
        calving date 分娩日
    bdate : datetime.datetime
        basedate 基準日
    wbN : str
        Excelfile's name : 'AB_AI.xlsx'
    sN : str
        sheet name : 'AB_AI'
    idNo_coln : int
        the column's number of cowidNo at sheet sN : 2
        sheet sN上の　個体識別番号のある列数
    clvd_coln : int
        the column's number of calving_date at sheet sN : 7 
        sheet sN上の　のある列数
    lstAI_coln : int
        the column's number of lastAI_times at sheet sN : 10 
        sheet sN上の　最終授精日(授精日)のある列数
    AIt_coln : int
        the column's number of AI_times at sheet sN : 11 
        sheet sN上の　授精回数のある列数
    PT_coln : str
        the column's number of PT at sheet sN : 16 
        sheet sN上の　PTのある列数
    eDofnc_coln : datetime.datetime
        the column's number of expDateofnextCalving at sheet sN : 18 
        sheet sN上の　分娩予定日のある列数

    Returns
    -------
    lists'list :
    an individual AI list after the latest calving at base date
    [[lastAI_date, AI_times, PT, expDateofnextCalving], ...]
    
    """
    wb = openpyxl.load_workbook(wbN)
    s = wb[sN]
    
    ind_clvd_AIlist = [] 
    #individual AI list after the latest calving at base date : default
    #[[lastAI_date, AItimes, PT, expDateofnextCalving],...]
    ind_AIdata = [] 
    #individual AI data for AB_AI.xlsx/AB_AI : default
    #[lastAI_date,AI_times, PT, expDateofnextCalving] 
    
    for i in range(2,s.max_row):
        
        idNo_ = fmstls.fpygetCell_value(s, i, idNo_coln)
        clvd_ = fmstls.fpygetCell_value(s, i, clvd_coln)
        lastAI_date = fmstls.fpygetCell_value(s, i, lstAI_coln)
        daysfrmlstAI = bdate - lastAI_date
        
        
        if idNo_ == idNo and clvd_ == clvd: #個体識別番号と分娩日が等しい
            if lastAI_date <= bdate: #基準日より前の授精
                #lastAI_date
                ind_AIdata.append(lastAI_date)
                #AI_times
                AI_times = fmstls.fpygetCell_value(s, i, AIt_coln)
                ind_AIdata.append(AI_times)
                #PT
                if daysfrmlstAI.days < 30: #授精後30日以内なら PT 不明(None)
                    PT = None
                else:#30日以降なら sheet s上のデータを入力
                    PT = fmstls.fpygetCell_value(s, i, PT_coln)
                ind_AIdata.append(PT)
                #expDateofnextCalving　分娩予定日
                expDateofnextCalving = fmstls.fpygetCell_value(s, i, eDofnc_coln)
                ind_AIdata.append(expDateofnextCalving)
            
                ind_clvd_AIlist.append(ind_AIdata)
                ind_AIdata = []
            else: #基準日以降の授精
                continue
        else:
            continue
        
    ind_clvd_AIlist.sort(key = lambda x:x[0]) #, reverse=True
    #lists' listを lastAI_date(授精日) 昇順 でsort lambda関数を利用
        
    return ind_clvd_AIlist

#fpyext_idNo_clvd_s_AIlist_s###################################################
"""
fpyext_idNo_clvd_s_AIlist_s :
    extract an individual AI list after the latest calving at base date
    個体の、基準日直近の分娩以後のAIlistを抽出する
    sheet version
    ｖ1.0
    2014/3/20
    @author: jicc
    
"""
def fpyext_idNo_clvd_s_AIlist_s(idNo, clvd, bdate, sheet, idNo_coln, clvd_coln,
                              lstAI_coln, AIt_coln, PT_coln, eDofnc_coln):
    """
    extract an individual AI list after the latest calving at base date
    個体の、基準日直近の分娩以後のAIlistを抽出する
    sheet version
    Parameters
    ----------
    idNo : str
        cowidNo 個体識別番号 
    clvd : datetime.datetime
        calving date 分娩日
    bdate : datetime.datetime
        basedate 基準日
    sheet : worksheet.worksheet.Worksheet
         worksheet object
    idNo_coln : int
        the column's number of cowidNo at sheet sN : 2
        sheet sN上の　個体識別番号のある列数
    clvd_coln : int
        the column's number of calving_date at sheet sN : 7 
        sheet sN上の　のある列数
    lstAI_coln : int
        the column's number of lastAI_times at sheet sN : 10 
        sheet sN上の　最終授精日(授精日)のある列数
    AIt_coln : int
        the column's number of AI_times at sheet sN : 11 
        sheet sN上の　授精回数のある列数
    PT_coln : str
        the column's number of PT at sheet sN : 16 
        sheet sN上の　PTのある列数
    eDofnc_coln : datetime.datetime
        the column's number of expDateofnextCalving at sheet sN : 18 
        sheet sN上の　分娩予定日のある列数

    Returns
    -------
    lists'list :
    an individual AI list after the latest calving at base date
    [[lastAI_date, AI_times, PT, expDateofnextCalving], ...]
    
    """
       
    ind_clvd_AIlist = [] 
    #individual AI list after the latest calving at base date : default
    #[[lastAI_date, AItimes, PT, expDateofnextCalving],...]
    ind_AIdata = [] 
    #individual AI data for AB_AI.xlsx/AB_AI : default
    #[lastAI_date,AI_times, PT, expDateofnextCalving] 
    
    for i in range(2,sheet.max_row):
        
        idNo_ = fmstls.fpygetCell_value(sheet, i, idNo_coln)
        clvd_ = fmstls.fpygetCell_value(sheet, i, clvd_coln)
        lastAI_date = fmstls.fpygetCell_value(sheet, i, lstAI_coln)
        daysfrmlstAI = bdate - lastAI_date
        
        
        if idNo_ == idNo and clvd_ == clvd: #個体識別番号と分娩日が等しい
            if lastAI_date <= bdate: #基準日より前の授精
                #lastAI_date
                ind_AIdata.append(lastAI_date)
                #AI_times
                AI_times = fmstls.fpygetCell_value(sheet, i, AIt_coln)
                ind_AIdata.append(AI_times)
                #PT
                if daysfrmlstAI.days < 30: #授精後30日以内なら PT 不明(None)
                    PT = None
                else:#30日以降なら sheet s上のデータを入力
                    PT = fmstls.fpygetCell_value(sheet, i, PT_coln)
                ind_AIdata.append(PT)
                #expDateofnextCalving　分娩予定日
                expDateofnextCalving = fmstls.fpygetCell_value(sheet, i, eDofnc_coln)
                ind_AIdata.append(expDateofnextCalving)
            
                ind_clvd_AIlist.append(ind_AIdata)
                ind_AIdata = []
            else: #基準日以降の授精
                continue
        else:
            continue
        
    ind_clvd_AIlist.sort(key = lambda x:x[0]) #, reverse=True
    #lists' listを lastAI_date(授精日) 昇順 でsort lambda関数を利用
        
    return ind_clvd_AIlist

#fpyinput_AIdt_into_ymdcow##################################################
"""
fpyinput_AIdt_into_ymdcow : 
    input AIdata intosheet yyyymmddCow00
    Group, Stage, 授精回数, 初回授精日, 初回授精日数, 最終授精日, 授精後日数, 
    分娩予定日 を入力する
    v1.0
    2024/3/19
    妊娠+ 後に再AIしてしまった場合のAIの処理を追加した。 *) ~ *)
    対象のAIに対して、PT='error' を入力するルールとする。
    v1.1
    2024/6/19
    @author: jicc
    
"""
def fpyinput_AIdt_into_ymdcow(wb0N, s0N, wb1N, s1N, VWP):
    """
    input AIdata into sheet yyyymmddCow00

    Parameters
    ----------
    wb0N : str
        Excelfile's name : 'AB_rpdd.xlsx'
    s0N : str
        sheet name : 'yyyymmddCow00'
    wb1N : str
        Excelfile's name : 'AB_AI.xlsx'
    s1N : str
        sheet name : 'AB_AI'
    VWP : int
        volantary waiting period 
    Returns
    -------
    None.

    """
    wb0 = openpyxl.load_workbook(wb0N)
    s0 = wb0[s0N]
    
    wb1 = openpyxl.load_workbook(wb1N)
    s1 = wb1[s1N]
    
    for i in range(2,s0.max_row+1): #タイトル行を除く
        
        bdate = fmstls.fpygetCell_value(s0, i, 18)   #基準日
        idNo = fmstls.fpygetCell_value(s0, i, 6)     #個体識別番号
        clvd = fmstls.fpygetCell_value(s0, i, 9)     #分娩日
        daysfrmcalving = fmstls.fpygetCell_value(s0, i, 10) #分娩後日数
        # or =(bdate - clvd).days
        print(idNo)
        ind_clvd_AIlist = fpyext_idNo_clvd_s_AIlist_s(idNo, 
           clvd, bdate, s1, 2, 7, 10, 11, 16, 18)
        #[[lastAI_date, AI_times, PT, expDateofnextCalving], ...]
        # lastAI_date(AIdata) ascending order
        print(ind_clvd_AIlist)
        lind_clvd_AIlist = len(ind_clvd_AIlist)
        
        if lind_clvd_AIlist == 0: # not inseminated 未授精
            if daysfrmcalving < VWP:
                fmstls.fpyinputCell_value(s0, i, 2, 1)  #Group 1
                fmstls.fpyinputCell_value(s0, i, 3, '待機') #Stage 
                fmstls.fpyinputCell_value(s0, i, 11, 0) #AItimes 授精回数 0
            elif daysfrmcalving >= VWP:
                fmstls.fpyinputCell_value(s0, i, 2, 2)  #Group 2
                fmstls.fpyinputCell_value(s0, i, 3, 'AI待ち 未授精') #Stage 
                fmstls.fpyinputCell_value(s0, i, 11, 0) #AItimes 授精回数 0
            else:
                continue
        else:                    #inseminated 授精済
            j=-1
            PT = ind_clvd_AIlist[j][2]
            if PT == '-':
                fmstls.fpyinputCell_value(s0, i, 2, 3)  #Group 3
                fmstls.fpyinputCell_value(s0, i, 3, 'AI待ち 授精済') #Stage
            elif PT == None or PT == '?' or PT == '+?' or PT == '-?':
                fmstls.fpyinputCell_value(s0, i, 2, 4)  #Group 4
                fmstls.fpyinputCell_value(s0, i, 3, '妊娠鑑定予定') #Stage
            elif PT == '+':
                fmstls.fpyinputCell_value(s0, i, 2, 5)  #Group 5
                fmstls.fpyinputCell_value(s0, i, 3, '妊娠鑑定＋') #Stage
                fmstls.fpyinputCell_value(s0, i, 16, ind_clvd_AIlist[j][3])
                #分娩予定日
            elif PT == 'error':     # *)
                #listの負のindexを使用し、末尾のAIdataからみていく。
                #lind_clvd_AIlist = 1 で PT =='error'の場合はないはずだが、
                #かりにそのような場合は、for roop を飛び越えて下に行き、 Group, Stage の
                #値が入らないことになる。　これに対する処理はしていない。
                for j in range(-2, -lind_clvd_AIlist-1, -1): #再AIの前のAIに戻る
                    PT = ind_clvd_AIlist[j][2]
                    if PT == '-':
                        fmstls.fpyinputCell_value(s0, i, 2, 3)  #Group 3
                        fmstls.fpyinputCell_value(s0, i, 3, 'AI待ち 授精済') #Stage
                        break
                    elif PT == None or PT == '?' or PT == '+?' or PT == '-?':
                        fmstls.fpyinputCell_value(s0, i, 2, 4)  #Group 4
                        fmstls.fpyinputCell_value(s0, i, 3, '妊娠鑑定予定') #Stage
                        break
                    elif PT == '+':
                        fmstls.fpyinputCell_value(s0, i, 2, 5)  #Group 5
                        fmstls.fpyinputCell_value(s0, i, 3, '妊娠鑑定＋') #Stage
                        fmstls.fpyinputCell_value(s0, i, 16, ind_clvd_AIlist[j][3])
                        #分娩予定日
                        break
                    elif PT == 'error': #error が続いた場合は、もうひとつAIdataを戻る
                        continue                   # *)

            else:
                continue
                
            fmstls.fpyinputCell_value(s0, i, 12, ind_clvd_AIlist[0][0]) 
            #firstAI_date 初回授精日
            fmstls.fpyinputCell_value(s0, i, 11, ind_clvd_AIlist[j][1])
            #AItimes 授精回数
            fmstls.fpyinputCell_value(s0, i, 14, ind_clvd_AIlist[j][0])
            #lastAI_date 最終授精日
            
            fstAIdaysfrmcalving = (ind_clvd_AIlist[0][0]-clvd).days
            fmstls.fpyinputCell_value(s0, i, 13, fstAIdaysfrmcalving) 
            #fstAIdaysfrmcalving 初回授精日数
            daysfrmlstAI = (bdate - ind_clvd_AIlist[j][0]).days
            fmstls.fpyinputCell_value(s0, i, 15, daysfrmlstAI) 
            #daysfrmlstAI 授精後日数
    
    wb0.save(wb0N)

#fpyinput_AIdt_into_ymdcow_s##################################################
"""
fpyinput_AIdt_into_ymdcow_s : 
    input AIdata into sheet yyyymmddCow00
    AIdata を sheet yyyymmdd00 に入力
    Group, Stage, 授精回数, 初回授精日, 初回授精日数, 最終授精日, 授精後日数, 
    分娩予定日 を入力する
    v1.0
    2024/3/19
    sheet version
    v1.0
    2024/6/5
    妊娠+ 後に再AIしてしまった場合のAIの処理を追加した。 *) ~ *)
    対象のAIに対して、PT='error' を入力するルールとする。
    v1.1
    2024/6/19
    @author: jicc
    
"""
def fpyinput_AIdt_into_ymdcow_s(s0, s1, VWP):
    """
    input AIdata into sheet yyyymmddCow00
    AIdata を sheet yyyymmdd00 に入力
    Group, Stage, 授精回数, 初回授精日, 初回授精日数, 最終授精日, 授精後日数, 
    分娩予定日 を入力する
    sheet version
    Parameters
    ----------
    s0 : worksheet.worksheet.Worksheet
         worksheet object : 'sheet yyyymmddCow00'
    s1 : worksheet.worksheet.Worksheet
         worksheet object: 'sheet AB_AI'
    VWP : int
        volantary waiting period 
    Returns
    -------
    None.

    """
    for i in range(2,s0.max_row+1): #タイトル行を除く
        
        bdate = fmstls.fpygetCell_value(s0, i, 18)   #基準日
        idNo = fmstls.fpygetCell_value(s0, i, 6)     #個体識別番号
        clvd = fmstls.fpygetCell_value(s0, i, 9)     #分娩日
        daysfrmcalving = fmstls.fpygetCell_value(s0, i, 10) #分娩後日数
        # or =(bdate - clvd).days
        print(idNo)
        ind_clvd_AIlist = fpyext_idNo_clvd_s_AIlist_s(idNo, 
           clvd, bdate, s1, 2, 7, 10, 11, 16, 18)
        #[[lastAI_date, AI_times, PT, expDateofnextCalving], ...]
        # lastAI_date(AIdata) ascending order
        print(ind_clvd_AIlist)
        lind_clvd_AIlist = len(ind_clvd_AIlist)
        
        if lind_clvd_AIlist == 0: # not inseminated 未授精
            if daysfrmcalving < VWP:
                fmstls.fpyinputCell_value(s0, i, 2, 1)  #Group 1
                fmstls.fpyinputCell_value(s0, i, 3, '待機') #Stage 
                fmstls.fpyinputCell_value(s0, i, 11, 0) #AItimes 授精回数 0
            elif daysfrmcalving >= VWP:
                fmstls.fpyinputCell_value(s0, i, 2, 2)  #Group 2
                fmstls.fpyinputCell_value(s0, i, 3, 'AI待ち 未授精') #Stage 
                fmstls.fpyinputCell_value(s0, i, 11, 0) #AItimes 授精回数 0
            else:
                continue
        else:                    #inseminated 授精済
            j=-1
            PT = ind_clvd_AIlist[j][2]
            if PT == '-':
                fmstls.fpyinputCell_value(s0, i, 2, 3)  #Group 3
                fmstls.fpyinputCell_value(s0, i, 3, 'AI待ち 授精済') #Stage
            elif PT == None or PT == '?' or PT == '+?' or PT == '-?':
                fmstls.fpyinputCell_value(s0, i, 2, 4)  #Group 4
                fmstls.fpyinputCell_value(s0, i, 3, '妊娠鑑定予定') #Stage
            elif PT == '+':
                fmstls.fpyinputCell_value(s0, i, 2, 5)  #Group 5
                fmstls.fpyinputCell_value(s0, i, 3, '妊娠鑑定＋') #Stage
                fmstls.fpyinputCell_value(s0, i, 16, ind_clvd_AIlist[j][3])
                #分娩予定日
            elif PT == 'error':     # *)
                #listの負のindexを使用し、末尾のAIdataからみていく。
                #lind_clvd_AIlist = 1 で PT =='error'の場合はないはずだが、
                #かりにそのような場合は、for roop を飛び越えて下に行き、 Group, Stage の
                #値が入らないことになる。　これに対する処理はしていない。
                for j in range(-2, -lind_clvd_AIlist-1, -1): #再AIの前のAIに戻る
                    PT = ind_clvd_AIlist[j][2]
                    if PT == '-':
                        fmstls.fpyinputCell_value(s0, i, 2, 3)  #Group 3
                        fmstls.fpyinputCell_value(s0, i, 3, 'AI待ち 授精済') #Stage
                        break
                    elif PT == None or PT == '?' or PT == '+?' or PT == '-?':
                        fmstls.fpyinputCell_value(s0, i, 2, 4)  #Group 4
                        fmstls.fpyinputCell_value(s0, i, 3, '妊娠鑑定予定') #Stage
                        break
                    elif PT == '+':
                        fmstls.fpyinputCell_value(s0, i, 2, 5)  #Group 5
                        fmstls.fpyinputCell_value(s0, i, 3, '妊娠鑑定＋') #Stage
                        fmstls.fpyinputCell_value(s0, i, 16, ind_clvd_AIlist[j][3])
                        #分娩予定日
                        break
                    elif PT == 'error': #error が続いた場合は、もうひとつAIdataを戻る
                        continue                   # *)

            else:
                continue
                
            fmstls.fpyinputCell_value(s0, i, 12, ind_clvd_AIlist[0][0]) 
            #firstAI_date 初回授精日
            fmstls.fpyinputCell_value(s0, i, 11, ind_clvd_AIlist[j][1])
            #AItimes 授精回数
            fmstls.fpyinputCell_value(s0, i, 14, ind_clvd_AIlist[j][0])
            #lastAI_date 最終授精日
            
            fstAIdaysfrmcalving = (ind_clvd_AIlist[0][0]-clvd).days
            fmstls.fpyinputCell_value(s0, i, 13, fstAIdaysfrmcalving) 
            #fstAIdaysfrmcalving 初回授精日数
            daysfrmlstAI = (bdate - ind_clvd_AIlist[j][0]).days
            fmstls.fpyinputCell_value(s0, i, 15, daysfrmlstAI) 
            #daysfrmlstAI 授精後日数
    
    return s0

#fpychg_grNo_6_7_ifn########################################################
"""
fpychg_grNo_6_7_ifn :
    change Group and Stage to '6 乾乳' and '7 繁殖対象外', if necessary
    Group と Stage に '6 乾乳' and '7 繁殖対象外'　の分類を加える。
    umotion の　データ　sheet yyyymmddCow_uorg を参照する。
    v1.0
    2024/3/21
    参照ファイルと分娩日がずれている場合が生じ、
    sheet yyyymmddCow00 の　1. 待機　の分類を優先する　if節　#* を追加した。
    v1.01
    2024/5/13
    @author: jicc
    
"""
def fpychg_grNo_6_7_ifn(wbN, sheetN, srefN):
    """
    change Group and Stage to '6 乾乳' and '7 繁殖対象外', if necessary
    Group と Stage に '6 乾乳' and '7 繁殖対象外'　の分類を加える。
    umotion の　データ　sheet yyyymmddCow_uorg を参照する。

    Parameters
    ----------
    wbN : str
        Excel file's name : 'AB_rpdd.xlsx'
    sheetN : str
        sheet name : 'yyyymmddCow00'
    srefN : str
        sheet name for reference : yyyymmddCow_uorg

    Returns
    -------
    None.

    """
    wb = openpyxl.load_workbook(wbN)
    sheet = wb[sheetN]
    sref = wb[srefN]
    
    for i in range(2,sheet.max_row):
        idNo = fmstls.fpygetCell_value(sheet, i, 6)     #idNo 個体識別番号
        Stage = fmstls.fpygetCell_value(sheet, i, 3)
        
        for j in range(2,sref.max_row):
            idNo_ = fmstls.fpygetCell_value(sref, j, 5) #idNo for reference sheet
            if idNo_ == idNo:
                Stage_ = fmstls.fpygetCell_value(sref, j, 3) #状態
                if Stage_ == '乾乳前期' or Stage_ == '乾乳後期' :
                    if Stage != '待機' :   #* v1.01    
                        fmstls.fpyinputCell_value(sheet, i, 2, 6) #Group
                        fmstls.fpyinputCell_value(sheet, i, 3, '乾乳') #Stage
                    else:
                        continue
                
                elif Stage_ == '繁殖除外' :
                    fmstls.fpyinputCell_value(sheet, i, 2, 7) #Group
                    fmstls.fpyinputCell_value(sheet, i, 3, '繁殖対象外') #Stage
                
                else:
                    continue
                
            else:
                continue
            
    wb.save(wbN)


#fpychg_grNo_6_7_ifn_s########################################################
"""
fpychg_grNo_6_7_ifn_s :
    change Group and Stage to '6 乾乳' and '7 繁殖対象外', if necessary
    Group と Stage に '6 乾乳' and '7 繁殖対象外'　の分類を加える。
    umotion の　データ　sheet yyyymmddCow_uorg を参照する。
    v1.0
    2024/3/21
    参照ファイルと分娩日がずれている場合が生じ、
    sheet yyyymmddCow00 の　1. 待機　の分類を優先する　if節　#* を追加した。
    v1.01
    2024/5/13
    
    sheet version
    v1.0
    2024/6/24
    @author: jicc
    
"""
def fpychg_grNo_6_7_ifn_s(sheet, sref):
    """
    change Group and Stage to '6 乾乳' and '7 繁殖対象外', if necessary
    Group と Stage に '6 乾乳' and '7 繁殖対象外'　の分類を加える。
    umotion の　データ　sheet yyyymmddCow_uorg を参照する。

    sheet version

    Parameters
    ----------
    sheet : worksheet.worksheet.Worksheet
         worksheet object　
    sref : worksheet.worksheet.Worksheet
         worksheet object　

    Returns
    -------
    sheet worksheet.worksheet.Worksheet
         worksheet object　
    """
    
    for i in range(2,sheet.max_row):
        idNo = fmstls.fpygetCell_value(sheet, i, 6)     #idNo 個体識別番号
        Stage = fmstls.fpygetCell_value(sheet, i, 3)
        
        for j in range(2,sref.max_row):
            idNo_ = fmstls.fpygetCell_value(sref, j, 5) #idNo for reference sheet
            if idNo_ == idNo:
                Stage_ = fmstls.fpygetCell_value(sref, j, 3) #状態
                if Stage_ == '乾乳前期' or Stage_ == '乾乳後期' :
                    if Stage != '待機' :   #* v1.01    
                        fmstls.fpyinputCell_value(sheet, i, 2, 6) #Group
                        fmstls.fpyinputCell_value(sheet, i, 3, '乾乳') #Stage
                    else:
                        continue
                
                elif Stage_ == '繁殖除外' :
                    fmstls.fpyinputCell_value(sheet, i, 2, 7) #Group
                    fmstls.fpyinputCell_value(sheet, i, 3, '繁殖対象外') #Stage
                
                else:
                    continue
                
            else:
                continue
            
    return sheet

#fpysep_a_stagefrmothr#####################################################
"""
fpysep_a_stagefrmothr :
    separate a stage from other stages
    あるステージを他のステージと分ける
    v1.0
    2024/3/24    
    @author: jicc
    
"""
def fpysep_a_stagefrmothr(xllists, index, vl):
    """
    separate a stage from other stages
    あるステージを他のステージと分ける

    Parameters
    ----------
    xllists : list
        lists'list from Wxcel sheet 'yyyymmddCow00'
    index : TYPEint
        index of 'Stage' : 1
    vl : int
        a value of a Stage : 7 (繁殖対象外)

    Returns
    -------
    separated_list
    [[xllists_out], [xllists_ymd]]

    """
    xllists_out = []    #7 繁殖対象外list default
    xllists_ymd = []    # other stages 繁殖対象list default
    xllists_ = []
    lxllists = len(xllists)
    
    for i in range(0, lxllists):
        Stage = xllists[i][index]
        
        if Stage == vl:
            xllists_out.append(xllists[i])
        else:
            xllists_ymd.append(xllists[i])
            
    xllists_.append(xllists_out)
    xllists_.append(xllists_ymd)
            
    return xllists_

#fpydelete_rows_of_xls#####################################################
"""
fpydelete_rows_of_xls :
    clear a list data of xls's sheet
    Excel sheet の　リストデータを消去する
    行を削除する
    v1.0
    2024/3/24
    @author: jicc

"""
def fpydelete_rows_of_xls(wbN, sheetN):
    """
    clear a list data of xls's sheet
    Excel sheet の　リストデータを消去する
    行を削除する
    **sheet version にするなら、 idx,amountを引数にできる。
    Parameters
    ----------
    wbN : str
        Excel file's name : 'AB_rpdd.xlsx'
    sheetN : str
        sheet name : 'yyyymmddCow00'

    Returns
    -------
    None.

    """
    wb = openpyxl.load_workbook(wbN)
    sheet = wb[sheetN]
    
    sheet.delete_rows(idx=2, amount=sheet.max_row-1) #タイトル行を残す
    
    wb.save(wbN)
    
#fpydelete_rows_of_xls_s#####################################################
"""
fpydelete_rows_of_xls_s :
    clear a list data of xls's sheet
    Excel sheet の　リストデータを消去する
    行を削除する
    sheet version
    v1.0
    2024/3/29
    @author: jicc

"""
def fpydelete_rows_of_xls_s(sheet, r, amt):
    """
    clear a list data of xls's sheet
    Excel sheet の　リストデータを消去する
    行を削除する
    sheet version
    Parameters
    ----------
    sheet : worksheet.worksheet.Worksheet
         worksheet object
    r : int
        first row number to delete
    amt : the number of rows to delete

    Returns
    -------
    None.

    """
        
    sheet.delete_rows(idx=r, amount=amt) #タイトル行を残す
    
    return sheet

#fpysep_out_frm_00############################################################
"""
fpysep_out_frm_00 : 
    separate cows not fall within the breeding from breeding cows
    繁殖対象外の個体を、分離する。
    v1.0
    2024/3/25
    @author: jicc
    
"""
def fpysep_out_frm_00(wbN, sheetN, ncol):
    """
    separate cows not fall within the breeding from breeding cows
    繁殖対象外の個体を、分離する。

    Parameters
    ----------
    wbN : str
        Excel file's name : 'AB_rpdd.xlsx'
    sheetN : str
        sheet name : 'yyyymmddCow00'
    ncol :  int
        number of columns

    Returns
    -------
    None.

    """
    wb = openpyxl.load_workbook(wbN)
    sheet = wb[sheetN]
    
    #change Excelfile's sheet 20240114Cow00 to lists'list
    xllists = chghistory.fpyxllist_to_list_s(sheet, ncol)

    #'7 繁殖対象外' の牛を　list xllsits_outとし、繁殖対象牛 list xllists_ymdと分ける
    index = 1 #Group
    vl = 7    #value : 7 (繁殖対象外)
    xllists_ = fpysep_a_stagefrmothr(xllists, index, vl) 
        #xllists_ = [[xllsits_out], [xllists_ymd]]
    
    #sheet 'yyyymmddCowout'を作成する
    sheetoutN = sheetN.strip('00') + 'out'
    scolN = 'columns'
    r = 1
    sheetout = fmstls.fpyNewSheet_w(wb, sheetoutN, scolN, r)
    
    #sheet 'yyyymmddCow00' のデータを消去する
    r_ = 2
    amt = sheet.max_row -1 #タイトル行を残す
    fpydelete_rows_of_xls_s(sheet, r_, amt)
    
    #xllists_out → sheet yyyymmddout, xllist_ymd → sheet yyyymmddCow00 にデータ移行する
    chghistory.fpylisttoxls_s_(xllists_[1], 1, sheet)
    chghistory.fpylisttoxls_s_(xllists_[0], 1, sheetout)
        
    wb.save(wbN)

#fpysep_out_frm_00_w############################################################
"""
fpysep_out_frm_00_w : 
    separate cows not fall within the breeding from breeding cows
    繁殖対象外の個体を、分離する。
    workbook version
    v1.0
    2024/6/26
    @author: jicc
    
"""
def fpysep_out_frm_00_w(wb, sheetN, ncol):
    """
    separate cows not fall within the breeding from breeding cows
    繁殖対象外の個体を、分離する。
    workbook version
    
    Parameters
    ----------
    wb : wb : 　workbook.workbook.workbook          
        workbook objevt
    sheetN : str
        sheet name : 'yyyymmddCow00'
    ncol :  int
        number of columns

    Returns
    -------
    None.

    """
    sheet = wb[sheetN]
    
    #change Excelfile's sheet 20240114Cow00 to lists'list
    xllists = chghistory.fpyxllist_to_list_s(sheet, ncol)

    #'7 繁殖対象外' の牛を　list xllsits_outとし、繁殖対象牛 list xllists_ymdと分ける
    index = 1 #Group
    vl = 7    #value : 7 (繁殖対象外)
    xllists_ = fpysep_a_stagefrmothr(xllists, index, vl) 
        #xllists_ = [[xllsits_out], [xllists_ymd]]
    
    #sheet 'yyyymmddCowout'を作成する
    sheetoutN = sheetN.strip('00') + 'out'
    scolN = 'columns'
    r = 1
    sheetout = fmstls.fpyNewSheet_w(wb, sheetoutN, scolN, r)
    
    #sheet 'yyyymmddCow00' のデータを消去する
    r_ = 2
    amt = sheet.max_row -1 #タイトル行を残す
    fpydelete_rows_of_xls_s(sheet, r_, amt)
    
    #xllists_out → sheet yyyymmddout, xllist_ymd → sheet yyyymmddCow00 にデータ移行する
    chghistory.fpylisttoxls_s_(xllists_[1], 1, sheet)
    chghistory.fpylisttoxls_s_(xllists_[0], 1, sheetout)
        
    return sheet, sheetout

#fpyinput_agein_dyandmnth_into_ymdheifer ####################################
"""
fpyinput_agein_dyandmnth_into_ymdheifer :
    input parity(=0) , age in day and manth into sheet yyyymmddHeifer00
    産次数(=0), 日齢、 月齢　を　sheet yyyymmddHeifer00 に入力する。
    v1.0
    2024/3/31
    @author: jicc
    
"""
def fpyinput_agein_dyandmnth_into_ymdheifer(wbN, sheetN):
    """
    input parity(=0) , age in day and manth into sheet yyyymmddHeifer00
    産次数(=0), 日齢、 月齢　を　sheet yyyymmddHeifer00 に入力する。

    Parameters
    ----------
    wbN : str
        Excel file's name : AB_rpdd.xlsx
    sheetN : str
        sheet name : yyyymmddHeifer00
    Returns
    -------
    None.

    """
    wb = openpyxl.load_workbook(wbN)
    sheet = wb[sheetN]
    
    #input 0 into column 8 parity(産次)
    for row_num in range(2,sheet.max_row+1):    #タイトル行を除く
        
        sheet.cell(row=row_num, column=8).value = 0
        #column 8 産次 に 0 を入力する。
    
    #input age in day(日齢) into column 9 age_in_day 日齢
    mh_rpdu.fpyAgeinDays_s(sheet)
    
    # input age in month(月齢) into column 10 age_in_month 月齢
    mh_rpdu.fpyAgeinMnths_s(sheet)
    
    wb.save(wbN)

#fpyinput_agein_dyandmnth_into_ymdheifer_s ####################################
"""
fpyinput_agein_dyandmnth_into_ymdheifer_s :
    input parity(=0) , age in day and manth into sheet yyyymmddHeifer00
    産次数(=0), 日齢、 月齢　を　sheet yyyymmddHeifer00 に入力する。
    v1.0
    2024/3/31
    sheet version
    2024/6/6
    v1.0
    @author: jicc
    
"""
def fpyinput_agein_dyandmnth_into_ymdheifer_s(sheet):
    """
    input parity(=0) , age in day and manth into sheet yyyymmddHeifer00
    産次数(=0), 日齢、 月齢　を　sheet yyyymmddHeifer00 に入力する。

    Parameters
    ----------
    sheetN: worksheet.worksheet.Worksheet
         worksheet object : yyyymmddHeifer00
    Returns
    -------
    worksheet object 
    sheet

    """
    #input 0 into column 8 parity(産次)
    for row_num in range(2,sheet.max_row+1):    #タイトル行を除く
        
        sheet.cell(row=row_num, column=8).value = 0
        #column 8 産次 に 0 を入力する。
    
    #input age in day(日齢) into column 9 age_in_day 日齢
    mh_rpdu.fpyAgeinDays_s(sheet)
    
    # input age in month(月齢) into column 10 age_in_month 月齢
    mh_rpdu.fpyAgeinMnths_s(sheet)
    
    return sheet


#fpyext_idNo_s_AIlist_heifer####################################################
"""
fpyext_idNo_s_AIlist_heifer :
    extract an individual AI list for heifers at base date
    未経産牛に関して、個体の、基準日におけるAIlistを抽出する
    ｖ1.0
    2014/4/3
    @author: jicc
    
"""
def fpyext_idNo_s_AIlist_heifer(idNo, bdate, wbN, sN, idNo_coln, lstAI_coln, 
                                AIt_coln, PT_coln, eDofnc_coln):
    """
    extract an individual AI list for heifers at base date
    未経産牛に関して、個体の、基準日におけるAIlistを抽出する

    Parameters
    ----------
    idNo : str
        cowidNo 個体識別番号 
    bdate : datetime.datetime
        basedate 基準日
    wbN : str
        Excelfile's name : 'AB_AI.xlsx'
    sN : str
        sheet name : 'AB_AI'
    idNo_coln : int
        the column's number of cowidNo at sheet sN : 2
        sheet sN上の　個体識別番号のある列数
    lstAI_coln : int
        the column's number of lastAI_times at sheet sN : 10 
        sheet sN上の　最終授精日(授精日)のある列数
    AIt_coln : int
        the column's number of AI_times at sheet sN : 11 
        sheet sN上の　授精回数のある列数
    PT_coln : str
        the column's number of PT at sheet sN : 16 
        sheet sN上の　PTのある列数
    eDofnc_coln : datetime.datetime
        the column's number of expDateofnextCalving at sheet sN : 18 
        sheet sN上の　分娩予定日のある列数

    Returns
    -------
    lists'list :
    an individual AI list after the latest calving at base date
    [[lastAI_date, AI_times, PT, expDateofnextCalving], ...]
    
    """
    wb = openpyxl.load_workbook(wbN)
    s = wb[sN]
    
    ind_AIlist = [] 
    #individual AI list at base date : default
    #[[lastAI_date, AItimes, PT, expDateofnextCalving],...]
    ind_AIdata = [] 
    #individual AI data for AB_AI.xlsx/AB_AI : default
    #[lastAI_date,AI_times, PT, expDateofnextCalving] 
    
    for i in range(2,s.max_row):
        
        idNo_ = fmstls.fpygetCell_value(s, i, idNo_coln)
        lastAI_date = fmstls.fpygetCell_value(s, i, lstAI_coln)
        daysfrmlstAI = bdate - lastAI_date
        
        
        if idNo_ == idNo: #個体識別番号が等しい
            if lastAI_date <= bdate: #基準日より前の授精
                #lastAI_date
                ind_AIdata.append(lastAI_date)
                #AI_times
                AI_times = fmstls.fpygetCell_value(s, i, AIt_coln)
                ind_AIdata.append(AI_times)
                #PT
                if daysfrmlstAI.days < 30: #授精後30日以内なら PT 不明(None)
                    PT = None
                else:#30日以降なら sheet s上のデータを入力
                    PT = fmstls.fpygetCell_value(s, i, PT_coln)
                ind_AIdata.append(PT)
                #expDateofnextCalving　分娩予定日
                expDateofnextCalving = fmstls.fpygetCell_value(s, i, eDofnc_coln)
                ind_AIdata.append(expDateofnextCalving)
            
                ind_AIlist.append(ind_AIdata)
                ind_AIdata = []
            else: #基準日以降の授精
                continue
        else:
            continue
        
    ind_AIlist.sort(key = lambda x:x[0]) #, reverse=True
    #lists' listを lastAI_date(授精日) 昇順 でsort lambda関数を利用
        
    return ind_AIlist
 
#fpyext_idNo_s_AIlist_heifer_s##################################################
"""
fpyext_idNo_s_AIlist_heifer_s :
    extract an individual AI list for heifers at base date
    未経産牛に関して、個体の、基準日におけるAIlistを抽出する
    ｖ1.0
    2014/4/3
    @author: jicc
    
"""
def fpyext_idNo_s_AIlist_heifer_s(idNo, bdate, sheet, idNo_coln, lstAI_coln, 
                                AIt_coln, PT_coln, eDofnc_coln):
    """
    extract an individual AI list for heifers at base date
    sheet version
    未経産牛に関して、個体の、基準日におけるAIlistを抽出する

    Parameters
    ----------
    idNo : str
        cowidNo 個体識別番号 
    bdate : datetime.datetime
        basedate 基準日
    sheet : worksheet.worksheet.Worksheet
         worksheet object
    idNo_coln : int
        the column's number of cowidNo at sheet sN : 2
        sheet sN上の　個体識別番号のある列数
    lstAI_coln : int
        the column's number of lastAI_times at sheet sN : 10 
        sheet sN上の　最終授精日(授精日)のある列数
    AIt_coln : int
        the column's number of AI_times at sheet sN : 11 
        sheet sN上の　授精回数のある列数
    PT_coln : str
        the column's number of PT at sheet sN : 16 
        sheet sN上の　PTのある列数
    eDofnc_coln : datetime.datetime
        the column's number of expDateofnextCalving at sheet sN : 18 
        sheet sN上の　分娩予定日のある列数

    Returns
    -------
    lists'list :
    an individual AI list after the latest calving at base date
    [[lastAI_date, AI_times, PT, expDateofnextCalving], ...]
    
    """
        
    ind_AIlist = [] 
    #individual AI list at base date : default
    #[[lastAI_date, AItimes, PT, expDateofnextCalving],...]
    ind_AIdata = [] 
    #individual AI data for AB_AI.xlsx/AB_AI : default
    #[lastAI_date,AI_times, PT, expDateofnextCalving] 
    
    for i in range(2,sheet.max_row):
        
        idNo_ = fmstls.fpygetCell_value(sheet, i, idNo_coln)
        lastAI_date = fmstls.fpygetCell_value(sheet, i, lstAI_coln)
        daysfrmlstAI = bdate - lastAI_date
        
        
        if idNo_ == idNo: #個体識別番号が等しい
            if lastAI_date <= bdate: #基準日より前の授精
                #lastAI_date
                ind_AIdata.append(lastAI_date)
                #AI_times
                AI_times = fmstls.fpygetCell_value(sheet, i, AIt_coln)
                ind_AIdata.append(AI_times)
                #PT
                if daysfrmlstAI.days < 30: #授精後30日以内なら PT 不明(None)
                    PT = None
                else:#30日以降なら sheet s上のデータを入力
                    PT = fmstls.fpygetCell_value(sheet, i, PT_coln)
                ind_AIdata.append(PT)
                #expDateofnextCalving　分娩予定日
                expDateofnextCalving = fmstls.fpygetCell_value(sheet, i, eDofnc_coln)
                ind_AIdata.append(expDateofnextCalving)
            
                ind_AIlist.append(ind_AIdata)
                ind_AIdata = []
            else: #基準日以降の授精
                continue
        else:
            continue
        
    ind_AIlist.sort(key = lambda x:x[0]) #, reverse=True
    #lists' listを lastAI_date(授精日) 昇順 でsort lambda関数を利用
        
    return ind_AIlist

#fpyinput_AIdt_into_ymdheifer##################################################
"""
fpyinput_AIdt_into_ymdheifer : 
    input AIdata into sheet yyyymmddHeifer00
    AIdata を sheet yyyymmddHeifer00 に入力
    v1.0
    2024/4/3
    @author: jicc
    
"""
def fpyinput_AIdt_into_ymdheifer(wb0N, s0N, wb1N, s1N, VWPm, VWPM):
    """
    input AIdata into sheet yyyymmddHeifer00
    AIdata を sheet yyyymmddHeifer00 に入力
    Parameters
    ----------
    wb0N : str
        Excelfile's name : 'AB_rpdd.xlsx'
    s0N : str
        sheet name : 'yyyymmddHeifer00'
    wb1N : str
        Excelfile's name : 'AB_AI.xlsx'
    s1N : str
        sheet name : 'AB_AI'
   VWPm : int
       age in month to start waiting 授精待機開始月齢
   VWPM : int
       age in month to start AI 授精開始月齢 
    Returns
    -------
    None.

    """
    wb0 = openpyxl.load_workbook(wb0N)
    s0 = wb0[s0N]
    
    wb1 = openpyxl.load_workbook(wb1N)
    s1 = wb1[s1N]
    
    for i in range(2,s0.max_row+1): #タイトル行を除く
        
        bdate = fmstls.fpygetCell_value(s0, i, 18)   #基準日
        idNo = fmstls.fpygetCell_value(s0, i, 6)     #個体識別番号
        birthday = fmstls.fpygetCell_value(s0, i, 7)     #生年月日
        #ageinday = fmstls.fpygetCell_value(s0, i, 9) #日齢
        ageinmonth = fmstls.fpygetCell_value(s0, i, 10) #月齢
        
               
        ind_AIlist = fpyext_idNo_s_AIlist_heifer_s(idNo, 
           bdate, s1, 2, 10, 11, 16, 18)
        #[[lastAI_date, AI_times, PT, expDateofnextCalving], ...]
        # lastAI_date(AIdata) ascending order
        
        if len(ind_AIlist) == 0: # not inseminated 未授精
            if ageinmonth < VWPm:
                fmstls.fpyinputCell_value(s0, i, 2, 8)  #Group 8
                fmstls.fpyinputCell_value(s0, i, 3, '哺育') #Stage 
                fmstls.fpyinputCell_value(s0, i, 11, 0) #AItimes 授精回数 0
            elif ageinmonth >= VWPm and ageinmonth < VWPM:
                fmstls.fpyinputCell_value(s0, i, 2, 1)  #Group 1
                fmstls.fpyinputCell_value(s0, i, 3, '待機') #Stage 
                fmstls.fpyinputCell_value(s0, i, 11, 0) #AItimes 授精回数 0
            elif ageinmonth >= VWPM:
                fmstls.fpyinputCell_value(s0, i, 2, 2)  #Group 2
                fmstls.fpyinputCell_value(s0, i, 3, 'AI待ち 未授精') #Stage 
                fmstls.fpyinputCell_value(s0, i, 11, 0) #AItimes 授精回数 0
            else:
                continue
 
        else:                       #inseminated 授精済
            PT = ind_AIlist[-1][2]  #最終の AI data の鑑定結果
            if PT == '-':
                fmstls.fpyinputCell_value(s0, i, 2, 3)  #Group 3
                fmstls.fpyinputCell_value(s0, i, 3, 'AI待ち 授精済') #Stage
            elif PT == None:
                fmstls.fpyinputCell_value(s0, i, 2, 4)  #Group 4
                fmstls.fpyinputCell_value(s0, i, 3, '妊娠鑑定予定') #Stage
            elif PT == '+':
                fmstls.fpyinputCell_value(s0, i, 2, 5)  #Group 5
                fmstls.fpyinputCell_value(s0, i, 3, '妊娠鑑定＋') #Stage
                fmstls.fpyinputCell_value(s0, i, 16, ind_AIlist[-1][3])
                #expDateofnextCalving 分娩予定日
            else:
                continue
            
            fmstls.fpyinputCell_value(s0, i, 12, ind_AIlist[0][0]) 
            #firstAI_date 初回授精日
            fmstls.fpyinputCell_value(s0, i, 11, ind_AIlist[-1][1])
            #AItimes 授精回数
            fmstls.fpyinputCell_value(s0, i, 14, ind_AIlist[-1][0])
            #lastAI_date 最終授精日
            
            fstAIdaysfrmbirthday = (ind_AIlist[0][0]-birthday).days
            fstAImonthsfrmbirthday = fstAIdaysfrmbirthday/30
            fmstls.fpyinputCell_value(s0, i, 13, fstAImonthsfrmbirthday) 
            #fstAIdaysfrmcalving 初回授精月齢
            daysfrmlstAI = (bdate - ind_AIlist[-1][0]).days
            fmstls.fpyinputCell_value(s0, i, 15, daysfrmlstAI) 
            #daysfrmlstAI 授精後日数
    
    wb0.save(wb0N)

#fpyinput_AIdt_into_ymdheifer_s################################################
"""
fpyinput_AIdt_into_ymdheifer_s : 
    input AIdata into sheet yyyymmddHeifer00
    AIdata を sheet yyyymmddHeifer00 に入力
    v1.0
    2024/4/3
    sheet version
    v1.0
    2024/6/6
    @author: jicc
    
"""
def fpyinput_AIdt_into_ymdheifer_s(s0, s1, VWPm, VWPM):
    """
    input AIdata into sheet yyyymmddHeifer00
    AIdata を sheet yyyymmddHeifer00 に入力
    sheet version
    
    Parameters
    ----------
    s0 : worksheet.worksheet.Worksheet
         worksheet object : 'yyyymmddHeifer00'
    s1 : worksheet.worksheet.Worksheet
         worksheet object : 'AB_AI'
   VWPm : int
       age in month to start waiting 授精待機開始月齢
   VWPM : int
       age in month to start AI 授精開始月齢 
    Returns
    -------
    worksheet object
    s0

    """
    for i in range(2,s0.max_row+1): #タイトル行を除く
        
        bdate = fmstls.fpygetCell_value(s0, i, 18)   #基準日
        idNo = fmstls.fpygetCell_value(s0, i, 6)     #個体識別番号
        birthday = fmstls.fpygetCell_value(s0, i, 7)     #生年月日
        #ageinday = fmstls.fpygetCell_value(s0, i, 9) #日齢
        ageinmonth = fmstls.fpygetCell_value(s0, i, 10) #月齢
        
               
        ind_AIlist = fpyext_idNo_s_AIlist_heifer_s(idNo, 
           bdate, s1, 2, 10, 11, 16, 18)
        #[[lastAI_date, AI_times, PT, expDateofnextCalving], ...]
        # lastAI_date(AIdata) ascending order
        
        if len(ind_AIlist) == 0: # not inseminated 未授精
            if ageinmonth < VWPm:
                fmstls.fpyinputCell_value(s0, i, 2, 8)  #Group 8
                fmstls.fpyinputCell_value(s0, i, 3, '哺育') #Stage 
                fmstls.fpyinputCell_value(s0, i, 11, 0) #AItimes 授精回数 0
            elif ageinmonth >= VWPm and ageinmonth < VWPM:
                fmstls.fpyinputCell_value(s0, i, 2, 1)  #Group 1
                fmstls.fpyinputCell_value(s0, i, 3, '待機') #Stage 
                fmstls.fpyinputCell_value(s0, i, 11, 0) #AItimes 授精回数 0
            elif ageinmonth >= VWPM:
                fmstls.fpyinputCell_value(s0, i, 2, 2)  #Group 2
                fmstls.fpyinputCell_value(s0, i, 3, 'AI待ち 未授精') #Stage 
                fmstls.fpyinputCell_value(s0, i, 11, 0) #AItimes 授精回数 0
            else:
                continue
 
        else:                       #inseminated 授精済
            PT = ind_AIlist[-1][2]  #最終の AI data の鑑定結果
            if PT == '-':
                fmstls.fpyinputCell_value(s0, i, 2, 3)  #Group 3
                fmstls.fpyinputCell_value(s0, i, 3, 'AI待ち 授精済') #Stage
            elif PT == None:
                fmstls.fpyinputCell_value(s0, i, 2, 4)  #Group 4
                fmstls.fpyinputCell_value(s0, i, 3, '妊娠鑑定予定') #Stage
            elif PT == '+':
                fmstls.fpyinputCell_value(s0, i, 2, 5)  #Group 5
                fmstls.fpyinputCell_value(s0, i, 3, '妊娠鑑定＋') #Stage
                fmstls.fpyinputCell_value(s0, i, 16, ind_AIlist[-1][3])
                #expDateofnextCalving 分娩予定日
            else:
                continue
            
            fmstls.fpyinputCell_value(s0, i, 12, ind_AIlist[0][0]) 
            #firstAI_date 初回授精日
            fmstls.fpyinputCell_value(s0, i, 11, ind_AIlist[-1][1])
            #AItimes 授精回数
            fmstls.fpyinputCell_value(s0, i, 14, ind_AIlist[-1][0])
            #lastAI_date 最終授精日
            
            fstAIdaysfrmbirthday = (ind_AIlist[0][0]-birthday).days
            fstAImonthsfrmbirthday = fstAIdaysfrmbirthday/30
            fmstls.fpyinputCell_value(s0, i, 13, fstAImonthsfrmbirthday) 
            #fstAIdaysfrmcalving 初回授精月齢
            daysfrmlstAI = (bdate - ind_AIlist[-1][0]).days
            fmstls.fpyinputCell_value(s0, i, 15, daysfrmlstAI) 
            #daysfrmlstAI 授精後日数
    
    return s0
    
#fpyinput_PT_##################################################################
"""
fpyinput_PT : 繁殖データに、AI台帳から最終の鑑定結果を入力する
    Cow版
    Umotion仕様：Group4 -> Group5の場合、空胎日数を書き換える
    rpdd version 
    AIData　参照sheet を　ｙｙｙｙ から、　AB_AI（牧場すべてのAIData）に変更
    ｖ2.0
    2024/4/6
    @author: jicc
    
"""

#! python3
# MH_GList.xlsx のデータを取得して、　MH_MQ.xlsx に書き込む

def fpyinput_PT_(wbAIＮ, sheetAIN, wbRPDN, sRPDN):
    """
    繁殖データに、AI台帳から最終の鑑定結果を入力する
    Cow版
    Umotion仕様：Group4 -> Group5の場合、空胎月数を書き換える

    Parameters
    ----------
    wbAIＮ : str 
        AIdata workbook name
        "AB_AI.xlsx"
    sheetAIN : str 
        AI data sheet name
        "AB_AI"
    wbRPDN : str 
        workbook name
        "AB_rpdd.xlsx"
    sRPDN : str 
        Cows' data sheet name
        "yyyymmddCow01"

    Returns
    -------
    None.

    """
     
    wbAI = openpyxl.load_workbook(wbAIN)        #MH_AI.xlsx
    sheetAI = wbAI[sheetAIN]      #.get_sheet_by_name(sheetAIN)

    wbRPD = openpyxl.load_workbook(wbRPDN)      #MH_RPD.xlsx
    sheetRPD = wbRPD[sRPDN]         #.get_sheet_by_name(sRPDN)


       #2行目からスタート
    for i in range(2, sheetRPD.max_row + 1):     #先頭行をスキップ
        
        #繁殖ファイルのGroupを取得
        Group = sheetRPD.cell(row=i, column=2).value   #Group
        
        if Group == 4:
            cowidNoRPD = sheetRPD.cell(row=i, column=6).value
            #繁殖ファイルの個体識別番号
            lstAIDtate = sheetRPD.cell(row=i, column=14).value
            #繁殖ファイルの最終授精日
            clvingDate = sheetRPD.cell(row=i, column=9).value
            #繁殖ファイルの分娩日
            
            for j in range(2, sheetAI.max_row + 1 ):    #sheetAI.max_row + 1
                    
                cowidNoAI = sheetAI.cell(row=j, column=2).value
                #AI台帳の耳標番号
                lstAI_date = sheetAI.cell(row=j, column=10).value
                #AI台帳の最終授精日
                if (cowidNoRPD == cowidNoAI) and (lstAIDtate == lstAI_date):
                    
                    PT = sheetAI.cell(row=j, column=16).value
                    #AI台帳のPT（鑑定結果、+, -, ?, NONE）を取得
                    if PT == '+':
                        expDayofnxtclving = sheetAI.cell(row=j, column=18).value
                        #分娩予定日
                        sheetRPD.cell(row=i, column=2).value = 5
                        #Groupを　5　に変更
                        sheetRPD.cell(row=i, column=3).value = '妊娠鑑定+'
                        #Stage　を　'妊娠鑑定+'に変更
                        sheetRPD.cell(row=i, column=16).value = expDayofnxtclving
                        #分娩予定日を入力
                        openDays = lstAIDtate - clvingDate #空胎日数
                        sheetRPD.cell(row=i, column=17).value = openDays.days
                        #空胎日数を入力
                        break
                    
                    elif PT == '-':
                        sheetRPD.cell(row=i, column=2).value = 3
                        #Groupを　3　に変更
                        sheetRPD.cell(row=i, column=3).value = 'AI待ち 授精済'
                        #Stage　を　'AI待ち 授精済'に変更

                        break
                    
                    else:   #None, ?, -?, +? など　何もしない
                        break
                else:   #何もしない
                    continue
        else:    #Group != 4 何もしない
            continue
        
    wbRPD.save(wbRPDN)

#fpyinput_PT__s##################################################################
"""
fpyinput_PT__s : 繁殖データに、AI台帳から最終の鑑定結果を入力する
    Cow版
    Umotion仕様：Group4 -> Group5の場合、空胎日数を書き換える
    rpdd version 
    AIData　参照sheet を　ｙｙｙｙ から、　AB_AI（牧場すべてのAIData）に変更
    ｖ2.0
    2024/4/6
    sheet version
    v1.0
    2024/6/5
    @author: jicc
    
"""

#! python3
# MH_GList.xlsx のデータを取得して、　MH_MQ.xlsx に書き込む

def fpyinput_PT__s(sheetAI, sheetRPD):
    """
    繁殖データに、AI台帳から最終の鑑定結果を入力する
    Cow版
    Umotion仕様：Group4 -> Group5の場合、空胎月数を書き換える
    sheet version
    v1.0
    2024/6/5
    Parameters
    ----------
    sheetAI : sworksheet.worksheet.Worksheet
         worksheet object  "AB_AI"
    sheetRPDN : worksheet.worksheet.Worksheet
         worksheet object "yyyymmddCow01"

    Returns
    -------
    None.

    """
    #2行目からスタート
    for i in range(2, sheetRPD.max_row + 1):     #先頭行をスキップ
        
        #繁殖ファイルのGroupを取得
        Group = sheetRPD.cell(row=i, column=2).value   #Group
        
        if Group == 4:
            cowidNoRPD = sheetRPD.cell(row=i, column=6).value
            #繁殖ファイルの個体識別番号
            lstAIDtate = sheetRPD.cell(row=i, column=14).value
            #繁殖ファイルの最終授精日
            clvingDate = sheetRPD.cell(row=i, column=9).value
            #繁殖ファイルの分娩日
            
            for j in range(2, sheetAI.max_row + 1 ):    #sheetAI.max_row + 1
                    
                cowidNoAI = sheetAI.cell(row=j, column=2).value
                #AI台帳の耳標番号
                lstAI_date = sheetAI.cell(row=j, column=10).value
                #AI台帳の最終授精日
                if (cowidNoRPD == cowidNoAI) and (lstAIDtate == lstAI_date):
                    
                    PT = sheetAI.cell(row=j, column=16).value
                    #AI台帳のPT（鑑定結果、+, -, ?, NONE）を取得
                    if PT == '+':
                        expDayofnxtclving = sheetAI.cell(row=j, column=18).value
                        #分娩予定日
                        sheetRPD.cell(row=i, column=2).value = 5
                        #Groupを　5　に変更
                        sheetRPD.cell(row=i, column=3).value = '妊娠鑑定+'
                        #Stage　を　'妊娠鑑定+'に変更
                        sheetRPD.cell(row=i, column=16).value = expDayofnxtclving
                        #分娩予定日を入力
                        openDays = lstAIDtate - clvingDate #空胎日数
                        sheetRPD.cell(row=i, column=17).value = openDays.days
                        #空胎日数を入力
                        break
                    
                    elif PT == '-':
                        sheetRPD.cell(row=i, column=2).value = 3
                        #Groupを　3　に変更
                        sheetRPD.cell(row=i, column=3).value = 'AI待ち 授精済'
                        #Stage　を　'AI待ち 授精済'に変更

                        break
                    
                    else:   #None, ?, -?, +? など　何もしない
                        break
                else:   #何もしない
                    continue
        else:    #Group != 4 何もしない
            continue
        
        
    return sheetRPD

#fpyinput_PTH_############################################################
"""
fpyinput_PTH : 繁殖データに、AI台帳から最終の鑑定結果を入力する
    Heifer版
    Umotion仕様：Group4 -> Group5の場合、空胎月数を書き換える
    rpdd version 
    AIData　参照sheet を　ｙｙｙｙ から、　AB_AI（牧場すべてのAIData）に変更
    ｖ2.0
    2024/4/6
    @author: jicc
    
"""
#! python3
# MH_GList.xlsx のデータを取得して、　MH_MQ.xlsx に書き込む

def fpyinput_PTH_(wbAIＮ, sheetAIN, wbRPDN, sRPDN):
    """
    繁殖データに、AI台帳から最終の鑑定結果を入力する
    Heifer版
    Umotion仕様：Group4 -> Group5の場合、空胎月数を書き換える

    Parameters
    ----------
    wbAIＮ : str 
        AIdata workbook name
        "AB_AI.xlsx"
    sheetAIN : str 
        AI data sheet name
        "AB_AI"
    wbRPDN : str 
        workbook name
        "MH_RPDu.xlsx"
    sRPDN : str 
        Heifers' data sheet name
        "yyyymmddHeifer01"

    Returns
    -------
    None.

    """

    wbAI = openpyxl.load_workbook(wbAIN)        #MH_AI.xlsx
    sheetAI = wbAI[sheetAIN]      #.get_sheet_by_name(sheetAIN)

    wbRPD = openpyxl.load_workbook(wbRPDN)      #MH_RPDu.xlsx
    sheetRPD = wbRPD[sRPDN]         #.get_sheet_by_name(sRPDN)


       #2行目からスタート
    for i in range(2, sheetRPD.max_row + 1):     #先頭行をスキップ
        
        #繁殖ファイルのGroupを取得
        Group = sheetRPD.cell(row=i, column=2).value   #Group
        
        if Group == 4:
            cowidNoRPD = sheetRPD.cell(row=i, column=6).value
            #繁殖ファイルの個体識別番号
            birthday = sheetRPD.cell(row=i, column=7).value
            #繁殖ファイルの生年月日
            lstAIDtate = sheetRPD.cell(row=i, column=14).value
            #繁殖ファイルの最終授精日
            
            
            for j in range(2, sheetAI.max_row + 1 ):    #sheetAI.max_row + 1
                    
                cowidNoAI = sheetAI.cell(row=j, column=2).value
                #AI台帳の耳標番号
                lstAI_date = sheetAI.cell(row=j, column=10).value
                #AI台帳の最終授精日
                if (cowidNoRPD == cowidNoAI) and (lstAIDtate == lstAI_date):
                    
                    PT = sheetAI.cell(row=j, column=16).value
                    #AI台帳のPT（鑑定結果、+, -, ?, NONE）を取得
                    if PT == '+':
                        expDayofnxtclving = sheetAI.cell(row=j, column=18).value
                        #分娩予定日
                        sheetRPD.cell(row=i, column=2).value = 5
                        #Groupを　5　に変更
                        sheetRPD.cell(row=i, column=3).value = '妊娠鑑定+'
                        #Stage　を　'妊娠鑑定+'に変更
                        sheetRPD.cell(row=i, column=16).value = expDayofnxtclving
                        #分娩予定日を入力
                        openDays = lstAIDtate - birthday 
                        #空胎日数=最終授精日 - 生年月日
                        openDays = openDays.days
                        openMonths = openDays/30
                        #空胎月数
                        sheetRPD.cell(row=i, column=17).value = openMonths
                        #空胎月数を入力
                        break
                    
                    elif PT == '-':
                        sheetRPD.cell(row=i, column=2).value = 3
                        #Groupを　3　に変更
                        sheetRPD.cell(row=i, column=3).value = 'AI待ち 授精済'
                        #Stage　を　'AI待ち 授精済'に変更
                        
                        break
                    
                    else:   #None, ?, -?, +? など　何もしない
                        break
                else:   #何もしない
                    continue
        else:    #Group != 4 何もしない
            continue
        
    wbRPD.save(wbRPDN)

#fpyinput_PTH__s############################################################
"""
fpyinput_PTH__s : 繁殖データに、AI台帳から最終の鑑定結果を入力する
    Heifer版
    Umotion仕様：Group4 -> Group5の場合、空胎月数を書き換える
    rpdd version 
    AIData　参照sheet を　ｙｙｙｙ から、　AB_AI（牧場すべてのAIData）に変更
    ｖ2.0
    2024/4/6
    sheet version
    v1.0
    2024/6/6
    @author: jicc
    
"""
#! python3
# MH_GList.xlsx のデータを取得して、　MH_MQ.xlsx に書き込む

def fpyinput_PTH__s(sheetAI, sheetRPD):
    """
    繁殖データに、AI台帳から最終の鑑定結果を入力する
    Heifer版
    Umotion仕様：Group4 -> Group5の場合、空胎月数を書き換える

    Parameters
    ----------
    sheetAI : sworksheet.worksheet.Worksheet
         worksheet object  "AB_AI"
    sheetRPDN : worksheet.worksheet.Worksheet
         worksheet object "yyyymmddHeifer01"

    Returns
    -------
    None.

    """
       #2行目からスタート
    for i in range(2, sheetRPD.max_row + 1):     #先頭行をスキップ
        
        #繁殖ファイルのGroupを取得
        Group = sheetRPD.cell(row=i, column=2).value   #Group
        
        if Group == 4:
            cowidNoRPD = sheetRPD.cell(row=i, column=6).value
            #繁殖ファイルの個体識別番号
            birthday = sheetRPD.cell(row=i, column=7).value
            #繁殖ファイルの生年月日
            lstAIDtate = sheetRPD.cell(row=i, column=14).value
            #繁殖ファイルの最終授精日
            
            
            for j in range(2, sheetAI.max_row + 1 ):    #sheetAI.max_row + 1
                    
                cowidNoAI = sheetAI.cell(row=j, column=2).value
                #AI台帳の耳標番号
                lstAI_date = sheetAI.cell(row=j, column=10).value
                #AI台帳の最終授精日
                if (cowidNoRPD == cowidNoAI) and (lstAIDtate == lstAI_date):
                    
                    PT = sheetAI.cell(row=j, column=16).value
                    #AI台帳のPT（鑑定結果、+, -, ?, NONE）を取得
                    if PT == '+':
                        expDayofnxtclving = sheetAI.cell(row=j, column=18).value
                        #分娩予定日
                        sheetRPD.cell(row=i, column=2).value = 5
                        #Groupを　5　に変更
                        sheetRPD.cell(row=i, column=3).value = '妊娠鑑定+'
                        #Stage　を　'妊娠鑑定+'に変更
                        sheetRPD.cell(row=i, column=16).value = expDayofnxtclving
                        #分娩予定日を入力
                        openDays = lstAIDtate - birthday 
                        #空胎日数=最終授精日 - 生年月日
                        openDays = openDays.days
                        openMonths = openDays/30
                        #空胎月数
                        sheetRPD.cell(row=i, column=17).value = openMonths
                        #空胎月数を入力
                        break
                    
                    elif PT == '-':
                        sheetRPD.cell(row=i, column=2).value = 3
                        #Groupを　3　に変更
                        sheetRPD.cell(row=i, column=3).value = 'AI待ち 授精済'
                        #Stage　を　'AI待ち 授精済'に変更
                        
                        break
                    
                    else:   #None, ?, -?, +? など　何もしない
                        break
                else:   #何もしない
                    continue
        else:    #Group != 4 何もしない
            continue
        
    return sheetRPD


#fpyinput_eartagno_into_ymdcow###############################################
"""
fpyinput_eartagno_into_ymdcow: 
    input eartagno into sheet yyyymmddCow
    耳標番号を　sheeet yyyymmddCow00 に cowslist から入力する
    v1.0
    2024/5/19
    @author: inoue
    
"""
def fpyinput_eartagno_into_ymdcow(wb0N, s0N, cid0n, cet0n,
                                              wb1N, s1N, cid1n, cet1n):
    """
    input eartagno into sheet yyyymmddCow

    Parameters
    ----------
    wb0N : str
        Excelfile's name : 'AB_rpdd.xlsx'
    s0N : str
        sheet name : 'yyyymmddCow00'
    cidon : int
        column number of cowidNo
    cet0n : int
        column number of '牛コード'
    wb1N : str
        Excelfile's name : 'AB_cowslist.xlsx'
    s1N : str
        sheet name : 'cowslistyyyymmdd'
    cid1n : int
        column number of cowidNo
    cet1n : int
        column number of eartagNo 

    Returns
    -------
    None.

    """
    #import openpyxl
    #import fmstls
    
    wb0 = openpyxl.load_workbook(wb0N)
    s0 = wb0[s0N]
    
    wb1 = openpyxl.load_workbook(wb1N)
    s1 = wb1[s1N]
    
    for i in range(2,s0.max_row+1): #sheet s0 タイトル行を除く
    
        idNo = fmstls.fpygetCell_value(s0, i, cid0n) #i行の個体の　idNo
        
        for j in range (2,s1.max_row+1): #sheet s1 タイトル行を除く
            
            idNo_ = fmstls.fpygetCell_value(s1, j, cid1n) #j行の個体の　idNo
            
            if idNo_ == idNo :
                
                eartagNo = fmstls.fpygetCell_value(s1, j, cet1n)
                fmstls.fpyinputCell_value(s0, i, cet0n, eartagNo)
                
            else:
                continue
    
    wb0.save(wb0N)

#fpyrpdd_Cow_manual00######################################################
"""
fpyrpdd_Cow_manual00 : 
        1.input calving data into sheet yyyymmddCow00
        2.transfer heifers at base date 
        from sheet yyyymmddCow00 to yyyymmddHeifer00
        3.calculate days from calving at sheet yyyymmddCow00
        4.input AIdata into sheet yyyymmddCow00 
         and group cows by their reproductive stage
        5.input opendays into sheet yyyymmddCow00
       v1.0 
       2024/6/5
       @author: inoue
       
"""
def fpyrpdd_Cow_manual00( wb0N, s0cN, s0hN, wb1N, s1N, wb2N, s2N, VWP ):
    """
        1.input calving data into sheet yyyymmddCow00
        2.transfer heifers at base date 
        from sheet yyyymmddCow00 to yyyymmddHeifer00
        3.calculate days from calving at sheet yyyymmddCow00
        4.input AIdata into sheet yyyymmddCow00 
         and group cows by their reproductive stage
        5.input opendays into sheet yyyymmddCow00
        v1.0 
        2024/6/5
        @author: inoue

    Parameters
    ----------
    wb0N : str
        Excelfile's name : 'AB_rpdd.xlsx'
    s0cN : str
        sheet name : 'yyyymmddCow00'
    s0hN : str
        sheet name : 'yyyymmddHeifer00'
    wb1N : str
        Excelfile's name : 'AB_calving.xlsx'
    s1N : str
        sheet name : 'calving'
    wb2N : str
        Excelfile's name : 'AB_AI.xlsx'
    s2N : str
        sheet name : 'AB_AI'
    VWP : int
        volantary waiting period     
    Returns
    -------
    None.

    """
    
    wb0 = openpyxl.load_workbook(wb0N)  #AB_rpdd.xlsx
    s0c = wb0[s0cN]                       #yyyymmddCow00
    s0h = wb0[s0hN]                       #yyyymmddHeifer00
   
    wb1 = openpyxl.load_workbook(wb1N)  #AB_calving.xlsx
    s1 = wb1[s1N]                       #calving
    
    wb2 = openpyxl.load_workbook(wb2N)  #AB_AI.xlsx
    s2 = wb2[s2N]                       #AB_AI
    
    #1.input calving data into sheet yyyymmddCow00
    #基準日における分娩日、産次数を sheet yyyymmddCow00 に入力する 
    fpyinput_clvdt_into_ymdcow_s(s0c, s1)
    
    #2.transfer heifers at base date 
    # from sheet yyyymmddCow00 to yyyymmddHeifer00
    #基準日以降に初産分娩した未経産牛を、
    #sheet yyyymmddCow00 から yyyymmddHeifer00 へ移動する
    s0c_, s0c, s0h = fpyheifers_in_cow_to_heifer_ws(wb0, s0cN, s0c, s0hN, s0h)
    
    #3.calculate days from calving at sheet yyyymmddCow00(s0c_)
    s0c_ = fpydaysfrmcalving_s(s0c_, 18, 9, 10)
        #base_date:18, calving_date:9, daysfrmcalving:10
        
    # 4.input AIdata into sheet yyyymmddCow00 
    #  and group cows by their reproductive stage
    #Group, Stage, 授精回数, 初回授精日, 初回授精日数, 最終授精日, 授精後日数, 
    #分娩予定日 を入力する
    s0c_ = fpyinput_AIdt_into_ymdcow_s(s0c_, s2, VWP)
    
    #5.input opendays into sheet yyyymmddCow00
    mh_rpdu.fpyopenDays_s(s0c_)

    wb0.save(wb0N)

#fpyrpdd_Cow_manual01######################################################
"""
fpyrpdd_Cow_manual01 : 
        1.copy sheet yyyymmddCow00 and make sheet yyyymmddCow01
        2.繁殖データに、AI台帳から最終の鑑定結果を入力する
        
        v1.0 
        2024/6/5
        @author: inoue
       
"""
def fpyrpdd_Cow_manual01( wb0N, s0cN, s0c_N, wb1N, s1N ):
    """
    1.copy sheet yyyymmddCow00 and make sheet yyyymmddCow01
    2.繁殖データに、AI台帳から最終の鑑定結果を入力する
    
    v1.0 
    2024/6/5
    @author: inoue

    Parameters
    ----------
    wb0N : str
        Excelfile's name : 'AB_rpdd.xlsx'
    s0cN : str
        sheet name : 'yyyymmddCow00'
    s0c_N : str
        sheet name : 'yyyymmddCow01'
    wb1N : str
        Excelfile's name : 'AB_AI.xlsx'
    s1N : str
        sheet name : 'AB_AI'
 
    Returns
    -------
    None.

    """
    
    wb0 = openpyxl.load_workbook(wb0N)  #AB_rpdd.xlsx
    s0c = wb0[s0cN]                       #yyyymmddCow00
    #s0c_ = wb0[s0c_N]                       #yyyymmddHeifer00
   
    wb1 = openpyxl.load_workbook(wb1N)  #AB_AI.xlsx
    s1 = wb1[s1N]                       #AB_AI
    
       
    #1.copy sheet yyyymmddCow00 and make sheet yyyymmddCow01
    s0c_ = fmstls.fpysheet_copy_ws( wb0, s0c, s0c_N )
    
    #2.繁殖データに、AI台帳から最終の鑑定結果を入力する
    fpyinput_PT__s(s1, s0c_)
    
    wb0.save(wb0N)

#fpyrpdd_Cow_manual00######################################################
"""
fpyrpdd_Heifer_manual00 : 
        1.input parity(=0) , age in day and manth into sheet yyyymmddHeifer00
            産次数(=0), 日齢、 月齢　を　sheet yyyymmddHeifer00 に入力する。
        2.input AIdata into sheet yyyymmddHeifer00
            and group heifers by their reproductive stage
            AIdata を sheet yyyymmddHeifer00 に入力
            Group, Stage, 授精回数, 初回授精日, 初回授精月齢, 最終授精日, 
            授精後日数, 分娩予定日 を入力する
        3.input openmonths into sheet yyyymmddHeifer00
            空胎月数を計算・入力 する
        v1.0 
        2024/6/6
        @author: inoue
       
"""
def fpyrpdd_Heifer_manual00( wb0N, s0hN, wb1N, s1N, VWPm, VWPM ):
    """
    1.input parity(=0) , age in day and manth into sheet yyyymmddHeifer00
        産次数(=0), 日齢、 月齢　を　sheet yyyymmddHeifer00 に入力する。
    2.input AIdata into sheet yyyymmddHeifer00
        and group heifers by their reproductive stage
        AIdata を sheet yyyymmddHeifer00 に入力
        Group, Stage, 授精回数, 初回授精日, 初回授精月齢, 最終授精日, 
        授精後日数, 分娩予定日 を入力する
    3.input openmonths into sheet yyyymmddHeifer00
        空胎月数を計算・入力 する
    v1.0 
    2024/6/5
    @author: inoue

    Parameters
    ----------
    wb0N : str
        Excelfile's name : 'AB_rpdd.xlsx'
    s0hN : str
        sheet name : 'yyyymmddHeifer00'
    wb1N : str
        Excelfile's name : 'AB_AI.xlsx'
    s1N : str
        sheet name : 'AB_AI'
    VWPm : int
           age in month to start waiting 授精待機開始月齢 : 13
    VWPM : int
           age in month to start AI 授精開始月齢  : 14
 
    Returns
    -------
    None.

    """
    
    wb0 = openpyxl.load_workbook(wb0N)  #AB_rpdd.xlsx
    s0h = wb0[s0hN]                       #yyyymmddHeifer00
   
    wb1 = openpyxl.load_workbook(wb1N)  #AB_AI.xlsx
    s1 = wb1[s1N]                       #AB_AI
    
       
    #1.input parity(=0) , age in day and manth into sheet yyyymmddHeifer00
    s0h = fpyinput_agein_dyandmnth_into_ymdheifer_s(s0h)
    
    #2.input AIdata into sheet yyyymmddHeifer00
    #    and group heifers by their reproductive stage
    #    AIdata を sheet yyyymmddHeifer00 に入力
    #    Group, Stage, 授精回数, 初回授精日, 初回授精月齢, 最終授精日, 
    #    授精後日数, 分娩予定日 を入力する
    s0h = fpyinput_AIdt_into_ymdheifer_s(s0h, s1, VWPm, VWPM)
    
    #3.input openmonths into sheet yyyymmddHeifer00
    #    空胎月数を計算・入力 する
    s0h = mh_rpdu.fpyopenMnths_s( s0h )
    
    wb0.save(wb0N)

#fpyrpdd_Heifer_manual01######################################################
"""
fpyrpdd_Heifer_manual01 : 
        1.copy sheet yyyymmddHeifer00 and make sheet yyyymmddHeifer01
        2.繁殖データに、AI台帳から最終の鑑定結果を入力する
        
        v1.0 
        2024/6/6
        @author: inoue
       
"""
def fpyrpdd_Heifer_manual01( wb0N, s0hN, s0h_N, wb1N, s1N ):
    """
    1.copy sheet yyyymmddHeifer00 and make sheet yyyymmddHeifer01
    2.繁殖データに、AI台帳から最終の鑑定結果を入力する
    
    v1.0 
    2024/6/6
    @author: inoue

    Parameters
    ----------
    wb0N : str
        Excelfile's name : 'AB_rpdd.xlsx'
    s0hN : str
        sheet name : 'yyyymmddHeifer00'
    s0h_N : str
        sheet name : 'yyyymmddHeifer01'
    wb1N : str
        Excelfile's name : 'AB_AI.xlsx'
    s1N : str
        sheet name : 'AB_AI'
 
    Returns
    -------
    None.

    """
    
    wb0 = openpyxl.load_workbook(wb0N)  #AB_rpdd.xlsx
    s0h = wb0[s0hN]                       #yyyymmddHeifer00
    #s0h_ = wb0[s0h_N]                       #yyyymmddHeifer01
   
    wb1 = openpyxl.load_workbook(wb1N)  #AB_AI.xlsx
    s1 = wb1[s1N]                       #AB_AI
    
       
    #1.copy sheet yyyymmddCow00 and make sheet yyyymmddCow01
    s0h_ = fmstls.fpysheet_copy_ws( wb0, s0h, s0h_N )
    
    #2.繁殖データに、AI台帳から最終の鑑定結果を入力する
    fpyinput_PTH__s(s1, s0h_)
    
    wb0.save(wb0N)

#fpyrpdd_MH_tools00######################################################
"""
fpyrpdd_MH_tools00 : 
        0.bdateに近い '経産牛一覧plusyyyymmdd.csv' をMH_rpdd.xlsxに移動し、
        sheeet ‘yyyymmddCow_uorg’ とする。 hand work 
        1.5列 個体識別番号を10桁文字列に変更
        2.change Group and Stage to '6 乾乳' and '7 繁殖対象外', if necessary 
         Group と Stage に '6 乾乳' and '7 繁殖対象外'　の分類を加える。
         umotion の　データ　sheet yyyymmddCow_uorg を参照する。
        3.copy sheet yyyymmddCow00 and make sheet name yyyymmddCow00all
         yyyymmddCow00 をコピーして、yyyymmddCow00allを作成 
        4.separate cows not fall within the breeding from breeding cows
         繁殖対象外の個体を、分離する。 
        v1.0 
       2024/6/25
       @author: inoue
       
"""
def fpyrpdd_MH_tools00( wb0N, s0cN, ncolc, s0uN, coluidNo, s1cN):
    """
        0.bdateに近い '経産牛一覧plusyyyymmdd.csv' をMH_rpdd.xlsxに移動し、
        sheeet ‘yyyymmddCow_uorg’ とする。 hand work 
        1.5列 個体識別番号を10桁文字列に変更
        2.change Group and Stage to '6 乾乳' and '7 繁殖対象外', if necessary 
         Group と Stage に '6 乾乳' and '7 繁殖対象外'　の分類を加える。
         umotion の　データ　sheet yyyymmddCow_uorg を参照する。
        3.copy sheet yyyymmddCow00 and make sheet name yyyymmddCow00all
         yyyymmddCow00 をコピーして、yyyymmddCow00allを作成 
        4.separate cows not fall within the breeding from breeding cows
         繁殖対象外の個体を、分離する。 
        v1.0 
        2024/6/25
        @author: inoue

    Parameters
    ----------
    wb0N : str
        Excelfile's name : 'AB_rpdd.xlsx'
    s0cN : str
        sheet name : 'yyyymmddCow00'
    ncolc : int
        number of columns of sheet s0cN 'yyyymmddCow00' :18
    s0uN : str
        sheet name : 'yyyymmddCow_uorg'
    coluidNo : int
        idNo's columun number' at sheet s0uN 'yyyymmddCow_uorg' :5
    s1cN : str
        sheet name : 'yyyymmddCow00all'
  
    Returns
    -------
    None.

    """
    
    wb0 = openpyxl.load_workbook(wb0N)  #AB_rpdd.xlsx
    s0c = wb0[s0cN]                       #yyyymmddCow00
    s0u = wb0[s0uN]                       #yyyymmddCow_uorg
   
    #1.5列 個体識別番号を10桁文字列に変更
    fmstls.fpyidNo_9to10_s(s0u, coluidNo)
    
    #2.change Group and Stage to '6 乾乳' and '7 繁殖対象外', if necessary 
    # Group と Stage に '6 乾乳' and '7 繁殖対象外'　の分類を加える。
    # umotion の　データ　sheet yyyymmddCow_uorg を参照する。
    s0c = fpychg_grNo_6_7_ifn_s(s0c, s0u)
   
    #3.copy sheet yyyymmddCow00 and make sheet name yyyymmddCow00all
    # yyyymmddCow00 をコピーして、yyyymmddCow00allを作成 
    fmstls.fpysheet_copy_ws( wb0, s0c, s1cN )
         
    # 4.separate cows not fall within the breeding from breeding cows
    # 繁殖対象外の個体を、分離する。
    
    fpysep_out_frm_00_w(wb0, s0cN, ncolc)
    
    wb0.save(wb0N)

#fpyrpdd_manual################################################################################
"""
fpyrpdd_manual:                        start manual
ｖ1.0
2024/4/5

@author: jicc
"""
def fpyrpdd_manual():
    
    print('-----fpyrpdd_manual---------------------------------------------------------v1.02-------')
    print(' ')
    print('1. fpymkst_Cow_Heifer00( wbN, scolN, colnc, colnh, bdate)')
    print('...............................................................................')
    print(' make two sheets for cows and heifers at base date ')
    print(' 2枚のsheet yyyymmddCow00, yyyymmddHeifer00を作成 ')
    print('.....')
    print(' PS> python ps_fpymkst_cow_heifer00_args.py wbN scolN colnc colnh bdate')
    print('.....')
    print(' wbN : AB_rpdd.xlsx, scolN : columns, colnc : 1 (row number of cow\'s sheet list title), ')
    print('colnh : 3 (row number fo heifer\'s sheet list title), bdate : yyyy/mm/dd')
    print(' ')
    print('2. cowslists = fpysrt_into_Cow_Heiferlst(wbN0, sheetN0, ncol0, wbN1, sheetN1, ncol1)')
    print('   fpyinput_cwlstd_to_Cow_Heifer00( cowslists, fstcol, wbN, s0N, s1N )')
    print('...............................................................................')
    print(' input cowslists\'data to xls\'two sheets yyyymmddCow00 and yyyymmddHeifer00')
    print(' cowslistyyyymmdd のデータを2枚のsheet yyyymmddCow00, yyyymmddHeifer00に振り分ける')
    print('.....')
    print('PS> python ps_fpyinput_cwlstd_to_cow_heifer00_args.py wbN0 sheetN0 ncol0' )
    print(' wbN1 sheetN1 ncol1 fstcol wbN s0N s1N enofetn')
    print('.....')
    print(' wbN0 : AB_cowslist.xlsx, sheetN0 : cowslistyyyymmdd,')
    print(' ncol0 : 20 (the number of columns of sheet cowslist*\'s list), ')
    print(' wbN1 : AB_calving.xlsx, sheetN1 : calving, ') 
    print(' ncol1 : 11 (the number of columns of sheet calving\'s list), ')
    print(' fstcol : 1 first column number to input data ') 
    print('  wbN : AB_rpdd.xlsx, s0N : yyyymmddHeifer00, s1N : yyyymmddCow00')    
    print(' enofetn : 2(eartagNo), 3(DHITNo)')
    print(' ')
    print('---------------------------------------------------------------2024/6/5 by jicc---------')

#fpyrpdd_Cow_manual################################################################################
"""
fpyrpdd_Cow_manual:                        manual for cows
ｖ1.0
2024/4/5

@author: jicc
"""
def fpyrpdd_Cow_manual():
    
    print('-----fpyrpdd_Cow_manual---------------------------------------------------------v1.04-------')
    print(' ')
    print('1. fpyinput_clvdt_into_ymdcow(wb0N, s0N, wb1N, s1N)')
    print(' input calving data into sheet yyyymmddCow00')
    print(' 基準日における分娩日、産次数を sheet yyyymmddCow00 に入力する ')
    print('  PS> python ps_fpyinput_clvdt_into_ymdcow_args.py wb0N s0N wb1N s1N')
    print('  wb0N : AB_rpdd.xlsx, s0N : yyyymmddCow00, wb1N : AB_calving.xlsx  s1N : calving')
    print(' ')
    print('2. fpyheifers_in_cow_to_heifer(wbN, s0N, s1N)')
    print(' transfer heifers from sheet yyyymmddCow00 to yyyymmddHeifer00')
    print(' 基準日以降に初産分娩した未経産牛を、sheet yyyymmddCow00 から yyyymmddHeifer00 へ移動する ')
    print('  PS> python ps_fpyheifers_in_cow_to_heifer_args.py wbN s0N s1N')
    print('  wbN : AB_rpdd.xlsx, s0N : yyyymmddCow00, s1N : yyyymmddHeifer00')
    print(' ')
    print('3. fpydaysfrmcalving(wbN, sheetN, col_bd, col_clv, col_dsfrmclv)')
    print(' calculate days from calving at sheet yyyymmddCow00')
    print(' sheet yyyymmddCow00 の分娩後日数を計算・入力する ')
    print('  PS> python ps_fpydaysfrmcalving_args.py wbN sheetN col_bd col_clv col_dsfrmclv')
    print(' wbN : AB_rpdd.xlsx, sheetN : yyyymmddCow00, col_bd : 18, col_clv : 9,  col_dsfrmclv :10')
    print(' ')
    print('4. fpyinput_AIdt_into_ymdcow(wb0N, s0N, wb1N, s1N, VWP)')
    print(' input AIdata into sheet yyyymmddCow00')
    print(' AIdata を sheet yyyymmdd00 に入力 ')
    print(' Group, Stage, 授精回数, 初回授精日, 初回授精日数, 最終授精日, 授精後日数, 分娩予定日 を入力する')
    print('  PS> python ps_fpyinput_aidt_into_ymdcow_args.py wb0N s0N wb1N s1N VWP')
    print(' wbN : AB_rpdd.xlsx, sheetN : yyyymmddCow00,  wb1N : AB_AI.xlsx, s1N : AB_AI, VWP :50')
    print(' ')
    print('5. fpyopenDays( wbN, sheetN )')
    print(' input opendays into sheet yyyymmddCow00')
    print(' 空胎日数を計算・入力 する')
    print('  PS> python ps_fpyopendays_args.py wbN sheetN')
    print(' wbN : AB_rpdd.xlsx, sheetN : yyyymmddCow00')
    print(' ')
    print(' **繁殖対象外を考慮する場合は、toolsへ')
    print('    PS>python ps_rpdd_mh_tools_args.py')
    print(' ')
    print('6. fmstls.fpysheet_copy( wbN, sheetN, sheetN_ )')
    print('copy sheet yyyymmddCow00 and make sheet name yyyymmddCow01')
    print('yyyymmddCow00 をコピーして、yyyymmddCow01を作成 ')
    print('   PS> python ps_fpysheet_copy_args.py wbN sheetN sheetN_')
    print(' wbN : AB_rpdd.xlsx, sheetN : yyyymmddCow00, sheetN_ : yyyymmddCow01')
    print(' ')
    print('7. fpyinput_PT_(wbAIＮ, sheetAIN, wbRPDN, sRPDN)')
    print('繁殖データに、AI台帳から最終の鑑定結果を入力する')
    print('Cow版 rpdd v. sheetAIN yyyy -> AB_AI に変更')
    print('   PS> python ps_fpyinput_pt__args.py wbAIN sheetAIN wbRPDN sRPDN')
    print(' wbAIN : AB_AI.xlsx, sheetAIN : AB_AI, wbRPDN : AB_rpdd.xlsx, sRPDN : yyyymmddCow01')
    print(' ')
    print('8. fpyinput_eartagno_into_ymdcow(wb0N, s0N, cid0n, cet0n, wb1N, s1N, cid1n, cet1n)')
    print('input eartagno into sheet yyyymmddCow')
    print('耳標番号を　sheeet yyyymmddCow00 に cowslist から入力する')
    print('   PS> python ps_fpyinput_eartagno_into_ymdcow_args.py ')
    print('      wb0N s0N cid0n cet0n wb1N s1N cid1n cet1n')
    print(' wb0N : AB_rpdd.xlsx, s0N : yyyymmddCow00, cid0n : 6, cet0n : 4, ')
    print('      wb1N : AB_cowslist.xlsx  s1N :  cowslist, cid1n : 2, cet1n : 3')
    print(' ')
    print('---------------------------------------------------------------2024/5/19 by jicc---------')

#fpyrpddCowmanual00################################################################################
"""
fpyrpddCowmanual00:                   manual for cows
ｖ1.0
2024/6/6

@author: jicc
"""
def fpyrpddCowmanual00():
    
    print('-----fpyrpddCowmanual00------------------------------------------------------v1.00-------')
    print(' ')
    print('1. fpyrpdd_Cow_manual00( wb0N, s0cN, s0hN, wb1N, s1N, wb2N, s2N, VWP )')
    print(' 1)input calving data into sheet yyyymmddCow00')
    print(' 2)transfer heifers at base date ')
    print(' 3)calculate days from calving at sheet yyyymmddCow00')
    print(' 4)input AIdata into sheet yyyymmddCow00  ')
    print(' and group cows by their reproductive stage')
    print(' 5)input opendays into sheet yyyymmddCow00')
    print('---------------')
    print(' PS> python ps_fpyrpdd_cow_manual00_args.py wb0N s0cN s0hN wb1N s1N wb2N s2N VWP ')
    print(' wb0N : AB_rpdd.xlsx, s0cN : yyyymmddCow00, s0hN : yyyymmddHeifer00, ')
    print(' wb1N : AB_calving.xlsx, s1N : calving, wb2N : AB_AI.xlsx, s2Nl : AB_AI, VWP : 30, 50')
    print(' ')
    print(' **繁殖対象外を考慮する場合は、toolsへ')
    print('    PS>python ps_rpdd_mh_tools_args.py')
    print(' ')
    print('2. fpyrpdd_Cow_manual01( wb0N, s0cN, s0c_N, wb1N, s1N )')
    print(' 1)copy sheet yyyymmddHeifer00 and make sheet yyyymmddHeifer01')
    print(' 2)繁殖データに、AI台帳から最終の鑑定結果を入力する ')
    print('---------------')
    print(' PS> python ps_fpyrpdd_cow_manual01_args.py wb0N s0cN s0c_N wb1N s1N ')
    print(' wb0N : AB_rpdd.xlsx, s0cN : yyyymmddCow00, s0c_N : yyyymmddCow01,')
    print(' wb1N : AB_AI.xlsx, s1N : AB_AI')
    print(' ')
    print('---------------------------------------------------------------2024/6/6 by jicc---------')

    
#fpyrpdd_Heifer_manual################################################################################
"""
fpyrpdd_Heifer_manual:                        manual for heifers
ｖ1.0
2024/4/5

@author: jicc
"""
def fpyrpdd_Heifer_manual():
    print('-----fpyrpdd_Heifer_manual--------------------------------------------------------v1.02-------')
    print(' ')
    print('1. fpyinput_agein_dyandmnth_into_ymdheifer(wbN, sheetN)')
    print('input parity(=0) , age in day and month into sheet yyyymmddHeifer00')
    print('産次数(=0), 日齢、 月齢　を　sheet yyyymmddHeifer00 に入力する。 ')
    print(' PS> python ps_fpyinput_agein_dyandmnth_into_ymdheifer_args.py wbN sheetN')
    print(' wbN : AB_rpdd.xlsx, sheetN : yyyymmddHeifer00')
    print(' ')
    print('2. fpyinput_AIdt_into_ymdheifer(wb0N, s0N, wb1N, s1N, VWPm, VWPM)')
    print(' input AIdata into sheet yyyymmddHeifer00')
    print(' AIdata を sheet yyyymmddHeifer00 に入力 ')
    print(' Group, Stage, 授精回数, 初回授精日, 初回授精月齢, 最終授精日, 授精後日数, 分娩予定日 を入力する')
    print('  PS> python ps_fpyinput_aidt_into_ymdheifer_args.py wb0N s0N wb1N s1N VWP')
    print(' wbN : AB_rpdd.xlsx, sheetN : yyyymmddHeifer00,  wb1N : AB_AI_.xlsx, s1N : AB_AI, VWPm :13, VWPM : 14')
    print(' ')
    print('3. mh_rpdu.fpyopenMnths( wbN, sheetN )')
    print(' input openmonths into sheet yyyymmddHeifer00')
    print(' 空胎月数を計算・入力 する')
    print('  PS> python ps_fpyopenmnths_args.py wbN sheetN')
    print(' wbN : AB_rpdd.xlsx, sheetN : yyyymmddHeifer00')
    print(' ')
    print('4. fmstls.fpysheet_copy( wbN, sheetN, sheetN_ )')
    print('copy sheet yyyymmddHeifer00 and make sheet name yyyymmddHeifer01')
    print('yyyymmddHeifer00 をコピーして、yyyymmddHeifer01を作成 ')
    print('   PS> python ps_fpysheet_copy_args.py wbN sheetN sheetN_')
    print(' wbN : AB_rpdd.xlsx, sheetN : yyyymmddHeifer00, sheetN_ : yyyymmddHeifer01')
    print(' ')
    print('5. fpyinput_PTH_(wbAIＮ, sheetAIN, wbRPDN, sRPDN)')
    print('繁殖データに、AI台帳から最終の鑑定結果を入力する')
    print('Heifer版 rpdd v. sheetAIN yyyy -> AB_AI に変更')
    print('   PS> python ps_fpyinput_pth__args.py wbAIN sheetAIN wbRPDN sRPDN')
    print(' wbAIN : AB_AI.xlsx, sheetAIN : AB_AI, wbRPDN : AB_rpdd.xlsx, sRPDN : yyyymmddHeifer01')
    print(' ')
    print('---------------------------------------------------------------2024/4/14 by jicc---------')

#fpyrpddHeifermanual00################################################################################
"""
fpyrpddHeifermanual00:                   manual for heifers
ｖ1.0
2024/6/6

@author: jicc
"""
def fpyrpddHeifermanual00():
    
    print('-----fpyrpddHeifermanual00-------------------------------------------------------v1.00-------')
    print(' ')
    print('1. fpyrpdd_Heifer_manual00( wb0N, s0hN, wb1N, s1N, VWPm, VWPM )')
    print(' 1)input parity(=0) , age in day and manth into sheet yyyymmddHeifer00')
    print('   産次数(=0), 日齢、 月齢　を　sheet yyyymmddHeifer00 に入力する。')
    print(' 2)input AIdata into sheet yyyymmddHeifer00')
    print('  and group heifers by their reproductive stage')
    print('  AIdata を sheet yyyymmddHeifer00 に入力')
    print('  Group, Stage, 授精回数, 初回授精日, 初回授精月齢, 最終授精日, ')
    print('  授精後日数, 分娩予定日 を入力する')
    print(' 3)input openmonths into sheet yyyymmddHeifer00')
    print('   空胎月数を計算・入力 する  ')
    print('------------------')
    print('  PS> python ps_fpyrpdd_heifer_manual00_args.py wb0N s0hN wb1N s1N VWPm VWPM')
    print('      wb0N : AB_rpdd.xlsx, s0hN : yyyymmddHeifer00,')
    print('      wb1N : AB_AI.xlsx, s1N : AB_AI, ')
    print('      VWPm : 13, VWPM : 14')
    print(' ')
    print('2. fpyrpdd_Heifer_manual01( wb0N, s0hN, s0h_N, wb1N, s1N )')
    print(' 1)copy sheet yyyymmddHeifer00 and make sheet yyyymmddHeifer01')
    print(' 2)繁殖データに、AI台帳から最終の鑑定結果を入力する ')
    print('------------------')
    print('  PS> python ps_fpyrpdd_heifer_manual01_args.py wb0N s0hN s0h_N wb1N s1N ')
    print('      wb0N : AB_rpdd.xlsx, s0hN : yyyymmddHeifer00, ')
    print('      s0h_N : yyyymmddHeifer01,')
    print('      wb1N : AB_AI.xlsx, s1N : AB_AI')
    print(' ')
    print('---------------------------------------------------------------2024/6/6 by jicc---------')


#fpyrpddReference###################################################################
"""
fpyrpddReference:         reference of rpdd's functions
ｖ1.0
2024/3/2
@author: jicc
"""
def fpyrpddReference():
    
    
    print('-----rpddReference -----------------------------------------------------v1.19-----')
    print('** fpymkst_Cow_Heifer00( wbN, scolN, colnc, colnh, bdate)')
    print('make two sheets for cows and heifers at base date ')
    print('...................................................................................')
    print('** fpychgelmv_of_lsts_lst( lst, index, chgv )')
    print('change a element value of a lists\'list')
    print('基準日(yyyymmdd)のcowslist の検索年月日を基準日yyyy/mm/dd に変更する')
    print('...................................................................................')
    print('** fpysrt_into_Cow_Heiferlst(wbN0, sheetN0, ncol0, wbN1, sheetN1, ncol1)')
    print('sort cowslist into two lists Heifer and Cow')
    print('cowslistの個体を、Heifer と　Cow のリストに分ける')
    print('...................................................................................')
    print('** fpyinput_cwlstd_to_Cow_Heifer00( cowslists, fstcol, wbN, s0N, s1N enofetn )')
    print('input cowslists\'data to xls\'two sheets yyyymmddHeifer00 and yyyymmddCow00')
    print('fpysrt_into_Cow_Heiferlst の　cowslists が必要 ')
    print('...................................................................................')
    print('** fpyext_idNo_s_calvinglist( idNo, wbN, sN, idNo_coln, clvd_coln, prty_coln )')
    print('extract an individual calving list of two elements,calving_date and parity')
    print('**個別の分娩リスト[[calving_date, parity]...]を得る')
    print('...................................................................................')
    print('** fpyext_idNo_s_calvinglist_s( idNo, sheet, idNo_coln, clvd_coln, prty_coln )')
    print('extract an individual calving list of two elements,calving_date and parity')
    print('**個別の分娩リスト[[calving_date, parity]...]を得る')
    print('sheet version')
    print('...................................................................................')
    print('** fpysel_clvdate( bdate, ind_calving_list )')
    print('select the latest calving date at a base date')
    print('**基準日における最新の分娩日を選択する')
    print('...................................................................................')
    print('** fpyinput_clvdt_into_ymdcow(wb0N, s0N, wb1N, s1N)')
    print('input calving data into sheet yyyymmddCow00')
    print('**基準日における分娩日、産次数を sheet yyyymmddCow00 に入力する')
    print('...................................................................................')
    print('** fpyinput_clvdt_into_ymdcow_s(s0, s1)')
    print('input calving data into sheet yyyymmddCow00')
    print('**基準日における分娩日、産次数を sheet yyyymmddCow00 に入力する')
    print('sheet version')
    print('...................................................................................')
    print('** fpyheifers_in_cow_to_heifer(wbN, s0N, s1N)')
    print('transfer heifers from sheet yyyymmddCow00 to yyyymmddHeifer00')
    print('**基準日以降に初産分娩した未経産牛を、sheet yyyymmddCow00 から yyyymmddHeifer00 へ移動する')
    print('...................................................................................')
    print('** fpyheifers_in_cow_to_heifer_ws(wb, s0cN, s0c, s0hN, s0h)')
    print('transfer heifers from sheet yyyymmddCow00 to yyyymmddHeifer00')
    print('**基準日以降に初産分娩した未経産牛を、sheet yyyymmddCow00 から yyyymmddHeifer00 へ移動する')
    print('workbook sheet version')
    print('...................................................................................')
    print('** fpydaysfrmcalving(wbN, sheetN, col_bd, col_clv, col_dsfrmclv)')
    print('calculate and input days from calving at sheet yyyymmddCow00')
    print('**分娩後日数=基準日-分娩日を計算し、sheet yyyymmddCow00に入力する')
    print('...................................................................................')
    print('** fpyext_idNo_clvd_s_AIlist(idNo, clvd, wbN, sN, idNo_coln, clvd_coln,')
    print(                              'lstAI_coln, AIt_coln, PT_coln, eDofnc_coln)')
    print('extract an individual AI list after the latest calving at base date')
    print('**個体の、基準日直近の分娩以後のAIlistを抽出する')
    print('...................................................................................')
    print('** fpyext_idNo_clvd_s_AIlist_s(idNo, clvd, sheet, idNo_coln, clvd_coln,')
    print(                              'lstAI_coln, AIt_coln, PT_coln, eDofnc_coln)')
    print('extract an individual AI list after the latest calving at base date')
    print('sheet version')
    print('**個体の、基準日直近の分娩以後のAIlistを抽出する')
    print('...................................................................................')
    print('** fpyinput_AIdt_into_ymdcow(wb0N, s0N, wb1N, s1N, VWP)')
    print('input AIdata into sheet yyyymmddCow00')
    print('**AIdata を sheet yyyymmddCow00 に入力')
    print('...................................................................................')
    print('** fpyinput_AIdt_into_ymdcow_s(s0, s1, VWP)')
    print('input AIdata into sheet yyyymmddCow00')
    print('**AIdata を sheet yyyymmddCow00 に入力')
    print('sheet version')
    print('...................................................................................')
    print('** mh_rpdu.fpyopenDays( wbN, sheetN )')
    print('calculate opendays 空胎日数を計算する')
    print('**空胎日数を計算する, 分娩日　あるなし、　Groupによる場合分け')
    print('...................................................................................')
    print('** mh_rpdu.fpyopenDays_s( sheet )')
    print('calculate opendays 空胎日数を計算する')
    print('**空胎日数を計算する, 分娩日　あるなし、　Groupによる場合分け')
    print('sheet version')
    print('...................................................................................')
    print('** fpysep_a_stagefrmothr(xllists, index, vl)')
    print('separate a stage from other stages')
    print('**あるステージを他のステージと分ける')
    print('...................................................................................')
    print('**fpydelete_rows_of_xls(wbN, sheetN)')
    print('clear a list data of xls\'s sheet')
    print('**Excel sheet の　リストデータを消去する 行を削除する')
    print('...................................................................................')
    print('**fpydelete_rows_of_xls_s(sheet, r, amt)')
    print('clear a list data of xls\'s sheet')
    print('sheet version')
    print('**Excel sheet の　リストデータを消去する 行を削除する')
    print('...................................................................................')
    print('**fpysep_out_frm_00(wbN, sheetN, ncol)')
    print('separate cows not fall within the breeding from breeding cows')
    print('繁殖対象外の個体を、分離する。')
    print('...................................................................................')
    print('**fpyinput_agein_dyandmnth_into_ymdheifer(wbN, sheetN)')
    print('input parity(=0) , age in day and manth into sheet yyyymmddHeifer00')
    print('産次数(=0), 日齢、 月齢　を　sheet yyyymmddHeifer00 に入力する。')
    print('...................................................................................')
    print('**fpyinput_agein_dyandmnth_into_ymdheife_s(sheet)')
    print('input parity(=0) , age in day and manth into sheet yyyymmddHeifer00')
    print('産次数(=0), 日齢、 月齢　を　sheet yyyymmddHeifer00 に入力する。')
    print('sheet version')
    print('...................................................................................')
    print('**fpyext_idNo_s_AIlist_heifer(idNo, bdate, wbN, sN, idNo_coln, lstAI_coln,')
    print('                           AIt_coln, PT_coln, eDofnc_coln)')
    print('extract an individual AI list for heifers at base date')
    print('未経産牛に関して、個体の、基準日におけるAIlistを抽出する')
    print('...................................................................................')
    print('**fpyext_idNo_s_AIlist_heifer_s(idNo, bdate, sheet, idNo_coln, lstAI_coln,')
    print('                           AIt_coln, PT_coln, eDofnc_coln)')
    print('extract an individual AI list for heifers at base date')
    print('sheet version')
    print('未経産牛に関して、個体の、基準日におけるAIlistを抽出する')
    print('...................................................................................')
    print('**fpyinput_AIdt_into_ymdheifer(wb0N, s0N, wb1N, s1N, VWPm, VWPM)')
    print(' input AIdata into sheet yyyymmddHeifer00')
    print('AIdata を sheet yyyymmddHeifer00 に入力')
    print('...................................................................................')
    print('**fpyinput_AIdt_into_ymdheifer_s(s0, s1, VWPm, VWPM)')
    print(' input AIdata into sheet yyyymmddHeifer00')
    print('AIdata を sheet yyyymmddHeifer00 に入力')
    print('sheet version')
    print('...................................................................................')
    print(' calculate and input openmonths into sheet yyyymmddHeifer00')
    print('空胎月数を計算・入力する。')
    print('...................................................................................')
    print('**fpyinput_PT_(wbAIＮ, sheetAIN, wbRPDN, sRPDN)')
    print(' 繁殖データに、AI台帳から最終の鑑定結果を入力する')
    print('Cow版 rpdd version : sheetAIN yyyy -> AB_AI に変更')
    print('...................................................................................')
    print('**fpyinput_PT__s(sheetAI, sheetRPDN)')
    print(' 繁殖データに、AI台帳から最終の鑑定結果を入力する')
    print('Cow版 rpdd version : sheetAIN yyyy -> AB_AI に変更')
    print('sheet version')
    print('...................................................................................')
    print('**fpyinput_PTH_(wbAIＮ, sheetAIN, wbRPDN, sRPDN)')
    print(' 繁殖データに、AI台帳から最終の鑑定結果を入力する')
    print('Heifer版 rpdd version : sheetAIN yyyy -> AB_AI に変更')
    print('...................................................................................')
    print('**fpyinput_PTH__s(sheetAI, sheetRPD)')
    print(' 繁殖データに、AI台帳から最終の鑑定結果を入力する')
    print('Heifer版 rpdd version : sheetAIN yyyy -> AB_AI に変更')
    print('sheet version')
    print('----------------------------------------------------------2024/6/6　by jicc---------')
    
#fpyrpdd_MH_tools################################################################################
"""
fpyrpdd_MH_tools:                        tools
ｖ1.0
2024/3/21

@author: jicc
"""
def fpyrpdd_MH_tools():
    
    print('-----fpyrpdd_MH_tools---------------------------------------------------------v1.02-------')
    print('sheet yyyymmddCow00 繁殖対象外の個体を除去する' )
    print('1. bdateに近い \'経産牛一覧plusyyyymmdd.csv\' をMH_rpdd.xlsxに移動し、')
    print('sheeet ‘yyyymmddCow_uorg’ とする。 hand work ')
    print('#fpyidNo_9to10( wbN, sheetN, col )')
    print('5列 個体識別番号を10桁文字列に変更')
    print('   PS> python ps_fpyidno_9to10_args.py wbN sheetN col')
    print(' wbN : AB_rpdd.xlsx, sheetN : yyyymmddCow_uorg, col : 5')
    print(' ')
    print('2. fpychg_grNo_6_7_ifn(wbN, sheetN, srefN)')
    print('change Group and Stage to \'6 乾乳\' and \'7 繁殖対象外\', if necessary')
    print('Group と Stage に \'6 乾乳\' and \'7 繁殖対象外\'　の分類を加える。')
    print(' umotion の　データ　sheet yyyymmddCow_uorg を参照する。')
    print('   PS> python ps_fpychg_grno_6_7_ifn_args.py wbN sheetN srefN')
    print(' wbN : AB_rpdd.xlsx, sheetN : yyyymmddCow00, srefN : yyyymmddCow_uorg')
    print(' ')
    print('3. fmstls.fpysheet_copy( wbN, sheetN, sheetN_ )')
    print('copy sheet yyyymmddCow00 and make sheet name yyyymmddCow00all')
    print('yyyymmddCow00 をコピーして、yyyymmddCow00allを作成 ')
    print('   PS> python ps_fpysheet_copy_args.py wbN sheetN sheetN_')
    print(' wbN : AB_rpdd.xlsx, sheetN : yyyymmddCow00, sheetN_ : yyyymmddCow00all')
    print(' ')
    print('4. fpysep_out_frm_00(wbN, sheetN, ncol)')
    print('separate cows not fall within the breeding from breeding cows')
    print('繁殖対象外の個体を、分離する。 ')
    print('   PS> python ps_fpysep_out_frm_00_args.py wbN sheetN ncol')
    print(' wbN : AB_rpdd.xlsx, sheetN : yyyymmddCow00, ncol : 18')
    print(' ')
    print('---------------------------------------------------------------2024/3/29 by jicc---------')
    
    
#fpyrpdd_MH_tools00################################################################################
"""
fpyrpdd_MH_tools00:                        tools
ｖ1.0
2024/6/26

@author: jicc
"""
def fpyrpddMHtools00():
    
    print('-----fpyrpdd_MH_tools00------------------------------------------------v1.00-------')
    print(' ')
    print('fpyrpdd_MH_tools00( wb0N, s0cN, ncolc, s0uN, coluidNo, s1cN)')
    print(' ')
    print('sheet yyyymmddCow00 繁殖対象外の個体を除去する' )
    print('0. bdateに近い \'経産牛一覧plusyyyymmdd.csv\' をMH_rpdd.xlsxに移動し、')
    print('sheeet ‘yyyymmddCow_uorg’ とする。 hand work ')
    print(' ')
    print('1. 5列 個体識別番号を10桁文字列に変更')
    print(' ')
    print('2.change Group and Stage to \'6 乾乳\' and \'7 繁殖対象外\, if necessary ')
    print('Group と Stage に \'6 乾乳\ and \'7 繁殖対象外\'　の分類を加える。')
    print(' umotion の　データ　sheet yyyymmddCow_uorg を参照する。')
    print(' ')
    print('3. copy sheet yyyymmddCow00 and make sheet name yyyymmddCow00all')
    print('yyyymmddCow00 をコピーして、yyyymmddCow00allを作成 ')
    print('4. separate cows not fall within the breeding from breeding cows)')
    print('繁殖対象外の個体を、分離する。 ')
    print('------------------------------------------------------')
    print('PS> python ps_fpyrpdd_mh_tools00_args.py wb0N s0cN ncolc s0uN coluidNo s1cN')
    print('wb0N : AB_rpdd.xlsx, s0cN : yyyymmddCow00, ncolc : 18')
    print(' s0uN : yyyymmddCow_uorg, coluidNo : 5, s1cN: yyyymmddCow00all')
    print(' ')
    print('---------------------------------------------------------------2024/6/26 by jicc---------')